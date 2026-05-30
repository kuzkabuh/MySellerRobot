"""version: 1.1.0
description: Excel/CSV import service for WB auto promotion conditions.
    Imports entry price conditions from user-uploaded Excel or CSV files.
updated: 2026-05-23
"""

import asyncio
import csv
import io
import logging
import tempfile
from collections.abc import Iterator, Sequence
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

import openpyxl
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.domain import (
    Product,
    WbAutoPromotionCondition,
)

logger = logging.getLogger(__name__)

COLUMN_MAP_RU = {
    "nmid": "wb_nm_id",
    "артикул продавца": "seller_article",
    "название товара": "title",
    "название автоакции": "promotion_name",
    "цена для участия": "required_price",
    "плановая цена для акции": "required_price",
    "текущая цена wb": "current_wb_price",
    "текущая розничная цена": "current_full_price",
    "текущая скидка на сайте, %": "current_discount",
    "загружаемая скидка для участия в акции": "wb_upload_discount_percent",
    "товар уже участвует в акции": "is_participating",
    "артикул wb": "wb_nm_id",
    "артикул поставщика": "seller_article",
    "наименование": "title",
    "статус": "wb_status",
    "участвует": "is_participating",
    "комментарий": "comment",
}

COLUMN_MAP_TECH = {
    "wb_nm_id": "wb_nm_id",
    "seller_article": "seller_article",
    "title": "title",
    "promotion_name": "promotion_name",
    "required_price": "required_price",
    "current_wb_price": "current_wb_price",
    "current_full_price": "current_full_price",
    "current_discount": "current_discount",
    "current_discounted_price": "current_discounted_price",
    "wb_upload_discount_percent": "wb_upload_discount_percent",
    "wb_status": "wb_status",
    "is_participating": "is_participating",
    "comment": "comment",
}

REQUIRED_COLUMNS = {"wb_nm_id"}


@dataclass(slots=True)
class AutoPromoImportPreview:
    import_id: int
    user_id: int
    marketplace_account_id: int
    total_rows: int
    valid_rows: int
    warning_rows: int
    error_rows: int
    source: str = "web"
    original_file_name: str | None = None
    created_at: datetime | None = None


@dataclass(slots=True)
class AutoPromoImportRow:
    id: int
    import_id: int
    wb_nm_id: int | None
    seller_article: str | None
    title: str | None
    promotion_name: str | None
    required_price: Decimal | None
    current_wb_price: Decimal | None
    current_full_price: Decimal | None
    current_discount: Decimal | None
    current_discounted_price: Decimal | None
    wb_upload_discount_percent: Decimal | None
    condition_type: str
    is_participating: bool | None
    product_id: int | None
    status: str
    message: str | None


class WbAutoPromoImportService:
    """Import auto promotion conditions from Excel files."""

    def __init__(self, session: AsyncSession) -> None:
        self.session = session

    async def generate_template(self, user_id: int) -> Path:
        """Generate an Excel template for auto promotion conditions import."""
        wb = Workbook()
        ws = wb.active
        if ws is None:
            raise RuntimeError("Excel workbook has no active worksheet")
        ws.title = "Условия автоакций"

        headers = [
            "nmID",
            "Артикул продавца",
            "Название товара",
            "Название автоакции",
            "Цена для участия",
            "Текущая цена WB",
            "Участвует",
            "Комментарий",
        ]
        ws.append(headers)

        ws.append(
            [
                345455998,
                "ARTICLE-001",
                "Пример товара",
                "Автоакция WB",
                980,
                1000,
                "нет",
                "Пример строки",
            ]
        )

        tmp = Path(tempfile.gettempdir()) / f"auto_promo_template_{user_id}.xlsx"
        wb.save(str(tmp))
        return tmp

    async def create_preview(
        self,
        file_path: Path,
        user_id: int,
        marketplace_account_id: int,
        source: str = "web",
        original_file_name: str | None = None,
    ) -> tuple[AutoPromoImportPreview, list[dict[str, Any]]]:
        """Parse Excel or CSV file and create a preview of import results."""
        logger.info(
            "wb_auto_promo_conditions_import_started",
            extra={
                "user_id": user_id,
                "marketplace_account_id": marketplace_account_id,
                "file": original_file_name or file_path.name,
            },
        )

        suffix = file_path.suffix.lower()

        try:
            if suffix in (".xlsx", ".xlsm"):
                return await self._parse_xlsx(
                    file_path,
                    user_id,
                    marketplace_account_id,
                    source,
                    original_file_name,
                )
            elif suffix == ".csv":
                return await self._parse_csv(
                    file_path,
                    user_id,
                    marketplace_account_id,
                    source,
                    original_file_name,
                )
            else:
                raise ValueError("Файл должен быть .xlsx или .csv")
        except BadZipFile:
            logger.warning(
                "wb_auto_promo_import_bad_zip_file",
                extra={"file": original_file_name or file_path.name},
            )
            raise ValueError("Excel-файл повреждён или сохранён не в формате XLSX") from None
        except ValueError:
            raise
        except Exception as exc:
            logger.exception("wb_auto_promo_import_unexpected_error")
            raise ValueError(f"Ошибка при чтении файла: {exc}") from exc

    async def _parse_xlsx(
        self,
        file_path: Path,
        user_id: int,
        marketplace_account_id: int,
        source: str,
        original_file_name: str | None,
    ) -> tuple[AutoPromoImportPreview, list[dict[str, Any]]]:
        """Parse XLSX/XLSM file."""
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb["Отчёт по скидкам"] if "Отчёт по скидкам" in wb.sheetnames else wb.active
            if ws is None:
                raise ValueError("Excel-файл не содержит листов")
            return await self._process_rows(
                ws.iter_rows(values_only=True),
                user_id,
                marketplace_account_id,
                source,
                original_file_name,
            )
        finally:
            wb.close()

    async def _parse_csv(
        self,
        file_path: Path,
        user_id: int,
        marketplace_account_id: int,
        source: str,
        original_file_name: str | None,
    ) -> tuple[AutoPromoImportPreview, list[dict[str, Any]]]:
        """Parse CSV file."""
        content = await asyncio.to_thread(file_path.read_text, encoding="utf-8-sig")

        reader = csv.reader(io.StringIO(content))
        rows_iter = iter(reader)

        return await self._process_rows(
            rows_iter,
            user_id,
            marketplace_account_id,
            source,
            original_file_name,
        )

    async def _process_rows(
        self,
        rows_iter: Iterator[Sequence[Any]],
        user_id: int,
        marketplace_account_id: int,
        source: str,
        original_file_name: str | None,
    ) -> tuple[AutoPromoImportPreview, list[dict[str, Any]]]:
        """Process rows from any source (XLSX or CSV)."""
        raw_headers = next(rows_iter, None)
        if raw_headers is None:
            raise ValueError("Файл пустой")

        headers = self._resolve_headers(raw_headers)
        missing = REQUIRED_COLUMNS - set(headers.values())
        if missing:
            raise ValueError(f"Нет обязательных колонок: {', '.join(missing)}")
        has_price_condition = bool(
            {"required_price", "wb_upload_discount_percent"} & set(headers.values())
        )
        if not has_price_condition:
            raise ValueError(
                "Нет обязательных колонок: required_price или wb_upload_discount_percent"
            )

        preview_rows: list[dict[str, Any]] = []
        valid_count = 0
        warning_count = 0
        error_count = 0

        products_cache: dict[int, Product] = {}

        for row_idx, row_values in enumerate(rows_iter, start=2):
            row_data: dict[str, Any] = {}
            for col_idx, header_key in headers.items():
                if col_idx < len(row_values):
                    row_data[header_key] = row_values[col_idx]

            wb_nm_id = self._parse_int(row_data.get("wb_nm_id"))
            required_price = self._parse_decimal(row_data.get("required_price"))
            current_wb_price = self._parse_decimal(row_data.get("current_wb_price"))
            current_full_price = self._parse_decimal(row_data.get("current_full_price"))
            current_discount = self._parse_decimal(row_data.get("current_discount"))
            wb_upload_discount_percent = self._parse_decimal(
                row_data.get("wb_upload_discount_percent")
            )
            current_discounted_price = self._parse_decimal(row_data.get("current_discounted_price"))
            if current_discounted_price is None:
                current_discounted_price = self._discounted_price(
                    current_full_price,
                    current_discount,
                )
            if current_wb_price is None:
                current_wb_price = current_discounted_price
            is_participating = self._parse_bool(row_data.get("is_participating"))
            seller_article = str(row_data.get("seller_article") or "").strip()
            title = str(row_data.get("title") or "").strip()
            promotion_name = str(row_data.get("promotion_name") or "").strip()
            wb_status = str(row_data.get("wb_status") or "").strip()
            condition_type = "max_price" if required_price is not None else "discount_projection"
            fallback_price = self._discounted_price(
                current_full_price,
                wb_upload_discount_percent,
            )

            status = "valid"
            message: str | None = None
            product = None

            if wb_nm_id is None:
                status = "error"
                message = "nmID не указан или некорректен"
                error_count += 1
            elif required_price is not None and required_price <= 0:
                status = "error"
                message = "Плановая цена для акции <= 0"
                error_count += 1
            elif required_price is None and (fallback_price is None or fallback_price <= 0):
                status = "error"
                message = (
                    "Плановая цена не указана, а загружаемую скидку нельзя " "пересчитать в цену"
                )
                error_count += 1
            else:
                product = await self._find_product_by_nm_id(
                    marketplace_account_id,
                    wb_nm_id,
                    products_cache,
                )
                if product is None:
                    status = "warning"
                    message = "Товар не найден в базе, условие будет сохранено"
                    warning_count += 1
                else:
                    valid_count += 1

            preview_rows.append(
                {
                    "row_num": row_idx,
                    "wb_nm_id": wb_nm_id,
                    "seller_article": seller_article or None,
                    "title": title or None,
                    "promotion_name": promotion_name or None,
                    "required_price": required_price,
                    "current_wb_price": current_wb_price,
                    "current_full_price": current_full_price,
                    "current_discount": current_discount,
                    "current_discounted_price": current_discounted_price,
                    "wb_upload_discount_percent": wb_upload_discount_percent,
                    "fallback_discounted_price": fallback_price,
                    "condition_type": condition_type,
                    "wb_status": wb_status or None,
                    "is_participating": is_participating,
                    "product_id": product.id if product else None,
                    "status": status,
                    "message": message,
                    "raw_payload": {
                        "row_num": row_idx,
                        "wb_status": wb_status or None,
                        "plan_price": str(required_price) if required_price is not None else None,
                        "current_full_price": (
                            str(current_full_price) if current_full_price is not None else None
                        ),
                        "current_discount": (
                            str(current_discount) if current_discount is not None else None
                        ),
                        "current_discounted_price": (
                            str(current_discounted_price)
                            if current_discounted_price is not None
                            else None
                        ),
                        "wb_upload_discount_percent": (
                            str(wb_upload_discount_percent)
                            if wb_upload_discount_percent is not None
                            else None
                        ),
                        "fallback_discounted_price": (
                            str(fallback_price) if fallback_price is not None else None
                        ),
                        "wb_upload_discount_is_diagnostic": required_price is not None,
                    },
                }
            )

        total_rows = len(preview_rows)
        preview = AutoPromoImportPreview(
            import_id=0,
            user_id=user_id,
            marketplace_account_id=marketplace_account_id,
            total_rows=total_rows,
            valid_rows=valid_count,
            warning_rows=warning_count,
            error_rows=error_count,
            source=source,
            original_file_name=original_file_name,
            created_at=datetime.now(tz=UTC),
        )

        logger.info(
            "wb_auto_promo_conditions_import_preview_created",
            extra={
                "user_id": user_id,
                "total_rows": total_rows,
                "valid_rows": valid_count,
                "warning_rows": warning_count,
                "error_rows": error_count,
            },
        )

        return preview, preview_rows

    async def apply_import(
        self,
        preview_rows: list[dict[str, Any]],
        user_id: int,
        marketplace_account_id: int,
    ) -> int:
        """Apply the import: save conditions to the database."""
        now_utc = datetime.now(tz=UTC)
        saved_count = 0

        for row_data in preview_rows:
            if row_data["status"] == "error":
                continue

            wb_nm_id = row_data.get("wb_nm_id")
            required_price = row_data.get("required_price")
            current_full_price = row_data.get("current_full_price")
            upload_discount = row_data.get("wb_upload_discount_percent")
            fallback_price = row_data.get("fallback_discounted_price")
            condition_type = row_data.get("condition_type") or "unknown"
            if wb_nm_id is None:
                continue
            if required_price is None and fallback_price is None:
                continue

            existing = await self.session.execute(
                select(WbAutoPromotionCondition).where(
                    WbAutoPromotionCondition.marketplace_account_id == marketplace_account_id,
                    WbAutoPromotionCondition.wb_nm_id == wb_nm_id,
                    WbAutoPromotionCondition.source == "file_import",
                    WbAutoPromotionCondition.promotion_name
                    == (row_data.get("promotion_name") or ""),
                )
            )
            condition = existing.scalar_one_or_none()

            if condition is None:
                condition = WbAutoPromotionCondition(
                    user_id=user_id,
                    marketplace_account_id=marketplace_account_id,
                    wb_nm_id=wb_nm_id,
                    source="file_import",
                )
                self.session.add(condition)

            condition.seller_article = row_data.get("seller_article")
            condition.title = row_data.get("title")
            condition.promotion_name = row_data.get("promotion_name")
            condition.required_price = required_price
            condition.max_auto_promo_price = required_price
            condition.current_wb_price = row_data.get("current_wb_price")
            condition.wb_condition_discount_percent = upload_discount
            condition.current_full_price = current_full_price
            condition.current_discount = self._decimal_to_int(row_data.get("current_discount"))
            condition.current_discounted_price = row_data.get("current_discounted_price")
            condition.candidate_discounted_price = required_price or fallback_price
            condition.condition_type = condition_type
            condition.is_participating = row_data.get("is_participating")
            condition.raw_payload = row_data.get("raw_payload")
            condition.synced_at = now_utc
            saved_count += 1

        await self.session.flush()

        logger.info(
            "wb_auto_promo_conditions_import_applied",
            extra={
                "user_id": user_id,
                "marketplace_account_id": marketplace_account_id,
                "saved_count": saved_count,
            },
        )

        return saved_count

    async def _find_product_by_nm_id(
        self,
        marketplace_account_id: int,
        wb_nm_id: int,
        cache: dict[int, Product],
    ) -> Product | None:
        """Find a product by nmID, using cache."""
        if wb_nm_id in cache:
            return cache[wb_nm_id]

        result = await self.session.execute(
            select(Product)
            .where(
                Product.marketplace_account_id == marketplace_account_id,
                Product.marketplace == "wb",
                (Product.external_product_id == str(wb_nm_id))
                | (Product.marketplace_article == str(wb_nm_id)),
            )
            .limit(1)
        )
        product = result.scalar_one_or_none()
        if product:
            cache[wb_nm_id] = product
        return product

    @staticmethod
    def _resolve_headers(raw_headers: Sequence[Any]) -> dict[int, str]:
        """Map raw header strings to technical column names."""
        result: dict[int, str] = {}
        for idx, h in enumerate(raw_headers):
            h_str = str(h).strip().lower()
            if h_str in COLUMN_MAP_RU:
                result[idx] = COLUMN_MAP_RU[h_str]
            elif h_str in COLUMN_MAP_TECH:
                result[idx] = COLUMN_MAP_TECH[h_str]
            else:
                result[idx] = h_str
        return result

    @staticmethod
    def _parse_int(value: Any) -> int | None:
        if value is None:
            return None
        try:
            return int(str(value).replace("\xa0", "").replace(" ", "").strip())
        except (ValueError, TypeError):
            try:
                return int(Decimal(str(value)))
            except (InvalidOperation, ValueError, TypeError):
                return None

    @staticmethod
    def _parse_decimal(value: Any) -> Decimal | None:
        if value is None or value == "":
            return None
        try:
            normalized = (
                str(value)
                .replace("\xa0", "")
                .replace(" ", "")
                .replace("₽", "")
                .replace("%", "")
                .replace(",", ".")
                .strip()
            )
            return Decimal(normalized)
        except (InvalidOperation, ValueError, TypeError):
            return None

    @staticmethod
    def _parse_bool(value: Any) -> bool | None:
        if value is None:
            return None
        if isinstance(value, bool):
            return value
        s = str(value).strip().lower()
        if s in ("да", "yes", "true", "1", "участвует"):
            return True
        if s in ("нет", "no", "false", "0", "не участвует"):
            return False
        return None

    @staticmethod
    def _discounted_price(
        full_price: Decimal | None,
        discount_percent: Decimal | None,
    ) -> Decimal | None:
        if full_price is None or discount_percent is None:
            return None
        if full_price <= 0 or discount_percent < 0 or discount_percent >= 100:
            return None
        return full_price * (Decimal("1") - discount_percent / Decimal("100"))

    @staticmethod
    def _decimal_to_int(value: Any) -> int | None:
        if value is None:
            return None
        try:
            return int(Decimal(str(value)))
        except (InvalidOperation, ValueError, TypeError):
            return None
