"""version: 1.0.0
description: MRC bulk import/export service for Excel files.
    Shared between Telegram bot and WEB.
updated: 2026-05-21
"""

import logging
import math
import uuid
from dataclasses import dataclass, field
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import get_settings
from app.models.domain import MarketplaceAccount, Product, WbPromotion, WbPromotionNomenclature
from app.models.enums import Marketplace

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

TEMPLATE_COLUMNS = [
    "product_id",
    "wb_nm_id",
    "seller_sku",
    "barcode",
    "brand",
    "product_name",
    "current_wb_price",
    "current_mrc_price",
    "new_mrc_price",
    "min_price",
    "promo_name",
    "promo_plan_price",
    "calculated_price_preview",
    "comment",
]

REQUIRED_COLUMNS = {"product_id", "wb_nm_id", "new_mrc_price"}

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SERVICE_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")


@dataclass(slots=True)
class MrcImportRow:
    """Single row from import file."""

    row_number: int
    product_id: int | None
    wb_nm_id: int | None
    seller_sku: str | None
    product_name: str | None
    new_mrc_price_raw: str | None
    old_mrc_price: Decimal | None
    status: str = "pending"
    message: str = ""


@dataclass(slots=True)
class MrcImportPreview:
    """Preview result after parsing import file."""

    preview_id: str
    user_id: int
    file_name: str
    total_rows: int
    valid_rows: list[MrcImportRow] = field(default_factory=list)
    skipped_rows: list[MrcImportRow] = field(default_factory=list)
    warning_rows: list[MrcImportRow] = field(default_factory=list)
    error_rows: list[MrcImportRow] = field(default_factory=list)
    created_at: datetime = field(default_factory=lambda: datetime.now(tz=UTC))


@dataclass(slots=True)
class MrcImportResult:
    """Final result after applying import."""

    preview_id: str
    user_id: int
    updated_count: int
    cleared_count: int
    skipped_count: int
    error_count: int
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


class MrcImportService:
    """Service for bulk MRC import/export via Excel files."""

    def __init__(self, session: AsyncSession) -> None:
        self.session = session
        self.settings = get_settings()
        self._previews: dict[str, MrcImportPreview] = {}

    def _max_file_size(self) -> int:
        return self.settings.mrc_import_max_file_size_mb * 1024 * 1024

    def _max_rows(self) -> int:
        return self.settings.mrc_import_max_rows

    def _allow_clear(self) -> bool:
        return self.settings.mrc_import_allow_clear

    async def generate_mrc_template(
        self,
        user_id: int,
        account_id: int | None = None,
    ) -> Path:
        """Generate Excel template with user's WB products."""
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl not installed")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "МРЦ WB"

        # Headers
        service_cols = {0, 1, 2, 3}  # product_id, wb_nm_id, seller_sku, barcode
        input_col = 8  # new_mrc_price

        for col_idx, header in enumerate(TEMPLATE_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

            if col_idx - 1 in service_cols:
                cell.fill = SERVICE_FILL
            elif col_idx - 1 == input_col:
                cell.fill = INPUT_FILL

        # Fetch products
        query = (
            select(Product)
            .join(MarketplaceAccount, MarketplaceAccount.id == Product.marketplace_account_id)
            .where(MarketplaceAccount.user_id == user_id)
            .where(Product.marketplace == Marketplace.WB)
            .where(Product.is_active.is_(True))
            .order_by(Product.seller_article)
        )
        if account_id:
            query = query.where(Product.marketplace_account_id == account_id)

        result = await self.session.execute(query)
        products = list(result.scalars().all())

        # Fetch promo data for products
        promo_map = await self._build_promo_map(user_id, products)

        for row_idx, product in enumerate(products, 2):
            wb_nm_id = _extract_nm_id(product)
            ws.cell(row=row_idx, column=1, value=product.id).fill = SERVICE_FILL
            ws.cell(row=row_idx, column=2, value=wb_nm_id).fill = SERVICE_FILL
            ws.cell(row=row_idx, column=3, value=product.seller_article).fill = SERVICE_FILL
            ws.cell(row=row_idx, column=4, value="").fill = SERVICE_FILL
            ws.cell(row=row_idx, column=5, value=product.brand)
            ws.cell(row=row_idx, column=6, value=(product.title or "")[:100])
            ws.cell(row=row_idx, column=7, value="")
            ws.cell(row=row_idx, column=8, value=float(product.mrc_price) if product.mrc_price else "")
            ws.cell(row=row_idx, column=9, value="").fill = INPUT_FILL
            ws.cell(row=row_idx, column=10, value="").fill = SERVICE_FILL

            promo_key = (product.marketplace_account_id, wb_nm_id)
            promo_info = promo_map.get(promo_key)
            if promo_info:
                ws.cell(row=row_idx, column=11, value=promo_info["name"])
                ws.cell(row=row_idx, column=12, value=float(promo_info["plan_price"]) if promo_info["plan_price"] else "")

        # Auto-filter and freeze panes
        ws.auto_filter.ref = f"A1:{_col_letter(len(TEMPLATE_COLUMNS))}{len(products) + 1}"
        ws.freeze_panes = "A2"

        # Column widths
        widths = [12, 14, 18, 18, 16, 40, 14, 14, 14, 14, 24, 16, 18, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[_col_letter(i)].width = w

        # Instructions sheet
        ws_instr = wb.create_sheet("Инструкция")
        instructions = [
            "Инструкция по заполнению МРЦ",
            "",
            "1. Заполните только колонку 'new_mrc_price' (жёлтый фон).",
            "2. МРЦ — это целевая цена товара со скидкой на Wildberries.",
            "3. Цена продавца до скидки будет рассчитана автоматически как МРЦ × 4.",
            "4. Если товар участвует в подходящей акции WB, цена может быть снижена, но не более чем на 10% от МРЦ.",
            "5. Не меняйте product_id и wb_nm_id — это служебные поля.",
            "6. Если оставить new_mrc_price пустым, МРЦ по товару не изменится.",
            "7. Чтобы очистить МРЦ, укажите CLEAR в колонке new_mrc_price.",
            "8. Значение 0 или отрицательное число вызовет ошибку.",
        ]
        for i, line in enumerate(instructions, 1):
            ws_instr.cell(row=i, column=1, value=line)

        # Save
        output_dir = Path("/tmp")
        output_dir.mkdir(exist_ok=True)
        file_path = output_dir / f"mrc_prices_wb_{datetime.now(tz=UTC).strftime('%Y-%m-%d')}.xlsx"
        wb.save(str(file_path))
        return file_path

    async def parse_mrc_import_file(
        self,
        file_path: Path,
        user_id: int,
    ) -> MrcImportPreview:
        """Parse import file and validate rows."""
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl not installed")

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        # Check headers
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        headers = [str(h).strip() if h else "" for h in headers]

        missing = REQUIRED_COLUMNS - set(headers)
        if missing:
            raise ValueError(f"Отсутствуют обязательные колонки: {', '.join(missing)}")

        col_map = {h: i for i, h in enumerate(headers, 1)}

        preview = MrcImportPreview(
            preview_id=str(uuid.uuid4()),
            user_id=user_id,
            file_name=file_path.name,
            total_rows=0,
        )

        seen_product_ids: set[int] = set()
        seen_wb_nm_ids: set[int] = set()

        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for col_name, col_idx in col_map.items():
                row_data[col_name] = ws.cell(row=row_idx, column=col_idx).value

            product_id = _safe_int(row_data.get("product_id"))
            wb_nm_id = _safe_int(row_data.get("wb_nm_id"))
            new_mrc_raw = _safe_str(row_data.get("new_mrc_price"))

            row = MrcImportRow(
                row_number=row_idx,
                product_id=product_id,
                wb_nm_id=wb_nm_id,
                seller_sku=_safe_str(row_data.get("seller_sku")),
                product_name=_safe_str(row_data.get("product_name")),
                new_mrc_price_raw=new_mrc_raw,
                old_mrc_price=None,
            )

            preview.total_rows += 1

            if preview.total_rows > self._max_rows():
                row.status = "error"
                row.message = f"Превышен лимит строк (макс. {self._max_rows()})."
                preview.error_rows.append(row)
                continue

            # Empty new_mrc_price → skip
            if not new_mrc_raw or new_mrc_raw.strip() == "":
                row.status = "skipped_empty"
                row.message = "Пустое значение, МРЦ не изменится."
                preview.skipped_rows.append(row)
                continue

            # CLEAR
            if new_mrc_raw.strip().upper() == "CLEAR":
                if not self._allow_clear():
                    row.status = "error"
                    row.message = "Очистка МРЦ запрещена настройкой."
                    preview.error_rows.append(row)
                    continue
                row.status = "valid_clear"
                preview.valid_rows.append(row)
                continue

            # Validate number
            try:
                mrc_val = Decimal(new_mrc_raw.replace(",", "."))
            except (InvalidOperation, ValueError):
                row.status = "error"
                row.message = "МРЦ должна быть числом."
                preview.error_rows.append(row)
                continue

            if mrc_val <= 0:
                row.status = "error"
                row.message = "МРЦ должна быть больше 0."
                preview.error_rows.append(row)
                continue

            # Validate product_id
            if not product_id:
                row.status = "error"
                row.message = "product_id должен быть числом."
                preview.error_rows.append(row)
                continue

            # Check duplicates
            if product_id in seen_product_ids:
                row.status = "error"
                row.message = "Дубликат product_id в файле."
                preview.error_rows.append(row)
                continue
            seen_product_ids.add(product_id)

            if wb_nm_id and wb_nm_id in seen_wb_nm_ids:
                row.status = "error"
                row.message = "Дубликат wb_nm_id в файле."
                preview.error_rows.append(row)
                continue
            if wb_nm_id:
                seen_wb_nm_ids.add(wb_nm_id)

            # Verify product exists and belongs to user
            product = await self.session.get(Product, product_id)
            if product is None:
                row.status = "error"
                row.message = "Товар не найден в базе."
                preview.error_rows.append(row)
                continue

            if product.user_id != user_id:
                row.status = "error"
                row.message = "Товар принадлежит другому пользователю."
                preview.error_rows.append(row)
                continue

            if product.marketplace != Marketplace.WB:
                row.status = "error"
                row.message = "Товар не относится к Wildberries."
                preview.error_rows.append(row)
                continue

            # Verify wb_nm_id matches
            product_nm_id = _extract_nm_id(product)
            if wb_nm_id and product_nm_id and wb_nm_id != product_nm_id:
                row.status = "warning"
                row.message = f"wb_nm_id в файле ({wb_nm_id}) не совпадает с товаром ({product_nm_id})."
                preview.warning_rows.append(row)
                continue

            row.old_mrc_price = product.mrc_price

            # Check if value changed
            if product.mrc_price and mrc_val == product.mrc_price:
                row.status = "skipped_no_change"
                row.message = "МРЦ не изменилась."
                preview.skipped_rows.append(row)
                continue

            row.status = "valid"
            preview.valid_rows.append(row)

        self._previews[preview.preview_id] = preview
        return preview

    async def apply_mrc_import(
        self,
        preview_id: str,
        user_id: int,
    ) -> MrcImportResult:
        """Apply validated MRC import to database."""
        preview = self._previews.get(preview_id)
        if preview is None:
            raise ValueError("Предварительный просмотр не найден или истёк.")

        if preview.user_id != user_id:
            raise ValueError("Доступ запрещён.")

        result = MrcImportResult(
            preview_id=preview_id,
            user_id=user_id,
            updated_count=0,
            cleared_count=0,
            skipped_count=0,
            error_count=0,
        )

        for row in preview.valid_rows:
            if not row.product_id:
                result.error_count += 1
                continue

            product = await self.session.get(Product, product_id=row.product_id)
            if product is None or product.user_id != user_id:
                result.error_count += 1
                continue

            old_mrc = product.mrc_price

            if row.status == "valid_clear":
                product.mrc_price = None
                result.cleared_count += 1
            else:
                try:
                    mrc_val = Decimal(row.new_mrc_price_raw.replace(",", "."))
                    product.mrc_price = mrc_val.quantize(Decimal("0.01"))
                    result.updated_count += 1
                except (InvalidOperation, ValueError):
                    result.error_count += 1
                    continue

            logger.info(
                "mrc_price_updated_import",
                extra={
                    "user_id": user_id,
                    "product_id": product.id,
                    "wb_nm_id": _extract_nm_id(product),
                    "old_mrc_price": str(old_mrc),
                    "new_mrc_price": str(product.mrc_price),
                    "source": "import",
                },
            )

        for row in preview.skipped_rows:
            result.skipped_count += 1

        for row in preview.error_rows:
            result.error_count += 1

        await self.session.commit()

        # Clean up preview
        self._previews.pop(preview_id, None)

        return result

    async def _build_promo_map(
        self,
        user_id: int,
        products: list[Product],
    ) -> dict[tuple[int, int | None], dict[str, Any]]:
        """Build map of (account_id, wb_nm_id) → promo info for template."""
        from datetime import UTC, datetime

        now_utc = datetime.now(tz=UTC)

        # Get active promotions for user's accounts
        account_ids = {p.marketplace_account_id for p in products}
        result = await self.session.execute(
            select(WbPromotionNomenclature, WbPromotion.name)
            .join(
                WbPromotion,
                (WbPromotion.wb_promotion_id == WbPromotionNomenclature.wb_promotion_id)
                & (WbPromotion.marketplace_account_id == WbPromotionNomenclature.marketplace_account_id),
            )
            .where(
                WbPromotionNomenclature.marketplace_account_id.in_(account_ids),
                WbPromotion.is_active_today.is_(True),
                WbPromotion.start_datetime <= now_utc,
                WbPromotion.end_datetime >= now_utc,
                WbPromotionNomenclature.plan_price.isnot(None),
                WbPromotionNomenclature.plan_price > 0,
            )
        )

        promo_map: dict[tuple[int, int | None], dict[str, Any]] = {}
        for nom, promo_name in result.all():
            key = (nom.marketplace_account_id, nom.wb_nm_id)
            if key not in promo_map:
                promo_map[key] = {"name": promo_name, "plan_price": nom.plan_price}

        return promo_map

    def generate_import_report(self, result: MrcImportResult) -> Path:
        """Generate Excel report of import results."""
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl not installed")

        preview = self._previews.get(result.preview_id)
        if not preview:
            raise ValueError("Preview not found")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Результат импорта"

        report_cols = [
            "row_number", "product_id", "wb_nm_id", "seller_sku",
            "product_name", "old_mrc_price", "new_mrc_price", "status", "message",
        ]

        for col_idx, header in enumerate(report_cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        all_rows = (
            preview.valid_rows
            + preview.skipped_rows
            + preview.warning_rows
            + preview.error_rows
        )
        all_rows.sort(key=lambda r: r.row_number)

        for row_idx, row in enumerate(all_rows, 2):
            ws.cell(row=row_idx, column=1, value=row.row_number)
            ws.cell(row=row_idx, column=2, value=row.product_id)
            ws.cell(row=row_idx, column=3, value=row.wb_nm_id)
            ws.cell(row=row_idx, column=4, value=row.seller_sku)
            ws.cell(row=row_idx, column=5, value=row.product_name)
            ws.cell(row=row_idx, column=6, value=float(row.old_mrc_price) if row.old_mrc_price else "")
            ws.cell(row=row_idx, column=7, value=row.new_mrc_price_raw)
            ws.cell(row=row_idx, column=8, value=row.status)
            ws.cell(row=row_idx, column=9, value=row.message)

        output_dir = Path("/tmp")
        output_dir.mkdir(exist_ok=True)
        ts = datetime.now(tz=UTC).strftime("%Y-%m-%d_%H-%M")
        file_path = output_dir / f"mrc_import_result_{ts}.xlsx"
        wb.save(str(file_path))
        return file_path


def _extract_nm_id(product: Product) -> int | None:
    """Extract WB nmID from product."""
    if product.marketplace_article and product.marketplace_article.isdigit():
        return int(product.marketplace_article)
    if product.external_product_id and product.external_product_id.isdigit():
        return int(product.external_product_id)
    return None


def _safe_int(value: Any) -> int | None:
    """Safely convert to int."""
    if value is None:
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


def _safe_str(value: Any) -> str | None:
    """Safely convert to string."""
    if value is None:
        return None
    return str(value).strip()


def _col_letter(n: int) -> str:
    """Convert column number to Excel letter (1→A, 2→B, 27→AA)."""
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result
