"""Import WB auto-promotion condition files from seller cabinet reports."""

import asyncio
import csv
import io
import re
from collections.abc import Iterable
from dataclasses import dataclass, field
from datetime import UTC, datetime
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.domain import (
    WbAutoPromoFileImport,
    WbAutoPromoFileImportRow,
    WbAutoPromotionCondition,
)

MAX_IMPORT_FILE_SIZE = 15 * 1024 * 1024
REPORT_SHEET_NAME = "Отчёт по скидкам"

HEADER_ALIASES = {
    "товар уже участвует в акции": "already_participating",
    "бренд": "brand",
    "предмет": "subject",
    "наименование": "title",
    "артикул поставщика": "seller_article",
    "артикул wb": "wb_nm_id",
    "последний баркод": "barcode",
    "плановая цена для акции": "plan_price",
    "текущая розничная цена": "current_full_price",
    "валюта": "currency",
    "текущая скидка на сайте, %": "current_discount_percent",
    "загружаемая скидка для участия в акции": "wb_upload_discount_percent",
    "статус": "wb_status",
}

INDEX_FALLBACK = {
    0: "already_participating",
    1: "brand",
    2: "subject",
    3: "title",
    4: "seller_article",
    5: "wb_nm_id",
    6: "barcode",
    11: "plan_price",
    12: "current_full_price",
    13: "currency",
    14: "current_discount_percent",
    15: "wb_upload_discount_percent",
    16: "wb_status",
}


@dataclass(slots=True)
class AutoPromoFileRow:
    row_number: int
    brand: str | None
    subject: str | None
    title: str | None
    seller_article: str | None
    wb_nm_id: int | None
    barcode: str | None
    already_participating: bool | None
    plan_price: Decimal | None
    current_full_price: Decimal | None
    currency: str | None
    current_discount_percent: Decimal | None
    wb_upload_discount_percent: Decimal | None
    wb_status: str | None
    current_discounted_price: Decimal | None
    source: str = "wb_file"
    raw_payload: dict[str, Any] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def status(self) -> str:
        if self.errors:
            return "error"
        if self.warnings:
            return "warning"
        return "valid"

    @property
    def message(self) -> str | None:
        messages = [*self.errors, *self.warnings]
        return "; ".join(messages) if messages else None

    @property
    def candidate_discounted_price(self) -> Decimal | None:
        if self.plan_price is not None:
            return self.plan_price
        return WbAutoPromoFileImportService.discounted_price(
            self.current_full_price,
            self.wb_upload_discount_percent,
        )

    @property
    def condition_type(self) -> str:
        if self.plan_price is not None:
            return "max_price"
        if self.candidate_discounted_price is not None:
            return "discount_projection"
        return "unknown"


@dataclass(slots=True)
class AutoPromoFilePreview:
    total_rows: int
    valid_rows: int
    error_rows: int
    warning_rows: int
    already_participating_count: int
    not_participating_count: int
    with_plan_price_count: int
    without_plan_price_count: int
    import_id: int = 0
    promotion_name: str | None = None


@dataclass(slots=True)
class AutoPromoFileParseResult:
    preview: AutoPromoFilePreview
    rows: list[AutoPromoFileRow]


class WbAutoPromoFileImportService:
    """Parse, preview, and persist WB auto-promotion condition reports."""

    def __init__(self, session: AsyncSession | None = None) -> None:
        self.session = session

    async def parse_file(self, file_path: Path) -> AutoPromoFileParseResult:
        self._validate_file(file_path)
        suffix = file_path.suffix.lower()
        try:
            if suffix in (".xlsx", ".xlsm"):
                raw_rows = await asyncio.to_thread(self._read_xlsx_rows, file_path)
            elif suffix == ".csv":
                raw_rows = await asyncio.to_thread(self._read_csv_rows, file_path)
            else:
                raise ValueError("Файл должен быть .xlsx, .xlsm или .csv")
        except BadZipFile:
            raise ValueError("Excel-файл повреждён или сохранён не в формате XLSX") from None
        rows = self._parse_rows(raw_rows)
        return AutoPromoFileParseResult(preview=self._build_preview(rows), rows=rows)

    async def create_preview(
        self,
        file_path: Path,
        user_id: int,
        marketplace_account_id: int,
        original_file_name: str | None = None,
        promotion_name: str | None = None,
    ) -> tuple[AutoPromoFilePreview, list[AutoPromoFileRow]]:
        if self.session is None:
            raise RuntimeError("AsyncSession is required to create preview")
        parsed = await self.parse_file(file_path)
        resolved_name = promotion_name or self.extract_promotion_name(original_file_name)

        import_record = WbAutoPromoFileImport(
            user_id=user_id,
            marketplace_account_id=marketplace_account_id,
            original_file_name=original_file_name,
            promotion_name=resolved_name,
            status="preview",
            total_rows=parsed.preview.total_rows,
            valid_rows=parsed.preview.valid_rows,
            error_rows=parsed.preview.error_rows,
            warning_rows=parsed.preview.warning_rows,
        )
        self.session.add(import_record)
        await self.session.flush()

        for row in parsed.rows:
            self.session.add(self._to_import_row(import_record.id, row))
        await self.session.flush()

        parsed.preview.import_id = import_record.id
        parsed.preview.promotion_name = resolved_name
        return parsed.preview, parsed.rows

    async def apply_import(
        self,
        preview_rows: list[AutoPromoFileRow | dict[str, Any]],
        user_id: int,
        marketplace_account_id: int,
        promotion_name: str | None = None,
    ) -> int:
        rows = [self._coerce_row(row) for row in preview_rows]
        return await self.save_conditions(rows, user_id, marketplace_account_id, promotion_name)

    async def apply_import_record(
        self,
        import_id: int,
        user_id: int,
        promotion_name: str | None = None,
    ) -> int:
        if self.session is None:
            raise RuntimeError("AsyncSession is required to apply import")
        import_record = await self.session.get(WbAutoPromoFileImport, import_id)
        if import_record is None or import_record.user_id != user_id:
            raise ValueError("Импорт не найден")
        rows_result = await self.session.execute(
            select(WbAutoPromoFileImportRow)
            .where(WbAutoPromoFileImportRow.import_id == import_id)
            .order_by(WbAutoPromoFileImportRow.row_number)
        )
        rows = [self._from_import_row(row) for row in rows_result.scalars().all()]
        saved = await self.save_conditions(
            rows,
            user_id,
            import_record.marketplace_account_id,
            promotion_name or import_record.promotion_name,
        )
        import_record.status = "applied"
        import_record.applied_at = datetime.now(tz=UTC)
        await self.session.flush()
        return saved

    async def save_conditions(
        self,
        rows: list[AutoPromoFileRow],
        user_id: int,
        marketplace_account_id: int,
        promotion_name: str | None = None,
    ) -> int:
        if self.session is None:
            raise RuntimeError("AsyncSession is required to save conditions")
        saved = 0
        now = datetime.now(tz=UTC)
        for row in rows:
            if row.status == "error" or row.wb_nm_id is None:
                continue
            existing = await self.session.execute(
                select(WbAutoPromotionCondition).where(
                    WbAutoPromotionCondition.marketplace_account_id
                    == marketplace_account_id,
                    WbAutoPromotionCondition.wb_nm_id == row.wb_nm_id,
                    WbAutoPromotionCondition.source == "wb_file",
                    WbAutoPromotionCondition.promotion_name == (promotion_name or ""),
                )
            )
            condition = existing.scalar_one_or_none()
            if condition is None:
                condition = WbAutoPromotionCondition(
                    user_id=user_id,
                    marketplace_account_id=marketplace_account_id,
                    wb_nm_id=row.wb_nm_id,
                    source="wb_file",
                )
                self.session.add(condition)
            candidate = row.candidate_discounted_price
            condition.seller_article = row.seller_article
            condition.title = row.title
            condition.promotion_name = promotion_name or ""
            condition.required_price = row.plan_price
            condition.max_auto_promo_price = row.plan_price
            condition.wb_condition_discount_percent = row.wb_upload_discount_percent
            condition.current_wb_price = row.current_discounted_price
            condition.current_full_price = row.current_full_price
            condition.current_discount = self.decimal_to_int(row.current_discount_percent)
            condition.current_discounted_price = row.current_discounted_price
            condition.candidate_discounted_price = candidate
            condition.condition_type = row.condition_type
            condition.is_participating = row.already_participating
            condition.confidence = "high" if row.plan_price is not None else "medium"
            condition.raw_payload = row.raw_payload
            condition.synced_at = now
            saved += 1
        await self.session.flush()
        return saved

    async def load_preview(
        self,
        import_id: int,
        user_id: int,
    ) -> tuple[WbAutoPromoFileImport, list[WbAutoPromoFileImportRow]]:
        if self.session is None:
            raise RuntimeError("AsyncSession is required to load preview")
        import_record = await self.session.get(WbAutoPromoFileImport, import_id)
        if import_record is None or import_record.user_id != user_id:
            raise ValueError("Импорт не найден")
        rows_result = await self.session.execute(
            select(WbAutoPromoFileImportRow)
            .where(WbAutoPromoFileImportRow.import_id == import_id)
            .order_by(WbAutoPromoFileImportRow.row_number)
        )
        return import_record, list(rows_result.scalars().all())

    @staticmethod
    def extract_promotion_name(file_name: str | None) -> str | None:
        if not file_name:
            return None
        stem = Path(file_name).stem
        stem = re.sub(r"^Все товары подходящие для акции_", "", stem)
        stem = re.sub(r"_\d{2}\.\d{2}\.\d{4}.*$", "", stem)
        return stem.strip("_ ") or None

    def _read_xlsx_rows(self, file_path: Path) -> list[tuple[Any, ...]]:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb[REPORT_SHEET_NAME] if REPORT_SHEET_NAME in wb.sheetnames else wb.active
            return list(ws.iter_rows(values_only=True))
        finally:
            wb.close()

    @staticmethod
    def _read_csv_rows(file_path: Path) -> list[list[str]]:
        content = file_path.read_text(encoding="utf-8-sig")
        return list(csv.reader(io.StringIO(content)))

    def _parse_rows(self, raw_rows: list[Iterable[Any]]) -> list[AutoPromoFileRow]:
        iterator = iter(raw_rows)
        headers = next(iterator, None)
        if headers is None:
            raise ValueError("Файл пустой")
        mapping = self._resolve_headers(tuple(headers))
        rows: list[AutoPromoFileRow] = []
        for row_number, values in enumerate(iterator, start=2):
            values_tuple = tuple(values)
            if not any(value not in (None, "") for value in values_tuple):
                continue
            rows.append(self._parse_row(row_number, values_tuple, mapping))
        return rows

    def _parse_row(
        self,
        row_number: int,
        values: tuple[Any, ...],
        mapping: dict[int, str],
    ) -> AutoPromoFileRow:
        data = {
            key: values[idx] if idx < len(values) else None
            for idx, key in mapping.items()
        }
        current_full_price = self.parse_decimal(data.get("current_full_price"))
        current_discount = self.parse_decimal(data.get("current_discount_percent"))
        current_discounted = self.discounted_price(current_full_price, current_discount)
        plan_price = self.parse_decimal(data.get("plan_price"))
        upload_discount = self.parse_decimal(data.get("wb_upload_discount_percent"))
        row = AutoPromoFileRow(
            row_number=row_number,
            brand=self.clean_text(data.get("brand")),
            subject=self.clean_text(data.get("subject")),
            title=self.clean_text(data.get("title")),
            seller_article=self.clean_text(data.get("seller_article")),
            wb_nm_id=self.parse_int(data.get("wb_nm_id")),
            barcode=self.clean_text(data.get("barcode")),
            already_participating=self.parse_bool(data.get("already_participating")),
            plan_price=plan_price,
            current_full_price=current_full_price,
            currency=self.clean_text(data.get("currency")),
            current_discount_percent=current_discount,
            wb_upload_discount_percent=upload_discount,
            wb_status=self.clean_text(data.get("wb_status")),
            current_discounted_price=current_discounted,
            raw_payload=self._raw_payload(row_number, data, current_discounted),
        )
        if row.wb_nm_id is None:
            row.errors.append("Артикул WB не указан или некорректен")
        if row.plan_price is None and row.wb_upload_discount_percent is None:
            row.warnings.append("Плановая цена и скидка WB отсутствуют")
        elif row.plan_price is None:
            row.warnings.append(
                "Плановая цена отсутствует, будет использован fallback по скидке WB"
            )
        return row

    @staticmethod
    def _resolve_headers(raw_headers: tuple[Any, ...]) -> dict[int, str]:
        mapping: dict[int, str] = {}
        for idx, header in enumerate(raw_headers):
            normalized = WbAutoPromoFileImportService.normalize_header(header)
            if normalized in HEADER_ALIASES:
                mapping[idx] = HEADER_ALIASES[normalized]
        for idx, key in INDEX_FALLBACK.items():
            if key not in mapping.values():
                mapping[idx] = key
        return mapping

    @staticmethod
    def _build_preview(rows: list[AutoPromoFileRow]) -> AutoPromoFilePreview:
        return AutoPromoFilePreview(
            total_rows=len(rows),
            valid_rows=sum(1 for row in rows if row.status == "valid"),
            error_rows=sum(1 for row in rows if row.status == "error"),
            warning_rows=sum(1 for row in rows if row.status == "warning"),
            already_participating_count=sum(1 for row in rows if row.already_participating),
            not_participating_count=sum(1 for row in rows if row.already_participating is False),
            with_plan_price_count=sum(1 for row in rows if row.plan_price is not None),
            without_plan_price_count=sum(1 for row in rows if row.plan_price is None),
        )

    @staticmethod
    def _to_import_row(import_id: int, row: AutoPromoFileRow) -> WbAutoPromoFileImportRow:
        return WbAutoPromoFileImportRow(
            import_id=import_id,
            row_number=row.row_number,
            wb_nm_id=row.wb_nm_id,
            seller_article=row.seller_article,
            title=row.title,
            plan_price=row.plan_price,
            current_full_price=row.current_full_price,
            current_discount_percent=row.current_discount_percent,
            current_discounted_price=row.current_discounted_price,
            wb_upload_discount_percent=row.wb_upload_discount_percent,
            wb_status=row.wb_status,
            already_participating=row.already_participating,
            status=row.status,
            message=row.message,
            raw_payload=row.raw_payload,
        )

    @staticmethod
    def _from_import_row(row: WbAutoPromoFileImportRow) -> AutoPromoFileRow:
        raw = row.raw_payload or {}
        return AutoPromoFileRow(
            row_number=row.row_number,
            brand=raw.get("brand"),
            subject=raw.get("subject"),
            title=row.title,
            seller_article=row.seller_article,
            wb_nm_id=row.wb_nm_id,
            barcode=raw.get("barcode"),
            already_participating=row.already_participating,
            plan_price=row.plan_price,
            current_full_price=row.current_full_price,
            currency=raw.get("currency"),
            current_discount_percent=row.current_discount_percent,
            wb_upload_discount_percent=row.wb_upload_discount_percent,
            wb_status=row.wb_status,
            current_discounted_price=row.current_discounted_price,
            raw_payload=raw,
            errors=[row.message] if row.status == "error" and row.message else [],
            warnings=[row.message] if row.status == "warning" and row.message else [],
        )

    @staticmethod
    def _coerce_row(row: AutoPromoFileRow | dict[str, Any]) -> AutoPromoFileRow:
        if isinstance(row, AutoPromoFileRow):
            return row
        return AutoPromoFileRow(
            row_number=int(row.get("row_number") or row.get("row_num") or 0),
            brand=row.get("brand"),
            subject=row.get("subject"),
            title=row.get("title"),
            seller_article=row.get("seller_article"),
            wb_nm_id=WbAutoPromoFileImportService.parse_int(row.get("wb_nm_id")),
            barcode=row.get("barcode"),
            already_participating=WbAutoPromoFileImportService.parse_bool(
                row.get("already_participating")
            ),
            plan_price=WbAutoPromoFileImportService.parse_decimal(row.get("plan_price")),
            current_full_price=WbAutoPromoFileImportService.parse_decimal(
                row.get("current_full_price")
            ),
            currency=row.get("currency"),
            current_discount_percent=WbAutoPromoFileImportService.parse_decimal(
                row.get("current_discount_percent")
            ),
            wb_upload_discount_percent=WbAutoPromoFileImportService.parse_decimal(
                row.get("wb_upload_discount_percent")
            ),
            wb_status=row.get("wb_status"),
            current_discounted_price=WbAutoPromoFileImportService.parse_decimal(
                row.get("current_discounted_price")
            ),
            raw_payload=row.get("raw_payload") or {},
        )

    @staticmethod
    def _validate_file(file_path: Path) -> None:
        if file_path.suffix.lower() not in {".xlsx", ".xlsm", ".csv"}:
            raise ValueError("Файл должен быть .xlsx, .xlsm или .csv")
        if file_path.stat().st_size > MAX_IMPORT_FILE_SIZE:
            raise ValueError("Файл слишком большой")

    @staticmethod
    def _raw_payload(
        row_number: int,
        data: dict[str, Any],
        current_discounted_price: Decimal | None,
    ) -> dict[str, Any]:
        result: dict[str, Any] = {"row_number": row_number}
        for key, value in data.items():
            result[key] = str(value).strip() if value is not None else None
        result["current_discounted_price"] = (
            str(current_discounted_price) if current_discounted_price is not None else None
        )
        result["wb_upload_discount_is_diagnostic"] = data.get("plan_price") not in (None, "")
        return result

    @staticmethod
    def normalize_header(value: Any) -> str:
        return re.sub(r"\s+", " ", str(value or "").strip().lower())

    @staticmethod
    def clean_text(value: Any) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    @staticmethod
    def parse_decimal(value: Any) -> Decimal | None:
        if value is None or value == "":
            return None
        try:
            normalized = (
                str(value)
                .replace("\xa0", "")
                .replace(" ", "")
                .replace("₽", "")
                .replace("%", "")
                .replace(",", ".")
                .strip()
            )
            return Decimal(normalized).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        except (InvalidOperation, ValueError, TypeError):
            return None

    @staticmethod
    def parse_int(value: Any) -> int | None:
        parsed = WbAutoPromoFileImportService.parse_decimal(value)
        return int(parsed) if parsed is not None else None

    @staticmethod
    def parse_bool(value: Any) -> bool | None:
        if value is None:
            return None
        if isinstance(value, bool):
            return value
        text = str(value).strip().lower()
        if text in {"да", "yes", "true", "1", "участвует"}:
            return True
        if text in {"нет", "no", "false", "0", "не участвует"}:
            return False
        return None

    @staticmethod
    def discounted_price(
        full_price: Decimal | None,
        discount_percent: Decimal | None,
    ) -> Decimal | None:
        if full_price is None or discount_percent is None:
            return None
        if full_price <= 0 or discount_percent < 0 or discount_percent >= 100:
            return None
        return (full_price * (Decimal("1") - discount_percent / Decimal("100"))).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )

    @staticmethod
    def decimal_to_int(value: Decimal | None) -> int | None:
        return int(value) if value is not None else None
