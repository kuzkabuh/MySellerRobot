"""version: 1.1.0
description: Robust XLSX parser for WB daily and weekly detailed reports.
updated: 2026-06-08
"""

from __future__ import annotations

import hashlib
import json
import logging
import re
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

logger = logging.getLogger(__name__)

REPORT_NUMBER_PATTERN = re.compile(r"(\d{8,})_(?P<supplier>\d+)", re.IGNORECASE)


@dataclass(slots=True)
class WbDailyReportParsed:
    report_number: str
    report_type: str
    report_date: date | None
    report_period_start: date | None
    report_period_end: date | None
    source_filename: str | None = None
    rows: list[WbDailyReportParsedRow] = field(default_factory=list)
    skipped_rows: int = 0


@dataclass(slots=True)
class WbDailyReportParsedRow:
    row_number: int | None
    report_type: str
    sale_dt: datetime | None
    order_dt: datetime | None
    nm_id: int | None
    supplier_article: str | None
    product_name: str | None
    size: str | None
    barcode: str | None
    shk: str | None
    srid: str | None
    srid_normalized: str | None
    rid_normalized: str | None
    doc_type_name: str | None
    payment_reason: str | None
    subject_name: str | None
    brand_name: str | None
    quantity: int | None
    retail_price: Decimal | None
    retail_amount: Decimal | None
    for_pay: Decimal | None
    delivery_count: int | None
    return_count: int | None
    delivery_rub: Decimal | None
    penalty: Decimal | None
    storage_fee: Decimal | None
    acceptance: Decimal | None
    deduction: Decimal | None
    commission_rub: Decimal | None
    commission_correction_amount: Decimal | None
    reimbursement_amount: Decimal | None
    logistics_penalty_correction_type: str | None
    basket_id: str | None
    sale_method: str | None
    finance_operation_type: str
    finance_category: str
    order_required: bool
    raw: dict[str, object]

    def compute_hash(self) -> str:
        return self.compute_source_row_hash()

    def compute_source_row_hash(self) -> str:
        encoded = json.dumps(
            _normalize_hash_payload(self.raw),
            sort_keys=True,
            ensure_ascii=False,
            default=str,
        ).encode("utf-8")
        return hashlib.sha256(encoded).hexdigest()

    def compute_stable_business_key(
        self,
        *,
        marketplace_account_id: int,
        report_number: str,
    ) -> str:
        payload = {
            "marketplace_account_id": marketplace_account_id,
            "report_number": report_number,
            "report_type": self.report_type,
            "row_number": self.row_number,
            "nm_id": self.nm_id,
            "supplier_article": _stable_string(self.supplier_article),
            "barcode": _stable_string(self.barcode),
            "shk": _stable_string(self.shk),
            "srid": _stable_string(self.srid),
            "srid_normalized": _stable_string(self.srid_normalized),
            "basket_id": _stable_string(self.basket_id),
            "payment_reason": _stable_string(self.payment_reason),
            "doc_type_name": _stable_string(self.doc_type_name),
            "sale_dt": self.sale_dt.isoformat() if self.sale_dt else None,
        }
        encoded = json.dumps(payload, sort_keys=True, ensure_ascii=False).encode("utf-8")
        return hashlib.sha256(encoded).hexdigest()


COLUMN_ALIASES: dict[str, str] = {
    "№": "row_number",
    "Номер поставки": "report_number",
    "Номер отчёта": "report_number",
    "Предмет": "subject_name",
    "Код номенклатуры": "nm_id",
    "Бренд": "brand_name",
    "Артикул поставщика": "supplier_article",
    "Название": "product_name",
    "Размер": "size",
    "Баркод": "barcode",
    "ШК": "shk",
    "Srid": "srid",
    "SRID": "srid",
    "srid": "srid",
    "Тип документа": "doc_type_name",
    "Обоснование для оплаты": "payment_reason",
    "Дата заказа покупателем": "order_dt",
    "Дата продажи": "sale_dt",
    "Кол-во": "quantity",
    "Цена розничная": "retail_price",
    "Вайлдберриз реализовал Товар (Пр)": "retail_amount",
    "Согласованный продуктовый дисконт, %": "agreed_product_discount",
    "Промокод, %": "promo_discount",
    "Итоговая согласованная скидка, %": "final_discount",
    "Цена розничная с учетом согласованной скидки": "retail_price_discounted",
    "Скидка постоянного покупателя, %": "loyalty_discount",
    "Скидка продавца, %": "seller_discount",
    "Комиссия, %": "commission_percent",
    "Размер комиссии": "commission_rub",
    "Вознаграждение Вайлдберриз (ВВ), без НДС": "commission_rub",
    "НДС с Вознаграждения Вайлдберриз": "commission_vat",
    "К перечислению Продавцу за реализованный Товар": "for_pay",
    "Количество доставок": "delivery_count",
    "Количество возврата": "return_count",
    "Стоимость логистики": "delivery_rub",
    "Услуги по доставке товара покупателю": "delivery_rub",
    "Стоимость хранения": "storage_fee",
    "Хранение": "storage_fee",
    "Стоимость приёмки": "acceptance",
    "Операции на приемке": "acceptance",
    "Удержание": "deduction",
    "Удержания": "deduction",
    "Штраф": "penalty",
    "Общая сумма штрафов": "penalty",
    "Корректировка Вознаграждения Вайлдберриз (ВВ)": "commission_correction_amount",
    "Виды логистики, штрафов и корректировок ВВ": "logistics_penalty_correction_type",
    "Возмещение издержек по перевозке/по складским операциям с товаром": "reimbursement_amount",
    "Id корзины заказа": "basket_id",
    "Способы продажи и тип товара": "sale_method",
    "Компенсация платёжных услуг": "payment_services_amount",
    "Комиссия за интеграцию платёжных сервисов": "payment_services_amount",
    "Возмещение за выдачу и возврат товаров на ПВЗ": "pvz_amount",
    "Стоимость участия в программе лояльности": "loyalty_program_fee",
    "Сумма баллов, удержанных по программе лояльности": "loyalty_points_deduction",
    "Компенсация скидки по программе лояльности": "loyalty_discount_compensation",
}

WEEKLY_REPORT_EXPECTED_COLUMNS = 82

DATE_KEYS = {"order_dt", "sale_dt"}
REQUIRED_COLUMNS = {
    "nm_id",
    "supplier_article",
    "barcode",
    "payment_reason",
    "doc_type_name",
    "sale_dt",
    "quantity",
    "retail_amount",
    "commission_rub",
    "for_pay",
    "shk",
    "srid",
}


def iter_wb_daily_report_rows(
    xlsx_bytes: bytes,
    *,
    max_rows: int = 200_000,
    report_number_hint: str | None = None,
    report_type: str = "daily",
    source_filename: str | None = None,
) -> WbDailyReportParsed:
    """Parse a WB daily realisation-report XLSX file.

    Returns a WbDailyReportParsed object with the detected report number and
    normalized rows. The function uses openpyxl's read_only mode and falls
    back to ``worksheet.reset_dimensions()`` for files where the stored
    ``dimension = A1`` does not reflect the real data extent.
    """
    workbook = load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    worksheet = workbook.active
    if worksheet is None:
        raise ValueError("В файле отчёта WB нет активного листа")

    try:
        worksheet.reset_dimensions()
    except Exception:
        logger.debug("ws_reset_dimensions_failed", exc_info=True)

    rows_iter = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ValueError("Файл отчёта WB пустой или не содержит заголовки") from None

    headers = [_stringify(cell) for cell in header_row]
    if not any(headers):
        raise ValueError("Файл отчёта WB пустой или не содержит заголовки")

    mapped_columns = [_resolve_column(header) for header in headers]
    missing_columns = sorted(REQUIRED_COLUMNS - {column for column in mapped_columns if column})
    if missing_columns:
        raise ValueError(
            "В отчёте WB отсутствуют обязательные колонки: "
            + ", ".join(_public_column_name(column) for column in missing_columns)
        )

    report_number: str | None = _valid_report_number(report_number_hint)
    sale_dates: list[date] = []
    parsed_rows: list[WbDailyReportParsedRow] = []
    skipped = 0
    row_index = 0

    for raw_row in rows_iter:
        row_index += 1
        if row_index > max_rows:
            logger.warning("wb_daily_report_row_limit_exceeded", extra={"max_rows": max_rows})
            break
        if raw_row is None or all(cell is None for cell in raw_row):
            continue

        record: dict[str, object] = {}
        raw_columns: dict[str, object] = {}
        for header, column, cell in zip(headers, mapped_columns, raw_row, strict=False):
            if header:
                raw_columns[header] = cell
            if not column:
                continue
            record[column] = cell

        if "report_number" in record and not report_number:
            report_number = _valid_report_number(record.get("report_number"))
        if "sale_dt" in record:
            sale_dt = _coerce_datetime(record.get("sale_dt"))
            if sale_dt is not None:
                sale_dates.append(sale_dt.date())

        if report_number is None and row_index == 1:
            match = REPORT_NUMBER_PATTERN.search(_stringify(raw_row[0]) or "")
            if match:
                report_number = f"{match.group(1)}_{match.group('supplier')}"

        if not record:
            skipped += 1
            continue

        if report_number is not None:
            record["report_number"] = report_number
        record["_raw_columns"] = raw_columns
        record["_headers_count"] = len([header for header in headers if header])

        try:
            parsed = _build_row(record, row_index + 1, report_type=report_type)
        except Exception:
            logger.exception(
                "wb_daily_report_row_parse_failed",
                extra={"row_index": row_index},
            )
            skipped += 1
            continue
        parsed_rows.append(parsed)

    if report_number is None:
        raise ValueError(
            "Не удалось определить номер отчёта WB. "
            "Проверьте, что файл содержит колонку «Номер поставки»."
        )

    return WbDailyReportParsed(
        report_number=report_number,
        report_type=report_type,
        report_date=min(sale_dates) if sale_dates else None,
        report_period_start=min(sale_dates) if sale_dates else None,
        report_period_end=max(sale_dates) if sale_dates else None,
        source_filename=source_filename,
        rows=parsed_rows,
        skipped_rows=skipped,
    )


def parse_wb_daily_report_file(
    uploaded_path: Path,
    *,
    max_bytes: int = 50 * 1024 * 1024,
) -> WbDailyReportParsed:
    """Read an uploaded file, transparently handling ZIP-wrapped XLSX archives."""
    if not uploaded_path.exists():
        raise FileNotFoundError(f"Файл {uploaded_path} не найден")
    raw_bytes = uploaded_path.read_bytes()
    if len(raw_bytes) > max_bytes:
        raise ValueError("Файл превышает допустимый размер 50 МБ")

    return parse_wb_daily_report_upload(
        raw_bytes,
        filename=uploaded_path.name,
        max_bytes=max_bytes,
    )


def parse_wb_daily_report_upload(
    payload: bytes,
    *,
    filename: str,
    max_bytes: int = 50 * 1024 * 1024,
) -> WbDailyReportParsed:
    """Parse uploaded XLSX bytes or a ZIP archive containing one XLSX report."""
    if len(payload) > max_bytes:
        raise ValueError("Файл превышает допустимый размер 50 МБ")

    suffix = Path(filename).suffix.lower()
    report_number_hint = _report_number_from_filename(filename)
    report_type = _report_type_from_filename(filename)

    if suffix == ".zip":
        with zipfile.ZipFile(BytesIO(payload)) as archive:
            xlsx_name = _find_xlsx_in_zip(archive)
            with archive.open(xlsx_name) as xlsx_file:
                xlsx_bytes = xlsx_file.read()
        report_type = report_type or _report_type_from_filename(xlsx_name)
        return iter_wb_daily_report_rows(
            xlsx_bytes,
            report_number_hint=report_number_hint or _report_number_from_filename(xlsx_name),
            report_type=report_type or "daily",
            source_filename=filename,
        )

    if suffix == ".xlsx":
        return iter_wb_daily_report_rows(
            payload,
            report_number_hint=report_number_hint,
            report_type=report_type or "daily",
            source_filename=filename,
        )

    raise ValueError("Поддерживаются только файлы .xlsx или .zip")


def compute_file_hash(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _find_xlsx_in_zip(archive: zipfile.ZipFile) -> str:
    candidates = [
        name
        for name in archive.namelist()
        if name.lower().endswith(".xlsx") and not name.endswith("/")
    ]
    if not candidates:
        raise ValueError("В ZIP-архиве не найден XLSX-файл отчёта WB")
    candidates.sort()
    return candidates[0]


def _report_number_from_filename(filename: str) -> str | None:
    match = REPORT_NUMBER_PATTERN.search(filename)
    if not match:
        return None
    return f"{match.group(1)}_{match.group('supplier')}"


def _valid_report_number(value: object) -> str | None:
    text = _stringify(value)
    if not text or text in {"0", "0.0", "-"}:
        return None
    return text


def _report_type_from_filename(filename: str) -> str | None:
    lowered = filename.lower()
    if "еженедель" in lowered or "weekly" in lowered:
        return "weekly"
    if "ежеднев" in lowered or "daily" in lowered:
        return "daily"
    return None


def _resolve_column(header: str) -> str | None:
    cleaned = header.strip()
    if not cleaned:
        return None
    if cleaned in COLUMN_ALIASES:
        return COLUMN_ALIASES[cleaned]
    lowered = cleaned.lower()
    for source, target in COLUMN_ALIASES.items():
        if source.lower() == lowered:
            return target
    return None


def _build_row(
    record: dict[str, object],
    row_number: int,
    *,
    report_type: str,
) -> WbDailyReportParsedRow:
    sale_dt = _coerce_datetime(record.get("sale_dt")) if "sale_dt" in record else None
    order_dt = _coerce_datetime(record.get("order_dt")) if "order_dt" in record else None
    nm_id = _coerce_int(record.get("nm_id")) if "nm_id" in record else None
    quantity = _coerce_int(record.get("quantity")) if "quantity" in record else None
    payment_reason = _stringify(record.get("payment_reason")) or None
    finance_operation_type, finance_category = classify_payment_reason(payment_reason)
    srid = _stringify(record.get("srid")) or None
    srid_normalized = normalize_srid(srid)

    return WbDailyReportParsedRow(
        row_number=_coerce_int(record.get("row_number"), default=row_number),
        report_type=report_type,
        sale_dt=sale_dt,
        order_dt=order_dt,
        nm_id=nm_id,
        supplier_article=_stringify(record.get("supplier_article")) or None,
        product_name=_stringify(record.get("product_name")) or None,
        size=_stringify(record.get("size")) or None,
        barcode=_stringify(record.get("barcode")) or None,
        shk=_stringify(record.get("shk")) or None,
        srid=srid,
        srid_normalized=srid_normalized,
        rid_normalized=extract_rid_from_srid(srid_normalized),
        doc_type_name=_stringify(record.get("doc_type_name")) or None,
        payment_reason=payment_reason,
        subject_name=_stringify(record.get("subject_name")) or None,
        brand_name=_stringify(record.get("brand_name")) or None,
        quantity=quantity,
        retail_price=_coerce_decimal(record.get("retail_price")),
        retail_amount=_coerce_decimal(record.get("retail_amount")),
        for_pay=_coerce_decimal(record.get("for_pay")),
        delivery_count=_coerce_int(record.get("delivery_count")),
        return_count=_coerce_int(record.get("return_count")),
        delivery_rub=_coerce_decimal(record.get("delivery_rub")),
        penalty=_coerce_decimal(record.get("penalty")),
        storage_fee=_coerce_decimal(record.get("storage_fee")),
        acceptance=_coerce_decimal(record.get("acceptance")),
        deduction=_coerce_decimal(record.get("deduction")),
        commission_rub=_coerce_decimal(record.get("commission_rub")),
        commission_correction_amount=_coerce_decimal(record.get("commission_correction_amount")),
        reimbursement_amount=_coerce_decimal(record.get("reimbursement_amount")),
        logistics_penalty_correction_type=(
            _stringify(record.get("logistics_penalty_correction_type")) or None
        ),
        basket_id=_stringify(record.get("basket_id")) or None,
        sale_method=_stringify(record.get("sale_method")) or None,
        finance_operation_type=finance_operation_type,
        finance_category=finance_category,
        order_required=is_order_required(record, payment_reason),
        raw=record,
    )


def classify_payment_reason(reason: str | None) -> tuple[str, str]:
    text = (reason or "").strip().lower()
    if not text:
        return "unknown", "other"
    if "возврат" in text:
        return "return", "return"
    if "логист" in text or "достав" in text:
        return "expense", "logistics"
    if "хран" in text:
        return "expense", "storage"
    if "прием" in text or "приём" in text:
        return "expense", "paid_acceptance"
    if "штраф" in text:
        return "expense", "penalty"
    if "удерж" in text:
        return "expense", "deduction"
    if "компенсац" in text or "возмещ" in text or "доплат" in text:
        return "income", "compensation"
    if "коррект" in text and ("вознаграж" in text or "вв" in text):
        return "correction", "wb_commission"
    if "продаж" in text or "реализац" in text or "товар" in text:
        return "income", "revenue"
    return "unknown", "other"


def normalize_srid(value: str | None) -> str | None:
    text = (value or "").strip().lower()
    if not text:
        return None
    return re.sub(r"\s+", "", text)


def extract_rid_from_srid(value: str | None) -> str | None:
    text = normalize_srid(value)
    if not text:
        return None
    match = re.search(r"(?:rid|srid)[_:=/-]?([a-z0-9-]+)", text)
    if match:
        return match.group(1)
    if "." in text:
        tail = text.rsplit(".", 1)[-1]
        return tail or None
    return text


def is_order_required(record: dict[str, object], payment_reason: str | None) -> bool:
    for key in ("nm_id", "barcode", "supplier_article", "srid", "shk", "basket_id"):
        if _stringify(record.get(key)):
            return True
    text = (payment_reason or "").lower()
    markers = (
        "продаж",
        "возврат",
        "логист",
        "достав",
        "комисс",
        "вознаграж",
        "перечисл",
        "прием",
        "приём",
        "штраф",
        "удерж",
        "компенсац",
        "возмещ",
        "товар",
    )
    return any(marker in text for marker in markers)


def _coerce_datetime(value: object) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, (int, float)):
        try:
            return from_excel(value)
        except Exception:
            return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d",
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _coerce_int(value: object, *, default: int | None = None) -> int | None:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = str(value).strip()
    if not text:
        return default
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return default


def _coerce_decimal(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("\u00a0", "").replace(" ", "").replace(",", ".")
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _stringify(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _stable_string(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _stable_decimal(value: Decimal | None) -> str | None:
    if value is None:
        return None
    return format(value, "f")


def _normalize_hash_payload(value: object) -> object:
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, dict):
        return {str(key): _normalize_hash_payload(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_normalize_hash_payload(item) for item in value]
    return value


def _public_column_name(column: str) -> str:
    names = {
        "nm_id": "Код номенклатуры",
        "supplier_article": "Артикул поставщика",
        "barcode": "Баркод",
        "payment_reason": "Обоснование для оплаты",
        "doc_type_name": "Тип документа",
        "sale_dt": "Дата продажи",
        "quantity": "Кол-во",
        "retail_amount": "Вайлдберриз реализовал Товар (Пр)",
        "commission_rub": "Вознаграждение Вайлдберриз (ВВ), без НДС",
        "for_pay": "К перечислению Продавцу за реализованный Товар",
        "shk": "ШК",
        "srid": "Srid",
    }
    return names.get(column, column)
