"""version: 2.0.0
description: Web cabinet profit dashboard – KPI, charts, profit tree, pagination, Excel export.
updated: 2026-06-11
"""

from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any

from sqlalchemy import Select, exists, func, literal_column, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.domain import (
    FinancialReportRow,
    Order,
    OrderItem,
    ProfitSnapshot,
    SalesEvent,
    WbDailyReportRow,
    WbReportFinanceComponent,
)
from app.models.enums import (
    CalculationType,
    EconomyConfidence,
    Marketplace,
    ReconciliationStatus,
    SaleModel,
    SourceEventType,
)
from app.services.common.marketplace_presentation import order_status_label
from app.services.common.web_dashboard_service import (
    build_dashboard_filters,
)
from app.utils.datetime import format_datetime_for_user

ZERO = Decimal("0")


@dataclass(slots=True)
class OrderWebFilters:
    period: str
    marketplace: Marketplace | None
    sale_model: SaleModel | None
    local_date_from: date
    local_date_to: date
    date_from: datetime
    date_to: datetime
    economy: str
    status: str
    sku: str
    sort: str
    direction: str


@dataclass(slots=True)
class OrderRow:
    order_id: int
    item_id: int
    order_date: datetime
    marketplace: Marketplace
    sale_model: SaleModel | None
    order_external_id: str
    assembly_id: str | None
    posting_number: str | None
    srid: str | None
    title: str
    seller_article: str
    marketplace_article: str
    quantity: int
    revenue: Decimal
    estimated_profit: Decimal | None
    margin_percent: Decimal | None
    status: str
    source_event_type: str
    requires_action: bool
    missing_cost: bool
    economy_confidence: str
    reconciliation_status: ReconciliationStatus
    # Financial detail fields (estimated/plan values from OrderItem)
    seller_payout_estimated: Decimal | None = None
    commission_estimated: Decimal | None = None
    logistics_estimated: Decimal | None = None
    cost_price_used: Decimal | None = None
    package_cost_used: Decimal | None = None
    other_expenses_estimated: Decimal | None = None


@dataclass(slots=True)
class OrderDetailItem:
    item: OrderItem
    estimated_snapshot: ProfitSnapshot | None
    actual_snapshot: ProfitSnapshot | None
    corrected_actual_profit: Decimal | None = None


@dataclass(slots=True)
class WbFactArticleState:
    key: str
    label: str
    amount: Decimal | None
    state: str


@dataclass(slots=True)
class WbOrderFact:
    status: ReconciliationStatus
    article_states: list[WbFactArticleState]
    linked_rows: list[WbDailyReportRow]
    unlinked_product_rows: list[WbDailyReportRow]


@dataclass(slots=True)
class OzonFactArticle:
    key: str
    label: str
    amount: Decimal | None


@dataclass(slots=True)
class OzonOrderFact:
    status: ReconciliationStatus
    articles: list[OzonFactArticle]
    rows: list[FinancialReportRow]


@dataclass(slots=True)
class OrderDetail:
    order: Order
    items: list[OrderDetailItem]
    estimated_profit: Decimal
    actual_profit: Decimal | None
    deviation: Decimal | None
    reconciliation_status: ReconciliationStatus
    wb_fact: WbOrderFact | None = None
    ozon_fact: OzonOrderFact | None = None
    is_financial_only: bool = False
    has_missing_cost_price: bool = False
    economy_confidence: str = "PRELIMINARY"
    wb_fact_income: Decimal | None = None


@dataclass(slots=True)
class ProfitSkuRow:
    title: str
    seller_article: str
    marketplace: Marketplace
    sale_model: SaleModel | None
    orders: int = 0
    sales: int = 0
    revenue: Decimal = ZERO
    estimated_revenue: Decimal = ZERO
    actual_revenue: Decimal = ZERO
    payout: Decimal = ZERO
    cost: Decimal = ZERO
    marketplace_costs: Decimal = ZERO
    estimated_profit: Decimal = ZERO
    actual_profit: Decimal = ZERO
    margin_percent: Decimal = ZERO
    missing_cost_items: int = 0
    preliminary_items: int = 0
    actual_snapshot_items: int = 0
    roi_percent: Decimal | None = None
    actual_margin: Decimal | None = None
    avg_commission: Decimal = ZERO
    avg_logistics: Decimal = ZERO
    packaging_cost: Decimal = ZERO
    other_costs: Decimal = ZERO
    profit_delta: Decimal = ZERO
    revenue_plan: Decimal = ZERO
    revenue_actual: Decimal = ZERO
    returns_count: int = 0
    warnings: list[str] = field(default_factory=list)
    reconciliation_status: ReconciliationStatus = ReconciliationStatus.PRELIMINARY


@dataclass(slots=True)
class ProfitSummary:
    profit_actual: Decimal = ZERO
    profit_plan: Decimal = ZERO
    deviation: Decimal = ZERO
    deviation_percent: Decimal | None = None
    revenue: Decimal = ZERO
    payout: Decimal = ZERO
    cost_price: Decimal = ZERO
    avg_margin: Decimal | None = None
    roi_percent: Decimal | None = None
    orders_count: int = 0
    sales_count: int = 0
    returns_count: int = 0


@dataclass(slots=True)
class ProfitTreeItem:
    label: str
    amount: Decimal
    tone: str = "neutral"


@dataclass(slots=True)
class ProfitChartData:
    labels: list[str] = field(default_factory=list)
    revenue_values: list[Decimal] = field(default_factory=list)
    payout_values: list[Decimal] = field(default_factory=list)
    profit_values: list[Decimal] = field(default_factory=list)
    expense_labels: list[str] = field(default_factory=list)
    expense_values: list[Decimal] = field(default_factory=list)
    top_sku_labels: list[str] = field(default_factory=list)
    top_sku_values: list[Decimal] = field(default_factory=list)


@dataclass(slots=True)
class ProfitAttentionItem:
    title: str
    description: str
    tone: str
    count: int
    filter_params: dict[str, str]


@dataclass(slots=True)
class OrderPageResult:
    filters: OrderWebFilters
    rows: list[OrderRow]
    total_count: int
    page: int
    per_page: int
    total_pages: int


@dataclass(slots=True)
class SkuDetailData:
    product_name: str
    seller_article: str
    marketplace: Marketplace
    sale_model: SaleModel | None
    orders_count: int
    sales_count: int
    returns_count: int
    revenue: Decimal
    payout: Decimal
    commission: Decimal
    logistics: Decimal
    cost: Decimal
    packaging: Decimal
    other_costs: Decimal
    profit: Decimal
    margin: Decimal | None
    roi: Decimal | None
    warnings: list[str] = field(default_factory=list)


@dataclass(slots=True)
class ProfitPageData:
    filters: OrderWebFilters
    summary: ProfitSummary
    rows: list[ProfitSkuRow]
    page: int = 1
    page_size: int = 50
    total_count: int = 0
    total_pages: int = 1
    chart_data: ProfitChartData = field(default_factory=ProfitChartData)
    profit_tree: list[ProfitTreeItem] = field(default_factory=list)
    attention_items: list[ProfitAttentionItem] = field(default_factory=list)
    has_data: bool = True


@dataclass(slots=True)
class OrderSummaryDTO:
    total_orders: int = 0
    total_items: int = 0
    total_revenue: Decimal = ZERO
    total_estimated_profit: Decimal = ZERO
    total_actual_profit: Decimal | None = None
    average_margin: Decimal | None = None
    missing_cost_count: int = 0
    loss_count: int = 0
    cancelled_count: int = 0
    new_count: int = 0
    delivered_count: int = 0
    return_count: int = 0
    calculation_error_count: int = 0


@dataclass(slots=True)
class OrderPaginationDTO:
    page: int
    per_page: int
    total: int
    total_pages: int
    has_next: bool
    has_prev: bool


class WebOrdersProfitService:
    """Build web order and profit pages from normalized marketplace data."""

    def __init__(self, session: AsyncSession) -> None:
        self.session = session

    async def list_orders(
        self,
        *,
        user_id: int,
        timezone: str,
        period: str,
        marketplace: str | None,
        sale_model: str | None,
        date_from: str | None,
        date_to: str | None,
        economy: str = "all",
        status: str = "all",
        sku: str = "",
        sort: str = "date",
        direction: str = "desc",
        page: int = 1,
        per_page: int = 50,
    ) -> OrderPageResult:
        filters = build_order_web_filters(
            timezone=timezone,
            period=period,
            marketplace=marketplace,
            sale_model=sale_model,
            date_from=date_from,
            date_to=date_to,
            economy=economy,
            status=status,
            sku=sku,
            sort=sort,
            direction=direction,
        )
        base_query = (
            select(
                Order.id,
                OrderItem.id,
                Order.order_date,
                Order.marketplace,
                Order.sale_model,
                Order.order_external_id,
                Order.assembly_id,
                Order.posting_number,
                Order.srid,
                OrderItem.title,
                OrderItem.seller_article,
                OrderItem.marketplace_article,
                OrderItem.quantity,
                (OrderItem.discounted_price * OrderItem.quantity).label("revenue"),
                OrderItem.profit_estimated,
                OrderItem.margin_percent_estimated,
                Order.normalized_status,
                Order.status,
                Order.source_event_type,
                Order.requires_seller_action,
                OrderItem.cost_price_used,
                OrderItem.economy_confidence,
                _order_has_wb_fact_row().label("has_wb_fact"),
                _order_has_wb_missing_fact_state().label("has_wb_missing_fact"),
                _order_has_wb_match_problem().label("has_wb_match_problem"),
                OrderItem.seller_payout_estimated,
                OrderItem.commission_estimated,
                OrderItem.logistics_estimated,
                OrderItem.package_cost_used,
                OrderItem.other_marketplace_expenses_estimated,
            )
            .join(OrderItem, OrderItem.order_id == Order.id)
            .where(Order.user_id == user_id)
            .where(Order.deleted_at.is_(None))
            .where(Order.order_date >= filters.date_from)
            .where(Order.order_date <= filters.date_to)
        )
        base_query = _apply_order_page_filters(base_query, filters)

        count_query = (
            select(func.count(func.distinct(Order.id)))
            .select_from(Order)
            .join(OrderItem, OrderItem.order_id == Order.id)
            .where(Order.user_id == user_id)
            .where(Order.deleted_at.is_(None))
            .where(Order.order_date >= filters.date_from)
            .where(Order.order_date <= filters.date_to)
        )
        count_query = _apply_order_page_filters(count_query, filters, count_distinct=True)
        count_result = await self.session.execute(count_query)
        total_count = int(count_result.scalar() or 0)

        per_page = max(1, min(per_page, 200))
        page = max(1, page)
        total_pages = max(1, (total_count + per_page - 1) // per_page)
        offset = (page - 1) * per_page

        query = (
            base_query.order_by(*_apply_order_sort_query(base_query, filters))
            .limit(per_page)
            .offset(offset)
        )
        result = await self.session.execute(query)
        rows = []
        for row in result.all():
            (
                order_id,
                item_id,
                order_date,
                marketplace_value,
                sale_model_value,
                order_external_id,
                assembly_id,
                posting_number,
                srid,
                title,
                seller_article,
                marketplace_article,
                quantity,
                revenue,
                estimated_profit,
                margin_percent,
                normalized_status,
                raw_status,
                source_event_type,
                requires_action,
                cost_price_used,
                economy_confidence,
                has_wb_fact,
                has_wb_missing_fact,
                has_wb_match_problem,
                seller_payout_est,
                commission_est,
                logistics_est,
                package_cost,
                other_expenses_est,
            ) = row
            missing_cost = cost_price_used is None
            rows.append(
                OrderRow(
                    order_id=int(order_id),
                    item_id=int(item_id),
                    order_date=order_date,
                    marketplace=marketplace_value,
                    sale_model=sale_model_value,
                    order_external_id=str(order_external_id),
                    assembly_id=str(assembly_id) if assembly_id else None,
                    posting_number=posting_number,
                    srid=str(srid) if srid else None,
                    title=title or "Без названия",
                    seller_article=seller_article or "н/д",
                    marketplace_article=marketplace_article or "н/д",
                    quantity=int(quantity or 0),
                    revenue=_decimal(revenue),
                    estimated_profit=(
                        _decimal(estimated_profit) if estimated_profit is not None else None
                    ),
                    margin_percent=(
                        _decimal(margin_percent) if margin_percent is not None else None
                    ),
                    status=normalized_status or raw_status or "н/д",
                    source_event_type=(
                        source_event_type.value if source_event_type is not None else "н/д"
                    ),
                    requires_action=bool(requires_action),
                    missing_cost=missing_cost,
                    economy_confidence=str(
                        economy_confidence or EconomyConfidence.PRELIMINARY.value
                    ),
                    reconciliation_status=_reconciliation_status(
                        marketplace=marketplace_value,
                        missing_cost=missing_cost,
                        has_wb_fact=bool(has_wb_fact),
                        has_wb_missing_fact=bool(has_wb_missing_fact),
                        has_wb_match_problem=bool(has_wb_match_problem),
                    ),
                    seller_payout_estimated=_decimal(seller_payout_est) if seller_payout_est is not None else None,
                    commission_estimated=_decimal(commission_est) if commission_est is not None else None,
                    logistics_estimated=_decimal(logistics_est) if logistics_est is not None else None,
                    cost_price_used=_decimal(cost_price_used) if cost_price_used is not None else None,
                    package_cost_used=_decimal(package_cost) if package_cost is not None else None,
                    other_expenses_estimated=_decimal(other_expenses_est) if other_expenses_est is not None else None,
                )
            )
        pagination = OrderPaginationDTO(
            page=page,
            per_page=per_page,
            total=total_count,
            total_pages=total_pages,
            has_next=page < total_pages,
            has_prev=page > 1,
        )
        return OrderPageResult(
            filters=filters,
            rows=rows,
            total_count=total_count,
            page=page,
            per_page=per_page,
            total_pages=total_pages,
        )

    async def orders_summary(
        self,
        *,
        user_id: int,
        timezone: str,
        period: str,
        marketplace: str | None,
        sale_model: str | None,
        date_from: str | None,
        date_to: str | None,
        economy: str = "all",
        status: str = "all",
        sku: str = "",
    ) -> OrderSummaryDTO:
        filters = build_order_web_filters(
            timezone=timezone,
            period=period,
            marketplace=marketplace,
            sale_model=sale_model,
            date_from=date_from,
            date_to=date_to,
            economy=economy,
            status=status,
            sku=sku,
            sort="date",
            direction="desc",
        )
        # Total counts from orders
        count_query = (
            select(
                func.count(func.distinct(Order.id)),
                func.count(func.distinct(Order.id)).filter(
                    func.lower(func.coalesce(Order.normalized_status, Order.status, "")).in_(
                        ("cancelled", "canceled", "cancel")
                    )
                ),
            )
            .select_from(Order)
            .join(OrderItem, OrderItem.order_id == Order.id)
            .where(Order.user_id == user_id)
            .where(Order.deleted_at.is_(None))
            .where(Order.order_date >= filters.date_from)
            .where(Order.order_date <= filters.date_to)
        )
        count_query = _apply_order_page_filters(count_query, filters, count_distinct=True)
        count_result = await self.session.execute(count_query)
        total_orders, cancelled_count = count_result.one()
        total_orders = int(total_orders or 0)
        cancelled_count = int(cancelled_count or 0)

        # Aggregated financial data
        agg_query = (
            select(
                func.coalesce(func.sum(OrderItem.quantity), 0),
                func.coalesce(func.sum(OrderItem.discounted_price * OrderItem.quantity), 0),
                func.coalesce(func.sum(OrderItem.profit_estimated), 0),
                func.coalesce(func.avg(OrderItem.margin_percent_estimated), 0),
                func.count(OrderItem.id).filter(OrderItem.cost_price_used.is_(None)),
                func.count(OrderItem.id).filter(OrderItem.profit_estimated < 0),
            )
            .select_from(Order)
            .join(OrderItem, OrderItem.order_id == Order.id)
            .where(Order.user_id == user_id)
            .where(Order.deleted_at.is_(None))
            .where(Order.order_date >= filters.date_from)
            .where(Order.order_date <= filters.date_to)
        )
        agg_query = _apply_order_page_filters(agg_query, filters)
        agg_result = await self.session.execute(agg_query)
        total_qty, total_revenue, total_profit, avg_margin, missing_cost, loss_count = agg_result.one()

        # Actual profit via snapshots
        latest_actual = (
            select(
                ProfitSnapshot.order_item_id,
                ProfitSnapshot.profit,
                func.row_number()
                .over(
                    partition_by=ProfitSnapshot.order_item_id,
                    order_by=(ProfitSnapshot.calculated_at.desc(), ProfitSnapshot.id.desc()),
                )
                .label("rn"),
            )
            .where(ProfitSnapshot.calculation_type == CalculationType.ACTUAL)
            .subquery()
        )
        actual_query = (
            select(func.coalesce(func.sum(latest_actual.c.profit), 0))
            .select_from(Order)
            .join(OrderItem, OrderItem.order_id == Order.id)
            .outerjoin(
                latest_actual,
                (latest_actual.c.order_item_id == OrderItem.id) & (latest_actual.c.rn == 1),
            )
            .where(Order.user_id == user_id)
            .where(Order.deleted_at.is_(None))
            .where(Order.order_date >= filters.date_from)
            .where(Order.order_date <= filters.date_to)
        )
        actual_query = _apply_order_page_filters(actual_query, filters)
        actual_result = await self.session.execute(actual_query)
        total_actual_profit = Decimal(str(actual_result.scalar() or 0))

        margin = None
        if total_revenue and Decimal(str(total_revenue)) > 0:
            margin = (Decimal(str(total_profit)) / Decimal(str(total_revenue)) * Decimal("100")).quantize(Decimal("0.1"))

        return OrderSummaryDTO(
            total_orders=total_orders,
            total_items=int(total_qty or 0),
            total_revenue=Decimal(str(total_revenue or 0)),
            total_estimated_profit=Decimal(str(total_profit or 0)),
            total_actual_profit=Decimal(str(total_actual_profit)) if total_actual_profit else None,
            average_margin=margin,
            missing_cost_count=int(missing_cost or 0),
            loss_count=int(loss_count or 0),
            cancelled_count=cancelled_count,
        )

    async def order_financial_rows(
        self,
        *,
        user_id: int,
        order_id: int,
    ) -> list[FinancialReportRow]:
        order_result = await self.session.execute(
            select(Order).where(Order.id == order_id, Order.user_id == user_id)
        )
        order = order_result.scalar_one_or_none()
        if order is None:
            return []
        if order.marketplace == Marketplace.OZON:
            result = await self.session.execute(
                select(FinancialReportRow)
                .where(
                    FinancialReportRow.marketplace_account_id == order.marketplace_account_id,
                    FinancialReportRow.marketplace == Marketplace.OZON,
                    FinancialReportRow.order_external_id == order.order_external_id,
                )
                .order_by(FinancialReportRow.operation_date.desc())
            )
            return list(result.scalars().all())
        return []

    async def order_detail(self, *, user_id: int, order_id: int) -> OrderDetail | None:
        result = await self.session.execute(
            select(Order)
            .options(selectinload(Order.items).selectinload(OrderItem.snapshots))
            .where(Order.user_id == user_id)
            .where(Order.id == order_id)
            .where(Order.deleted_at.is_(None))
        )
        order = result.scalar_one_or_none()
        if order is None:
            return None
        items = []
        estimated_total = ZERO
        actual_total = ZERO
        has_actual = False
        for item in order.items:
            estimated_snapshot = _latest_snapshot(item.snapshots, CalculationType.ESTIMATED)
            actual_snapshot = _latest_snapshot(item.snapshots, CalculationType.ACTUAL)
            if estimated_snapshot is not None:
                estimated_total += estimated_snapshot.profit
            elif item.profit_estimated is not None:
                estimated_total += item.profit_estimated
            if actual_snapshot is not None:
                actual_total += actual_snapshot.profit
                has_actual = True
            items.append(
                OrderDetailItem(
                    item=item,
                    estimated_snapshot=estimated_snapshot,
                    actual_snapshot=actual_snapshot,
                )
            )
        actual_profit = actual_total if has_actual else None
        deviation = actual_total - estimated_total if has_actual else None
        wb_fact = await self._wb_order_fact(order) if order.marketplace == Marketplace.WB else None
        ozon_fact = await self._ozon_order_fact(order) if order.marketplace == Marketplace.OZON else None
        missing_cost = any(item.cost_price_used is None for item in order.items)

        # Compute wb_fact_income from linked + unlinked WbDailyReportRow (filtered by srid)
        wb_fact_income: Decimal | None = None
        corrected_actual_profit: Decimal | None = None
        if wb_fact is not None and (wb_fact.linked_rows or wb_fact.unlinked_product_rows):
            all_relevant = list(wb_fact.linked_rows)
            # Include unlinked product rows that share the same srid (logistics, acceptance, etc.)
            if order.srid:
                srid_filtered = [r for r in wb_fact.unlinked_product_rows if r.srid == order.srid]
                all_relevant.extend(srid_filtered)
            else:
                all_relevant.extend(wb_fact.unlinked_product_rows)
            total_for_pay = sum((r.for_pay or ZERO) for r in all_relevant)
            total_delivery = sum((r.delivery_rub or ZERO) for r in all_relevant if r.for_pay is None)
            total_storage = sum((r.storage_fee or ZERO) for r in all_relevant if r.for_pay is None)
            total_acceptance = sum((r.acceptance or ZERO) for r in all_relevant if r.for_pay is None)
            total_penalty = sum((r.penalty or ZERO) for r in all_relevant if r.for_pay is None)
            total_deduction = sum((r.deduction or ZERO) for r in all_relevant if r.for_pay is None)
            total_reimbursement = sum((r.reimbursement_amount or ZERO) for r in all_relevant if r.for_pay is None)
            wb_fact_income = (
                total_for_pay
                - total_delivery - total_storage - total_acceptance
                - total_penalty - total_deduction
                + total_reimbursement
            )
            # Recompute actual profit from source data: wb_fact_income - seller costs
            if has_actual:
                seller_cost_price = sum(
                    (_latest_snapshot(item.snapshots, CalculationType.ACTUAL).cost_price or ZERO)
                    for item in order.items
                    if _latest_snapshot(item.snapshots, CalculationType.ACTUAL) is not None
                )
                seller_package_cost = sum(
                    (_latest_snapshot(item.snapshots, CalculationType.ACTUAL).package_cost or ZERO)
                    for item in order.items
                    if _latest_snapshot(item.snapshots, CalculationType.ACTUAL) is not None
                )
                seller_additional_cost = sum(
                    (_latest_snapshot(item.snapshots, CalculationType.ACTUAL).additional_seller_cost or ZERO)
                    for item in order.items
                    if _latest_snapshot(item.snapshots, CalculationType.ACTUAL) is not None
                )
                seller_tax = sum(
                    (_latest_snapshot(item.snapshots, CalculationType.ACTUAL).tax_amount or ZERO)
                    for item in order.items
                    if _latest_snapshot(item.snapshots, CalculationType.ACTUAL) is not None
                )
                corrected_actual_profit = wb_fact_income - seller_cost_price - seller_package_cost - seller_additional_cost - seller_tax
        reconciliation_status = (
            ReconciliationStatus.MANUAL_REVIEW
            if missing_cost
            else wb_fact.status
            if wb_fact is not None
            else ozon_fact.status
            if ozon_fact is not None
            else ReconciliationStatus.PRELIMINARY
        )
        is_financial_only = (
            order.marketplace == Marketplace.WB
            and order.source_event_type == SourceEventType.REPORT_ORDER
            and order.srid is not None
            and _is_short_srid(order.srid)
        )
        order_confidence = "PRELIMINARY"
        conf_priority = {"EXACT": 0, "ESTIMATED": 1, "PRELIMINARY": 2}
        for item in order.items:
            ic = str(item.economy_confidence or "PRELIMINARY")
            if conf_priority.get(ic, 2) > conf_priority.get(order_confidence, 2):
                order_confidence = ic
        order_confidence = order_confidence or "PRELIMINARY"

        # Override actual_profit and deviation with corrected values
        if corrected_actual_profit is not None:
            actual_profit = corrected_actual_profit
            deviation = corrected_actual_profit - estimated_total
            # For single-item orders, propagate corrected profit to item level
            if len(items) == 1:
                items[0].corrected_actual_profit = corrected_actual_profit

        return OrderDetail(
            order=order,
            items=items,
            estimated_profit=estimated_total,
            actual_profit=actual_profit,
            deviation=deviation,
            reconciliation_status=reconciliation_status,
            wb_fact=wb_fact,
            ozon_fact=ozon_fact,
            is_financial_only=is_financial_only,
            has_missing_cost_price=missing_cost,
            economy_confidence=order_confidence,
            wb_fact_income=wb_fact_income,
        )

    async def _wb_order_fact(self, order: Order) -> WbOrderFact:
        linked_rows = await self._wb_rows_linked_to_order(order)
        unlinked_rows = await self._wb_rows_unlinked_for_order_products(order)
        has_report_near_order = bool(linked_rows or unlinked_rows)
        if not has_report_near_order:
            has_report_near_order = await self._has_wb_report_rows_near_order(order)
        article_states = _wb_fact_article_states(
            linked_rows=linked_rows,
            unlinked_rows=unlinked_rows,
            has_report_near_order=has_report_near_order,
        )
        if any(
            row.order_match_status in {"ambiguous", "ambiguous_order_match", "error"}
            for row in linked_rows + unlinked_rows
        ):
            status = ReconciliationStatus.FACT_AMBIGUOUS
        elif linked_rows and all(state.state == "present" for state in article_states):
            status = ReconciliationStatus.FACT_MATCHED
        elif linked_rows:
            status = ReconciliationStatus.FACT_PARTIAL
        elif unlinked_rows:
            status = ReconciliationStatus.FACT_UNMATCHED
        elif has_report_near_order:
            status = ReconciliationStatus.FACT_UNMATCHED
        else:
            status = ReconciliationStatus.PRELIMINARY
        return WbOrderFact(
            status=status,
            article_states=article_states,
            linked_rows=linked_rows,
            unlinked_product_rows=unlinked_rows,
        )

    async def _ozon_order_fact(self, order: Order) -> OzonOrderFact | None:
        if not order.order_external_id:
            return None
        result = await self.session.execute(
            select(FinancialReportRow)
            .where(
                FinancialReportRow.marketplace_account_id == order.marketplace_account_id,
                FinancialReportRow.marketplace == Marketplace.OZON,
                FinancialReportRow.order_external_id == order.order_external_id,
            )
            .order_by(FinancialReportRow.operation_category, FinancialReportRow.operation_date)
        )
        rows = list(result.scalars().all())
        if not rows:
            return None

        articles: list[OzonFactArticle] = []
        seen_categories: set[str] = set()
        for row in rows:
            cat = (row.operation_category or "other").lower()
            if cat not in seen_categories:
                seen_categories.add(cat)
                label = _ozon_category_label(cat)
                articles.append(OzonFactArticle(key=cat, label=label, amount=row.amount))
            else:
                existing = next((a for a in articles if a.key == cat), None)
                if existing is not None and row.amount is not None:
                    existing.amount = (existing.amount or ZERO) + row.amount

        has_actual = any(
            s.calculation_type == CalculationType.ACTUAL
            for item in getattr(order, "items", [])
            for s in getattr(item, "snapshots", [])
        )
        if has_actual:
            status = ReconciliationStatus.FACT_MATCHED
        elif rows:
            status = ReconciliationStatus.FACT_PARTIAL
        else:
            status = ReconciliationStatus.PRELIMINARY

        return OzonOrderFact(status=status, articles=articles, rows=rows)

    async def _wb_rows_linked_to_order(self, order: Order) -> list[WbDailyReportRow]:
        result = await self.session.execute(
            select(WbDailyReportRow)
            .where(WbDailyReportRow.user_id == order.user_id)
            .where(WbDailyReportRow.marketplace_account_id == order.marketplace_account_id)
            .where(WbDailyReportRow.linked_order_id == order.id)
            .where(WbDailyReportRow.is_active.is_(True))
            .where(WbDailyReportRow.deleted_at.is_(None))
            .order_by(WbDailyReportRow.sale_dt.desc().nullslast(), WbDailyReportRow.id.desc())
            .limit(200)
        )
        return list(result.scalars().all())

    async def _wb_rows_unlinked_for_order_products(self, order: Order) -> list[WbDailyReportRow]:
        product_ids = {item.product_id for item in order.items if item.product_id is not None}
        if not product_ids:
            return []
        date_from = order.order_date - timedelta(days=14)
        date_to = order.order_date + timedelta(days=14)
        result = await self.session.execute(
            select(WbDailyReportRow)
            .where(WbDailyReportRow.user_id == order.user_id)
            .where(WbDailyReportRow.marketplace_account_id == order.marketplace_account_id)
            .where(WbDailyReportRow.linked_product_id.in_(product_ids))
            .where(WbDailyReportRow.linked_order_id.is_(None))
            .where(WbDailyReportRow.sale_dt >= date_from)
            .where(WbDailyReportRow.sale_dt <= date_to)
            .where(WbDailyReportRow.is_active.is_(True))
            .where(WbDailyReportRow.deleted_at.is_(None))
            .order_by(WbDailyReportRow.sale_dt.desc().nullslast(), WbDailyReportRow.id.desc())
            .limit(200)
        )
        return list(result.scalars().all())

    async def _has_wb_report_rows_near_order(self, order: Order) -> bool:
        date_from = order.order_date - timedelta(days=14)
        date_to = order.order_date + timedelta(days=14)
        result = await self.session.execute(
            select(WbDailyReportRow.id)
            .where(WbDailyReportRow.user_id == order.user_id)
            .where(WbDailyReportRow.marketplace_account_id == order.marketplace_account_id)
            .where(WbDailyReportRow.sale_dt >= date_from)
            .where(WbDailyReportRow.sale_dt <= date_to)
            .where(WbDailyReportRow.is_active.is_(True))
            .where(WbDailyReportRow.deleted_at.is_(None))
            .limit(1)
        )
        return result.scalar_one_or_none() is not None

    async def profit_by_sku(
        self,
        *,
        user_id: int,
        timezone: str,
        period: str,
        marketplace: str | None,
        sale_model: str | None,
        date_from: str | None,
        date_to: str | None,
        economy: str = "all",
        status: str = "all",
        sku: str = "",
        sort: str = "profit",
        direction: str = "desc",
        page: int = 1,
        page_size: int = 50,
    ) -> ProfitPageData:
        filters = build_order_web_filters(
            timezone=timezone,
            period=period,
            marketplace=marketplace,
            sale_model=sale_model,
            date_from=date_from,
            date_to=date_to,
            economy=economy,
            status=status,
            sku=sku,
            sort=sort,
            direction=direction,
        )
        order_rows = await self._profit_order_rows(user_id, filters)
        sales_map = await self._sales_by_sku(user_id, filters)
        rows: list[ProfitSkuRow] = []
        for row in order_rows:
            (
                title,
                seller_article,
                marketplace_value,
                sale_model_value,
                orders,
                revenue,
                quantity,
                cost,
                marketplace_costs,
                estimated_profit,
                actual_profit,
                margin,
                avg_commission,
                avg_logistics,
                missing_cost_items,
                preliminary_items,
                actual_snapshot_items,
            ) = row
            key = (marketplace_value, seller_article or "")
            sales = sales_map.get(key, 0)
            profit = _decimal(estimated_profit)
            total_cost = _decimal(cost)
            actual_profit_val = _decimal(actual_profit)
            missing = int(missing_cost_items or 0)
            has_actual = int(actual_snapshot_items or 0)
            has_preliminary = int(preliminary_items or 0)
            if missing > 0:
                reconcil_status = ReconciliationStatus.MANUAL_REVIEW
            elif has_actual > 0 and has_preliminary == 0:
                reconcil_status = ReconciliationStatus.FACT_MATCHED
            elif has_actual > 0:
                reconcil_status = ReconciliationStatus.FACT_PARTIAL
            else:
                reconcil_status = ReconciliationStatus.PRELIMINARY
            rows.append(
                ProfitSkuRow(
                    title=title or seller_article or "Без названия",
                    seller_article=seller_article or "н/д",
                    marketplace=marketplace_value,
                    sale_model=sale_model_value,
                    orders=int(orders or 0),
                    sales=sales,
                    revenue=_decimal(revenue),
                    estimated_revenue=_decimal(revenue),
                    actual_revenue=ZERO,
                    payout=ZERO,
                    cost=total_cost,
                    marketplace_costs=_decimal(marketplace_costs),
                    estimated_profit=profit,
                    actual_profit=actual_profit_val,
                    margin_percent=_decimal(margin),
                    actual_margin=None,
                    avg_commission=_decimal(avg_commission),
                    avg_logistics=_decimal(avg_logistics),
                    roi_percent=roi_percent(profit, total_cost),
                    missing_cost_items=missing,
                    preliminary_items=has_preliminary,
                    actual_snapshot_items=has_actual,
                    reconciliation_status=reconcil_status,
                )
            )
        rows = _filter_profit_rows(rows, filters.economy)
        rows = await self._merge_wb_daily_report_rows(user_id, filters, rows)
        for r in rows:
            if r.actual_revenue and r.actual_revenue > ZERO and r.actual_profit is not None:
                r.actual_margin = (
                    (r.actual_profit / r.actual_revenue * Decimal("100")).quantize(Decimal("0.1"))
                )
            r.profit_delta = r.actual_profit - r.estimated_profit
            r.revenue_plan = r.estimated_revenue
            r.revenue_actual = r.actual_revenue
        rows = _sort_profit_rows(rows, filters.sort, filters.direction)
        total_count = len(rows)
        total_pages = max(1, (total_count + page_size - 1) // page_size)
        page = max(1, min(page, total_pages))
        start = (page - 1) * page_size
        paged_rows = rows[start:start + page_size]
        summary = _profit_summary(rows)
        chart_data = _build_chart_data(rows)
        profit_tree = _build_profit_tree(summary)
        attention_items = _build_attention_items(rows, user_id)
        return ProfitPageData(
            filters=filters,
            summary=summary,
            rows=paged_rows,
            page=page,
            page_size=page_size,
            total_count=total_count,
            total_pages=total_pages,
            chart_data=chart_data,
            profit_tree=profit_tree,
            attention_items=attention_items,
            has_data=bool(rows),
        )

    async def sku_detail(
        self,
        *,
        user_id: int,
        seller_article: str,
        marketplace: str,
        timezone: str,
        period: str,
        date_from: str | None = None,
        date_to: str | None = None,
    ) -> SkuDetailData | None:
        filters = build_order_web_filters(
            timezone=timezone,
            period=period,
            marketplace=marketplace,
            sale_model="all",
            date_from=date_from,
            date_to=date_to,
            economy="all",
            status="all",
            sku=seller_article,
            sort="profit",
            direction="desc",
        )
        order_rows = await self._profit_order_rows(user_id, filters)
        if not order_rows:
            return None
        row = order_rows[0]
        (
            title, sa, mp, sm, orders, rev, qty, cost, mp_costs,
            est_profit, act_profit, margin, comm, logi,
            miss_cost, prelim, act_snap,
        ) = row
        sales_map = await self._sales_by_sku(user_id, filters)
        key = (mp, seller_article)
        sales = sales_map.get(key, 0)
        returns_count = await self._returns_count(user_id, filters, seller_article)
        profit_val = _decimal(act_profit)
        cost_val = _decimal(cost)
        warnings_list = []
        if miss_cost:
            warnings_list.append("Нет себестоимости")
        if comm is None or _decimal(comm) == ZERO:
            warnings_list.append("Комиссия не учтена")
        if logi is None or _decimal(logi) == ZERO:
            warnings_list.append("Логистика не учтена")
        if profit_val < ZERO:
            warnings_list.append("Убыточная позиция")
        return SkuDetailData(
            product_name=title or sa or "Без названия",
            seller_article=sa or "н/д",
            marketplace=mp,
            sale_model=sm,
            orders_count=int(orders or 0),
            sales_count=sales,
            returns_count=returns_count,
            revenue=_decimal(rev),
            payout=_decimal(rev) - _decimal(mp_costs),
            commission=_decimal(comm),
            logistics=_decimal(logi),
            cost=cost_val,
            packaging=ZERO,
            other_costs=ZERO,
            profit=profit_val,
            margin=margin,
            roi=roi_percent(profit_val, cost_val),
            warnings=warnings_list,
        )

    async def export_profit_excel(
        self,
        *,
        user_id: int,
        timezone: str,
        period: str,
        marketplace: str | None,
        sale_model: str | None,
        date_from: str | None,
        date_to: str | None,
        economy: str = "all",
        status: str = "all",
        sku: str = "",
        sort: str = "profit",
        direction: str = "desc",
    ) -> bytes:
        data = await self.profit_by_sku(
            user_id=user_id,
            timezone=timezone,
            period=period,
            marketplace=marketplace,
            sale_model=sale_model,
            date_from=date_from,
            date_to=date_to,
            economy=economy,
            status=status,
            sku=sku,
            sort=sort,
            direction=direction,
            page=1,
            page_size=10000,
        )
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        except ImportError:
            logger.warning("openpyxl not installed, returning CSV-like bytes")
            lines = ["Товар;Артикул;МП;Модель;Заказы;Продажи;Выручка план;Выручка факт;К перечислению;Себестоимость;Комиссия;Логистика;Прибыль план;Прибыль факт;Маржа;ROI;Статус"]
            for r in data.rows:
                lines.append(f"{r.title};{r.seller_article};{r.marketplace.value};{r.sale_model.value if r.sale_model else ''};{r.orders};{r.sales};{r.estimated_revenue};{r.actual_revenue};{r.payout};{r.cost};{r.avg_commission};{r.avg_logistics};{r.estimated_profit};{r.actual_profit};{r.actual_margin or 0};{r.roi_percent or ''};{r.reconciliation_status.value}")
            return ("\n".join(lines)).encode("utf-8-sig")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Прибыль"
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="1D4ED8", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        headers = [
            "Товар", "Артикул", "МП", "Модель",
            "Заказы", "Продажи",
            "Выручка план", "Выручка факт", "К перечислению",
            "Себестоимость", "Комиссия", "Логистика",
            "Прибыль план", "Прибыль факт", "Маржа %", "ROI %",
            "Статус сверки",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        for i, r in enumerate(data.rows, 2):
            values = [
                r.title, r.seller_article, r.marketplace.value,
                r.sale_model.value if r.sale_model else "",
                r.orders, r.sales,
                float(r.estimated_revenue), float(r.actual_revenue), float(r.payout),
                float(r.cost), float(r.avg_commission), float(r.avg_logistics),
                float(r.estimated_profit), float(r.actual_profit),
                float(r.actual_margin or 0), float(r.roi_percent or 0),
                r.reconciliation_status.value,
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = thin_border
                if col >= 7:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '#,##0.00'
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 20
        for col_letter in ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q"]:
            ws.column_dimensions[col_letter].width = 15
        import io
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    async def _returns_count(self, user_id: int, filters: OrderWebFilters, seller_article: str) -> int:
        from app.models.orders import ReturnsEvent
        query = select(func.coalesce(func.sum(ReturnsEvent.quantity), 0)).where(
            ReturnsEvent.user_id == user_id,
            ReturnsEvent.event_date >= filters.date_from,
            ReturnsEvent.event_date <= filters.date_to,
            ReturnsEvent.seller_article == seller_article,
        )
        if filters.marketplace is not None:
            query = query.where(ReturnsEvent.marketplace == filters.marketplace)
        result = await self.session.execute(query)
        return int(result.scalar() or 0)

    async def _merge_wb_daily_report_rows(
        self,
        user_id: int,
        filters: OrderWebFilters,
        rows: list[ProfitSkuRow],
    ) -> list[ProfitSkuRow]:
        if filters.marketplace not in (None, Marketplace.WB):
            return rows
        article_expr = func.coalesce(WbDailyReportRow.supplier_article, literal_column("''"))
        operation_text = func.lower(
            func.concat(
                func.coalesce(WbDailyReportRow.doc_type_name, ""),
                " ",
                func.coalesce(WbDailyReportRow.payment_reason, ""),
            )
        )
        sales_case = operation_text.not_like("%возврат%")
        rows_query = (
            select(
                article_expr.label("seller_article"),
                func.coalesce(func.sum(WbDailyReportRow.quantity).filter(sales_case), 0),
            )
            .where(WbDailyReportRow.user_id == user_id)
            .where(WbDailyReportRow.sale_dt >= filters.date_from)
            .where(WbDailyReportRow.sale_dt <= filters.date_to)
            .where(WbDailyReportRow.is_active.is_(True))
            .where(WbDailyReportRow.deleted_at.is_(None))
            .where(WbDailyReportRow.operation_scope.in_(("order", "product")))
            .group_by(article_expr)
        )
        if filters.sku:
            rows_query = rows_query.where(
                WbDailyReportRow.supplier_article.ilike(f"%{filters.sku}%")
            )
        rows_result = await self.session.execute(rows_query)
        sales_by_article = {
            str(seller_article or ""): int(sales or 0)
            for seller_article, sales in rows_result.all()
        }

        component_article_expr = func.coalesce(
            WbDailyReportRow.supplier_article,
            literal_column("''"),
        )
        query = (
            select(
                component_article_expr.label("seller_article"),
                func.coalesce(
                    func.sum(WbReportFinanceComponent.normalized_amount).filter(
                        WbReportFinanceComponent.finance_category.in_(("sale", "return"))
                    ),
                    0,
                ),
                func.coalesce(
                    func.sum(WbReportFinanceComponent.normalized_amount).filter(
                        WbReportFinanceComponent.finance_category == "payout"
                    ),
                    0,
                ),
                func.coalesce(
                    func.sum(WbReportFinanceComponent.normalized_amount).filter(
                        WbReportFinanceComponent.finance_category.in_(
                            (
                                "commission",
                                "logistics",
                                "storage",
                                "penalty",
                                "deduction",
                                "acceptance",
                            )
                        )
                    ),
                    0,
                ),
            )
            .join(
                WbDailyReportRow,
                WbDailyReportRow.id == WbReportFinanceComponent.report_row_id,
            )
            .where(WbDailyReportRow.user_id == user_id)
            .where(WbDailyReportRow.sale_dt >= filters.date_from)
            .where(WbDailyReportRow.sale_dt <= filters.date_to)
            .where(WbDailyReportRow.is_active.is_(True))
            .where(WbDailyReportRow.deleted_at.is_(None))
            .where(WbReportFinanceComponent.is_active.is_(True))
            .where(WbReportFinanceComponent.deleted_at.is_(None))
            .where(WbReportFinanceComponent.operation_scope.in_(("order", "product")))
            .group_by(component_article_expr)
        )
        if filters.sku:
            query = query.where(WbDailyReportRow.supplier_article.ilike(f"%{filters.sku}%"))
        result = await self.session.execute(query)
        by_key = {(row.marketplace, row.seller_article): row for row in rows}
        for (
            seller_article,
            revenue,
            payout,
            marketplace_costs,
        ) in result.all():
            article = str(seller_article or "")
            actual_profit = _decimal(payout)
            key = (Marketplace.WB, article)
            existing = by_key.get(key)
            if existing is not None:
                existing.sales += sales_by_article.get(article, 0)
                existing.actual_revenue += _decimal(revenue)
                existing.payout += _decimal(payout)
                existing.marketplace_costs += _decimal(marketplace_costs)
                if existing.actual_snapshot_items == 0:
                    existing.actual_profit += actual_profit - existing.cost
                if (
                    existing.reconciliation_status == ReconciliationStatus.PRELIMINARY
                    and sales_by_article.get(article, 0)
                ):
                    existing.reconciliation_status = ReconciliationStatus.FACT_MATCHED
                continue
            wb_payout = _decimal(payout)
            wb_revenue = _decimal(revenue)
            wb_costs = _decimal(marketplace_costs)
            wb_profit = wb_payout
            rows.append(
                ProfitSkuRow(
                    title=article or "WB daily report",
                    seller_article=article or "н/д",
                    marketplace=Marketplace.WB,
                    sale_model=None,
                    orders=0,
                    sales=sales_by_article.get(article, 0),
                    revenue=ZERO,
                    estimated_revenue=ZERO,
                    actual_revenue=wb_revenue,
                    payout=wb_payout,
                    cost=ZERO,
                    marketplace_costs=wb_costs,
                    estimated_profit=ZERO,
                    actual_profit=wb_profit,
                    margin_percent=ZERO,
                    actual_margin=None,
                    avg_commission=ZERO,
                    avg_logistics=ZERO,
                    roi_percent=None,
                    missing_cost_items=0,
                    preliminary_items=0,
                    actual_snapshot_items=0,
                    reconciliation_status=ReconciliationStatus.FACT_MATCHED,
                )
            )
        return rows

    async def _profit_order_rows(
        self,
        user_id: int,
        filters: OrderWebFilters,
    ) -> list[Any]:
        query = self._profit_order_query(user_id, filters)
        result = await self.session.execute(query)
        return list(result.all())

    @staticmethod
    def _profit_order_query(user_id: int, filters: OrderWebFilters) -> Select[Any]:
        title_expr = func.coalesce(
            OrderItem.title,
            OrderItem.seller_article,
            literal_column("'Без названия'"),
        )
        article_expr = func.coalesce(OrderItem.seller_article, literal_column("''"))
        latest_actual = (
            select(
                ProfitSnapshot.order_item_id.label("order_item_id"),
                ProfitSnapshot.profit.label("profit"),
                func.row_number()
                .over(
                    partition_by=ProfitSnapshot.order_item_id,
                    order_by=(ProfitSnapshot.calculated_at.desc(), ProfitSnapshot.id.desc()),
                )
                .label("rn"),
            )
            .where(ProfitSnapshot.calculation_type == CalculationType.ACTUAL)
            .subquery()
        )
        query = (
            select(
                title_expr.label("title"),
                article_expr.label("seller_article"),
                Order.marketplace,
                Order.sale_model,
                func.count(func.distinct(Order.id)),
                func.coalesce(func.sum(OrderItem.discounted_price * OrderItem.quantity), 0),
                func.coalesce(func.sum(OrderItem.quantity), 0),
                func.coalesce(
                    func.sum(
                        (
                            func.coalesce(OrderItem.cost_price_used, 0)
                            + func.coalesce(OrderItem.package_cost_used, 0)
                        )
                        * OrderItem.quantity
                        + func.coalesce(OrderItem.tax_amount_estimated, 0)
                    ),
                    0,
                ),
                func.coalesce(
                    func.sum(
                        func.coalesce(OrderItem.commission_estimated, 0)
                        + func.coalesce(OrderItem.logistics_estimated, 0)
                        + func.coalesce(OrderItem.other_marketplace_expenses_estimated, 0)
                    ),
                    0,
                ),
                func.coalesce(func.sum(OrderItem.profit_estimated), 0),
                func.coalesce(func.sum(latest_actual.c.profit), 0),
                func.avg(OrderItem.margin_percent_estimated),
                func.avg(OrderItem.commission_estimated),
                func.avg(OrderItem.logistics_estimated),
                func.count(OrderItem.id).filter(OrderItem.cost_price_used.is_(None)),
                func.count(OrderItem.id).filter(
                    OrderItem.economy_confidence == EconomyConfidence.PRELIMINARY.value
                ),
                func.count(latest_actual.c.profit),
            )
            .join(OrderItem, OrderItem.order_id == Order.id)
            .outerjoin(
                latest_actual,
                (latest_actual.c.order_item_id == OrderItem.id) & (latest_actual.c.rn == 1),
            )
            .where(Order.user_id == user_id)
            .where(Order.deleted_at.is_(None))
            .where(Order.order_date >= filters.date_from)
            .where(Order.order_date <= filters.date_to)
            .group_by(
                title_expr,
                article_expr,
                Order.marketplace,
                Order.sale_model,
            )
        )
        return _apply_order_page_filters(query, filters, include_economy=False)

    async def _sales_by_sku(
        self,
        user_id: int,
        filters: OrderWebFilters,
    ) -> dict[tuple[Marketplace, str], int]:
        article_expr = func.coalesce(SalesEvent.seller_article, literal_column("''"))
        query = (
            select(
                SalesEvent.marketplace,
                article_expr.label("seller_article"),
                func.coalesce(func.sum(SalesEvent.quantity), 0),
            )
            .where(SalesEvent.user_id == user_id)
            .where(SalesEvent.event_date >= filters.date_from)
            .where(SalesEvent.event_date <= filters.date_to)
            .group_by(SalesEvent.marketplace, article_expr)
        )
        if filters.marketplace is not None:
            query = query.where(SalesEvent.marketplace == filters.marketplace)
        result = await self.session.execute(query)
        return {
            (marketplace, seller_article): int(quantity or 0)
            for marketplace, seller_article, quantity in result.all()
        }


def build_order_web_filters(
    *,
    timezone: str,
    period: str,
    marketplace: str | None,
    sale_model: str | None,
    date_from: str | None,
    date_to: str | None,
    economy: str,
    status: str,
    sku: str,
    sort: str,
    direction: str,
) -> OrderWebFilters:
    base = build_dashboard_filters(
        timezone=timezone,
        period=period,
        marketplace=marketplace,
        sale_model=sale_model,
        date_from=date_from,
        date_to=date_to,
    )
    return OrderWebFilters(
        period=base.period,
        marketplace=base.marketplace,
        sale_model=base.sale_model,
        local_date_from=base.local_date_from,
        local_date_to=base.local_date_to,
        date_from=base.date_from,
        date_to=base.date_to,
        economy=economy if economy in {"all", "profit", "loss", "missing_cost", "no_finance"} else "all",
        status=(
            status
            if status
            in {
                "all",
                "active",
                "cancelled",
                "new",
                "delivered",
                "return",
                "action_required",
                "fact_missing",
                "fact_partial",
                "fact_complete",
                "match_problem",
            }
            else "all"
        ),
        sku=sku.strip(),
        sort=sort if sort in {
            "date", "profit", "actual_profit", "revenue", "margin", "orders", "sales", "roi",
            "payout", "title",
        } else "date",
        direction="asc" if direction == "asc" else "desc",
    )


def roi_percent(profit: Decimal, cost: Decimal) -> Decimal | None:
    if cost == 0:
        return None
    return (profit / cost * Decimal("100")).quantize(Decimal("0.1"))


def _apply_order_page_filters(
    query: Select[Any],
    filters: OrderWebFilters,
    *,
    include_economy: bool = True,
    count_distinct: bool = False,
) -> Select[Any]:
    if filters.marketplace is not None:
        query = query.where(Order.marketplace == filters.marketplace)
    if filters.sale_model is not None:
        query = query.where(Order.sale_model == filters.sale_model)
    if filters.sku:
        pattern = f"%{filters.sku}%"
        query = query.where(
            (OrderItem.seller_article.ilike(pattern))
            | (OrderItem.marketplace_article.ilike(pattern))
            | (OrderItem.title.ilike(pattern))
        )
    if filters.status == "cancelled":
        query = query.where(
            func.lower(func.coalesce(Order.normalized_status, Order.status, "")).in_(
                ("cancelled", "canceled", "cancel")
            )
        )
    elif filters.status == "active":
        query = query.where(
            ~func.lower(func.coalesce(Order.normalized_status, Order.status, "")).in_(
                ("cancelled", "canceled", "cancel")
            )
        )
    elif filters.status == "new":
        query = query.where(
            func.lower(func.coalesce(Order.normalized_status, Order.status, "")).in_(
                ("new", "new_order", "created", "accepted")
            )
        )
    elif filters.status == "delivered":
        query = query.where(
            func.lower(func.coalesce(Order.normalized_status, Order.status, "")).in_(
                ("delivered", "completed", "sold", "sale")
            )
        )
    elif filters.status == "return":
        query = query.where(
            func.lower(func.coalesce(Order.normalized_status, Order.status, "")).like("%return%")
        )
    elif filters.status == "action_required":
        query = query.where(Order.requires_seller_action.is_(True))
    elif filters.status == "fact_missing":
        query = query.where(~_order_has_wb_fact_row())
    elif filters.status == "fact_partial":
        query = query.where(_order_has_wb_fact_row())
        query = query.where(_order_has_wb_missing_fact_state())
    elif filters.status == "fact_complete":
        query = query.where(_order_has_wb_fact_row())
        query = query.where(~_order_has_wb_missing_fact_state())
    elif filters.status == "match_problem":
        query = query.where(_order_has_wb_match_problem())
    if include_economy:
        if filters.economy == "loss":
            query = query.where(OrderItem.profit_estimated < 0)
        elif filters.economy == "profit":
            query = query.where(OrderItem.profit_estimated >= 0)
        elif filters.economy == "missing_cost":
            query = query.where(OrderItem.cost_price_used.is_(None))
    return query


def _order_has_wb_fact_row() -> Any:
    return exists(
        select(WbDailyReportRow.id).where(
            WbDailyReportRow.linked_order_id == Order.id,
            WbDailyReportRow.operation_scope == "order",
            WbDailyReportRow.deleted_at.is_(None),
            WbDailyReportRow.is_active.is_(True),
        )
    )


def _order_has_wb_missing_fact_state() -> Any:
    return exists(
        select(WbDailyReportRow.id).where(
            WbDailyReportRow.linked_order_id == Order.id,
            WbDailyReportRow.operation_scope == "order",
            WbDailyReportRow.deleted_at.is_(None),
            WbDailyReportRow.is_active.is_(True),
            WbDailyReportRow.order_match_status.in_(("order_pending_match", "pending", "partial")),
        )
    )


def _order_has_wb_match_problem() -> Any:
    return exists(
        select(WbDailyReportRow.id).where(
            WbDailyReportRow.marketplace_account_id == Order.marketplace_account_id,
            WbDailyReportRow.operation_scope == "order",
            WbDailyReportRow.deleted_at.is_(None),
            WbDailyReportRow.is_active.is_(True),
            WbDailyReportRow.order_match_status.in_(
                ("ambiguous", "ambiguous_order_match", "error")
            ),
        )
    )


def _reconciliation_status(
    *,
    marketplace: Marketplace,
    missing_cost: bool,
    has_wb_fact: bool,
    has_wb_missing_fact: bool,
    has_wb_match_problem: bool,
) -> ReconciliationStatus:
    if missing_cost:
        return ReconciliationStatus.MANUAL_REVIEW
    if marketplace != Marketplace.WB:
        return ReconciliationStatus.PRELIMINARY
    if has_wb_match_problem:
        return ReconciliationStatus.FACT_AMBIGUOUS
    if has_wb_fact and has_wb_missing_fact:
        return ReconciliationStatus.FACT_PARTIAL
    if has_wb_fact:
        return ReconciliationStatus.FACT_MATCHED
    return ReconciliationStatus.PRELIMINARY


def _apply_order_sort(query: Select[Any], filters: OrderWebFilters) -> Select[Any]:
    sort_map = {
        "date": Order.order_date,
        "profit": OrderItem.profit_estimated,
        "revenue": OrderItem.discounted_price * OrderItem.quantity,
        "margin": OrderItem.margin_percent_estimated,
    }
    expression = sort_map.get(filters.sort, Order.order_date)
    expression = expression.asc() if filters.direction == "asc" else expression.desc()
    return query.order_by(expression, Order.id.desc())


def _apply_order_sort_query(query: Select[Any], filters: OrderWebFilters) -> Any:
    """Return sort expression for use in .order_by() chaining."""
    sort_map = {
        "date": Order.order_date,
        "profit": OrderItem.profit_estimated,
        "revenue": OrderItem.discounted_price * OrderItem.quantity,
        "margin": OrderItem.margin_percent_estimated,
    }
    expression = sort_map.get(filters.sort, Order.order_date)
    if filters.direction == "asc":
        return expression.asc(), Order.id.desc()
    return expression.desc(), Order.id.desc()


def _latest_snapshot(
    snapshots: list[ProfitSnapshot],
    calculation_type: CalculationType,
) -> ProfitSnapshot | None:
    filtered = [item for item in snapshots if item.calculation_type == calculation_type]
    if not filtered:
        return None
    return max(filtered, key=lambda item: (item.calculated_at, item.id or 0))


def _ozon_category_label(cat: str) -> str:
    labels = {
        "sale": "Продажа",
        "payout": "Выплата",
        "commission": "Комиссия Ozon",
        "logistics": "Логистика",
        "other_marketplace_costs": "Прочие расходы МП",
        "return": "Возвраты",
        "storage": "Хранение",
        "penalty": "Штрафы",
        "deduction": "Удержания",
        "acceptance": "Приёмка",
        "compensation": "Компенсации",
    }
    return labels.get(cat, cat)


def _filter_profit_rows(rows: list[ProfitSkuRow], economy: str) -> list[ProfitSkuRow]:
    if economy == "loss":
        return [row for row in rows if row.estimated_profit < 0]
    if economy == "profit":
        return [row for row in rows if row.estimated_profit >= 0]
    if economy == "missing_cost":
        return [row for row in rows if row.missing_cost_items > 0]
    return rows


def _sort_profit_rows(
    rows: list[ProfitSkuRow],
    sort: str,
    direction: str,
) -> list[ProfitSkuRow]:
    reverse = direction != "asc"
    key_map = {
        "profit": lambda row: row.estimated_profit,
        "actual_profit": lambda row: row.actual_profit or ZERO,
        "revenue": lambda row: row.revenue,
        "margin": lambda row: row.margin_percent,
        "orders": lambda row: Decimal(row.orders),
        "sales": lambda row: Decimal(row.sales),
        "roi": lambda row: row.roi_percent or Decimal("-999999"),
        "title": lambda row: (row.title or "").lower(),
        "payout": lambda row: row.payout,
    }
    key = key_map.get(sort, key_map["profit"])
    return sorted(rows, key=key, reverse=reverse)


def _profit_summary(rows: list[ProfitSkuRow]) -> ProfitSummary:
    estimated = sum((row.estimated_profit for row in rows), ZERO)
    actual = sum((row.actual_profit for row in rows), ZERO)
    revenue = sum((row.revenue for row in rows), ZERO)
    cost = sum((row.cost for row in rows), ZERO)
    payout = sum((row.payout for row in rows), ZERO)
    quantity = sum((row.orders for row in rows), 0)
    sales_count = sum((row.sales for row in rows), 0)
    returns_count = sum((row.returns_count for row in rows), 0)
    deviation = actual - estimated
    return ProfitSummary(
        profit_actual=actual,
        profit_plan=estimated,
        deviation=deviation,
        deviation_percent=_percent(deviation, estimated),
        revenue=revenue,
        payout=payout,
        cost_price=cost,
        avg_margin=(
            (actual / revenue * Decimal("100")).quantize(Decimal("0.1")) if revenue else None
        ),
        roi_percent=roi_percent(actual, cost),
        orders_count=quantity,
        sales_count=sales_count,
        returns_count=returns_count,
    )


def _build_chart_data(rows: list[ProfitSkuRow]) -> ProfitChartData:
    sorted_by_profit = sorted(rows, key=lambda r: r.actual_profit, reverse=True)
    top_n = sorted_by_profit[:10]
    top_labels = [r.title[:20] for r in top_n]
    top_values = [r.actual_profit for r in top_n]
    expense_labels = ["Себестоимость", "Комиссия", "Логистика", "Прочие"]
    expense_values = [
        sum((r.cost for r in rows), ZERO),
        sum((r.avg_commission for r in rows), ZERO),
        sum((r.avg_logistics for r in rows), ZERO),
        sum((r.marketplace_costs - r.avg_commission - r.avg_logistics for r in rows), ZERO),
    ]
    return ProfitChartData(
        top_sku_labels=top_labels,
        top_sku_values=top_values,
        expense_labels=expense_labels,
        expense_values=expense_values,
    )


def _build_profit_tree(summary: ProfitSummary) -> list[ProfitTreeItem]:
    items = [
        ProfitTreeItem(label="Выручка", amount=summary.revenue, tone="neutral"),
    ]
    if summary.payout and summary.payout < summary.revenue:
        discount = summary.revenue - summary.payout
        items.append(ProfitTreeItem(label="Скидки и корректировки", amount=-discount, tone="warn"))
    items.append(ProfitTreeItem(label="К перечислению", amount=summary.payout, tone="neutral"))
    commission = summary.payout * Decimal("0.15") if summary.payout else ZERO
    logistics = summary.payout * Decimal("0.05") if summary.payout else ZERO
    items.append(ProfitTreeItem(label="Комиссия МП", amount=-commission, tone="bad"))
    items.append(ProfitTreeItem(label="Логистика", amount=-logistics, tone="bad"))
    items.append(ProfitTreeItem(label="Себестоимость", amount=-summary.cost_price, tone="bad"))
    profit_tone = "good" if summary.profit_actual >= ZERO else "bad"
    items.append(ProfitTreeItem(label="Фактическая прибыль", amount=summary.profit_actual, tone=profit_tone))
    return items


def _build_attention_items(rows: list[ProfitSkuRow], user_id: int) -> list[ProfitAttentionItem]:
    items = []
    no_cost = [r for r in rows if r.missing_cost_items > 0]
    if no_cost:
        items.append(ProfitAttentionItem(
            title="SKU без себестоимости",
            description=f"{len(no_cost)} товаров без себестоимости. Прибыль не может быть рассчитана корректно.",
            tone="warn",
            count=len(no_cost),
            filter_params={"economy": "missing_cost"},
        ))
    loss = [r for r in rows if r.actual_profit < ZERO]
    if loss:
        items.append(ProfitAttentionItem(
            title="Убыточные SKU",
            description=f"{len(loss)} товаров с отрицательной прибылью. Проверьте себестоимость и комиссии.",
            tone="bad",
            count=len(loss),
            filter_params={"economy": "loss"},
        ))
    no_finance = [r for r in rows if r.reconciliation_status == ReconciliationStatus.PRELIMINARY]
    if no_finance:
        items.append(ProfitAttentionItem(
            title="SKU без финансовых данных",
            description=f"{len(no_finance)} товаров ожидают загрузки финансовых отчётов.",
            tone="warn",
            count=len(no_finance),
            filter_params={"status": "fact_missing"},
        ))
    big_delta = [r for r in rows if abs(r.profit_delta) > Decimal("500")]
    if big_delta:
        items.append(ProfitAttentionItem(
            title="Большое отклонение план/факт",
            description=f"{len(big_delta)} товаров с отклонением более 500 ₽. Требуется анализ.",
            tone="warn",
            count=len(big_delta),
            filter_params={},
        ))
    if not items:
        items.append(ProfitAttentionItem(
            title="Всё в порядке",
            description="Критических проблем по данным за период не обнаружено.",
            tone="good",
            count=0,
            filter_params={},
        ))
    return items


def _percent(value: Decimal, base: Decimal) -> Decimal | None:
    if base == 0:
        return None
    return (value / abs(base) * Decimal("100")).quantize(Decimal("0.1"))


WB_FACT_ARTICLES: tuple[tuple[str, str, str], ...] = (
    ("sale", "Продажа", "retail_amount"),
    ("commission", "Комиссия WB", "commission_rub"),
    ("logistics", "Логистика", "delivery_rub"),
    ("storage", "Хранение", "storage_fee"),
    ("penalty", "Штрафы", "penalty"),
    ("deduction", "Удержания", "deduction"),
    ("acceptance", "Платная приемка FBS", "acceptance"),
    ("compensation", "Компенсации", "reimbursement_amount"),
    ("returns", "Возвраты", "return_count"),
    ("payout", "К перечислению", "for_pay"),
)


def _wb_fact_article_states(
    *,
    linked_rows: list[WbDailyReportRow],
    unlinked_rows: list[WbDailyReportRow],
    has_report_near_order: bool,
) -> list[WbFactArticleState]:
    states: list[WbFactArticleState] = []
    for key, label, attr in WB_FACT_ARTICLES:
        amount = _sum_wb_attr(linked_rows, attr)
        if amount is not None:
            state = "present"
        elif _sum_wb_attr(unlinked_rows, attr) is not None:
            state = "unlinked"
        elif has_report_near_order:
            state = "missing"
        else:
            state = "report_not_loaded"
        states.append(WbFactArticleState(key=key, label=label, amount=amount, state=state))
    return states


def _sum_wb_attr(rows: list[WbDailyReportRow], attr: str) -> Decimal | None:
    values: list[Decimal] = []
    for row in rows:
        if attr == "return_count":
            return_count = getattr(row, "return_count", None)
            if return_count is not None and int(return_count or 0) > 0:
                values.append(_decimal(getattr(row, "retail_amount", None)))
            continue
        value = getattr(row, attr, None)
        if value is not None:
            values.append(_decimal(value))
    if not values:
        return None
    return sum(values, ZERO)


def _decimal(value: object) -> Decimal:
    if isinstance(value, Decimal):
        return value
    if value is None:
        return ZERO
    return Decimal(str(value))


def localized_order_date(value: datetime, timezone: str) -> str:
    return format_datetime_for_user(value, timezone)


def order_state_label(status: str | None, requires_action: bool) -> str:
    return order_status_label(status, requires_action)


def _is_short_srid(srid: str | None) -> bool:
    """Check if srid looks like a short code rather than a real order ID.

    Real WB srids are typically longer strings or contain hyphens.
    Short all-numeric codes are likely financial report references.
    """
    if not srid:
        return False
    text = str(srid).strip()
    if len(text) < 8:
        return True
    if text.isdigit() and len(text) < 12:
        return True
    return False
