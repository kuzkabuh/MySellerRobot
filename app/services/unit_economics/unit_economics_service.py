"""Break-even price analytics service."""

from __future__ import annotations

from dataclasses import dataclass
from decimal import ROUND_HALF_UP, Decimal
from io import BytesIO, StringIO
from typing import Any

from sqlalchemy import String, and_, cast, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.domain import (
    BreakEvenExpenseSetting,
    Order,
    OrderItem,
    Product,
    ProductCostHistory,
)
from app.models.enums import Marketplace
from app.models.ozon_reports import OzonPriceSnapshot
from app.models.products import WbProductPrice

MONEY = Decimal("0.01")
PERCENT = Decimal("100")
RATE = Decimal("0.0001")
ZERO = Decimal("0")
DEFAULT_YELLOW_PROFIT = Decimal("100")
DEFAULT_BLUE_MARGIN = Decimal("35")


@dataclass(slots=True)
class BreakEvenExpenseProfile:
    tax_rate: Decimal = Decimal("0.0600")
    acquiring_rate: Decimal = Decimal("0.0150")
    advertising_rate: Decimal = Decimal("0.0500")
    packaging_cost: Decimal = ZERO
    storage_cost: Decimal = ZERO
    other_cost: Decimal = ZERO
    source: str = "default"


@dataclass(slots=True)
class BreakEvenRow:
    product_id: int | None
    title: str
    seller_article: str
    sku: str
    brand: str
    marketplace: Marketplace
    category: str
    image_url: str | None
    current_price: Decimal
    discounted_price: Decimal
    cost_price: Decimal
    commission_rate: Decimal
    commission_amount: Decimal
    logistics_cost: Decimal
    advertising_cost: Decimal
    acquiring_cost: Decimal
    storage_cost: Decimal
    tax_rate: Decimal
    tax_amount: Decimal
    packaging_cost: Decimal
    other_cost: Decimal
    break_even_price: Decimal
    min_profitable_price: Decimal
    current_margin_percent: Decimal
    current_profit: Decimal
    recommended_price: Decimal
    target_margin_price: Decimal
    simulated_price: Decimal
    simulated_profit: Decimal
    simulated_margin_percent: Decimal
    roi_percent: Decimal
    gross_profit: Decimal
    net_profit: Decimal
    status: str
    status_label: str
    recommendation: str
    data_warning: str | None = None


@dataclass(slots=True)
class BreakEvenSummary:
    total_products: int
    loss_products: int
    risky_products: int
    profitable_products: int
    high_margin_products: int
    average_margin_percent: Decimal
    average_profit: Decimal
    potential_lost_profit: Decimal
    additional_profit_after_optimization: Decimal


@dataclass(slots=True)
class BreakEvenPageResult:
    rows: list[BreakEvenRow]
    summary: BreakEvenSummary
    total_count: int
    filtered_count: int


class UnitEconomicsService:
    """Build break-even and price simulation data from product, price, and order economics."""

    def __init__(self, session: AsyncSession) -> None:
        self.session = session

    async def rows(
        self,
        *,
        user_id: int,
        target_margin_percent: Decimal = Decimal("20"),
        price_delta_percent: Decimal = Decimal("0"),
        limit: int = 80,
    ) -> list[BreakEvenRow]:
        data = await self.table(
            user_id=user_id,
            target_margin_percent=target_margin_percent,
            price_delta_percent=price_delta_percent,
            start=0,
            length=limit,
        )
        return data.rows

    async def table(
        self,
        *,
        user_id: int,
        target_margin_percent: Decimal = Decimal("20"),
        price_delta_percent: Decimal = Decimal("0"),
        search: str = "",
        marketplace: str = "all",
        status: str = "all",
        category: str = "",
        brand: str = "",
        min_profit: Decimal | None = None,
        max_profit: Decimal | None = None,
        min_margin: Decimal | None = None,
        max_margin: Decimal | None = None,
        min_price: Decimal | None = None,
        max_price: Decimal | None = None,
        start: int = 0,
        length: int = 50,
    ) -> BreakEvenPageResult:
        total_count = await self._count_products(user_id=user_id)
        base_filtered_count = await self._count_filtered_products(
            user_id=user_id,
            search=search,
            marketplace=marketplace,
            category=category,
            brand=brand,
        )
        products = await self._fetch_products(
            user_id=user_id,
            search=search,
            marketplace=marketplace,
            category=category,
            brand=brand,
            start=start,
            length=length,
        )
        product_ids = [p.id for p in products]
        cost_by_product = await self._latest_costs(product_ids)
        economics_by_product = await self._order_economics(product_ids)
        wb_prices = await self._latest_wb_prices(products)
        ozon_prices = await self._latest_ozon_prices(product_ids)
        settings = await self._expense_settings(user_id=user_id, products=products)

        rows = [
            self._row_from_product(
                product=product,
                latest_cost=cost_by_product.get(product.id),
                order_economics=economics_by_product.get(product.id, {}),
                current_price=self._current_price(product, wb_prices, ozon_prices),
                expense_profile=self._expense_profile_for(product, settings),
                target_margin_percent=target_margin_percent,
                price_delta_percent=price_delta_percent,
            )
            for product in products
        ]
        rows = self._filter_calculated_rows(
            rows,
            status=status,
            min_profit=min_profit,
            max_profit=max_profit,
            min_margin=min_margin,
            max_margin=max_margin,
            min_price=min_price,
            max_price=max_price,
        )
        summary = self._summary(rows)
        return BreakEvenPageResult(
            rows=rows,
            summary=summary,
            total_count=total_count,
            filtered_count=(
                len(rows)
                if _has_calculated_filters(
                    status=status,
                    min_profit=min_profit,
                    max_profit=max_profit,
                    min_margin=min_margin,
                    max_margin=max_margin,
                    min_price=min_price,
                    max_price=max_price,
                )
                else base_filtered_count
            ),
        )

    async def summary(
        self,
        *,
        user_id: int,
        target_margin_percent: Decimal = Decimal("20"),
    ) -> BreakEvenSummary:
        data = await self.table(
            user_id=user_id,
            target_margin_percent=target_margin_percent,
            start=0,
            length=500,
        )
        return data.summary

    async def detail(
        self,
        *,
        user_id: int,
        product_id: int,
        target_margin_percent: Decimal = Decimal("20"),
    ) -> dict[str, Any]:
        result = await self.session.execute(
            select(Product).where(Product.user_id == user_id, Product.id == product_id)
        )
        product = result.scalar_one_or_none()
        if product is None:
            return {}
        data = await self.table(
            user_id=user_id,
            target_margin_percent=target_margin_percent,
            start=0,
            length=1,
            search=str(product.id),
        )
        row = next((item for item in data.rows if item.product_id == product.id), None)
        if row is None:
            costs = await self._latest_costs([product.id])
            economics = await self._order_economics([product.id])
            wb_prices = await self._latest_wb_prices([product])
            ozon_prices = await self._latest_ozon_prices([product.id])
            settings = await self._expense_settings(user_id=user_id, products=[product])
            row = self._row_from_product(
                product=product,
                latest_cost=costs.get(product.id),
                order_economics=economics.get(product.id, {}),
                current_price=self._current_price(product, wb_prices, ozon_prices),
                expense_profile=self._expense_profile_for(product, settings),
                target_margin_percent=target_margin_percent,
                price_delta_percent=ZERO,
            )
        expense_total = (
            row.commission_amount
            + row.acquiring_cost
            + row.logistics_cost
            + row.storage_cost
            + row.advertising_cost
            + row.tax_amount
            + row.cost_price
            + row.packaging_cost
            + row.other_cost
        )
        return {
            "row": self.row_to_dict(row),
            "expense_structure": [
                self._expense_part("Комиссия MP", row.commission_amount, expense_total),
                self._expense_part("Эквайринг", row.acquiring_cost, expense_total),
                self._expense_part("Логистика", row.logistics_cost, expense_total),
                self._expense_part("Хранение", row.storage_cost, expense_total),
                self._expense_part("Реклама", row.advertising_cost, expense_total),
                self._expense_part("Налоги", row.tax_amount, expense_total),
                self._expense_part("Себестоимость", row.cost_price, expense_total),
                self._expense_part("Упаковка", row.packaging_cost, expense_total),
                self._expense_part("Прочие расходы", row.other_cost, expense_total),
            ],
            "sensitivity": self._sensitivity(row),
        }

    async def save_expense_setting(
        self,
        *,
        user_id: int,
        scope: str,
        category: str | None,
        product_id: int | None,
        tax_rate: Decimal,
        acquiring_rate: Decimal,
        advertising_rate: Decimal,
        packaging_cost: Decimal,
        storage_cost: Decimal,
        other_cost: Decimal,
    ) -> BreakEvenExpenseSetting:
        scope = scope if scope in {"global", "category", "product"} else "global"
        category = category.strip() if category else None
        if scope != "category":
            category = None
        if scope != "product":
            product_id = None
        query = select(BreakEvenExpenseSetting).where(
            BreakEvenExpenseSetting.user_id == user_id,
            BreakEvenExpenseSetting.scope == scope,
            BreakEvenExpenseSetting.category.is_(category)
            if category is None
            else BreakEvenExpenseSetting.category == category,
            BreakEvenExpenseSetting.product_id.is_(product_id)
            if product_id is None
            else BreakEvenExpenseSetting.product_id == product_id,
        )
        existing = (await self.session.execute(query)).scalar_one_or_none()
        row = existing or BreakEvenExpenseSetting(user_id=user_id, scope=scope)
        row.category = category
        row.product_id = product_id
        row.tax_rate = _rate_from_percent(tax_rate)
        row.acquiring_rate = _rate_from_percent(acquiring_rate)
        row.advertising_rate = _rate_from_percent(advertising_rate)
        row.packaging_cost = _money(packaging_cost)
        row.storage_cost = _money(storage_cost)
        row.other_cost = _money(other_cost)
        self.session.add(row)
        await self.session.flush()
        return row

    async def export_csv(self, *, user_id: int) -> str:
        data = await self.table(user_id=user_id, start=0, length=10000)
        output = StringIO()
        headers = [
            "Артикул продавца",
            "SKU",
            "Бренд",
            "Название",
            "Маркетплейс",
            "Категория",
            "Текущая цена",
            "Цена со скидкой",
            "Себестоимость",
            "Комиссия",
            "Логистика",
            "Реклама",
            "Налоги",
            "Прочие расходы",
            "Безубыточная цена",
            "Минимальная прибыльная цена",
            "Маржа %",
            "Прибыль",
            "Рекомендуемая цена",
            "Статус",
        ]
        output.write(";".join(headers) + "\n")
        for row in data.rows:
            values = [
                row.seller_article,
                row.sku,
                row.brand,
                row.title,
                row.marketplace.value,
                row.category,
                row.current_price,
                row.discounted_price,
                row.cost_price,
                row.commission_amount,
                row.logistics_cost,
                row.advertising_cost,
                row.tax_amount,
                row.other_cost,
                row.break_even_price,
                row.min_profitable_price,
                row.current_margin_percent,
                row.current_profit,
                row.recommended_price,
                row.status_label,
            ]
            output.write(";".join(str(value).replace(";", ",") for value in values) + "\n")
        return output.getvalue()

    async def export_xlsx(self, *, user_id: int) -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        data = await self.table(user_id=user_id, start=0, length=10000)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Break-even"
        headers = [
            "Артикул",
            "SKU",
            "Название",
            "Маркетплейс",
            "Цена",
            "Безубыток",
            "Прибыль",
            "Маржа %",
            "Рекомендация",
            "Статус",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="EAF2FF")
        for row in data.rows:
            ws.append(
                [
                    row.seller_article,
                    row.sku,
                    row.title,
                    row.marketplace.value,
                    float(row.current_price),
                    float(row.break_even_price),
                    float(row.current_profit),
                    float(row.current_margin_percent),
                    float(row.recommended_price),
                    row.status_label,
                ]
            )
        stream = BytesIO()
        wb.save(stream)
        return stream.getvalue()

    async def export_pdf(self, *, user_id: int) -> bytes:
        data = await self.table(user_id=user_id, start=0, length=200)
        lines = ["MP Control Break-even report", ""]
        for row in data.rows:
            lines.append(
                f"{row.seller_article[:18]} | {row.marketplace.value} | "
                f"price {row.discounted_price} | BE {row.break_even_price} | "
                f"profit {row.current_profit} | {row.status}"
            )
        return _simple_pdf(lines)

    def calculate_row(
        self,
        *,
        product_id: int | None,
        title: str,
        seller_article: str,
        marketplace: Marketplace,
        current_price: Decimal,
        cost_price: Decimal,
        commission_amount: Decimal,
        logistics_cost: Decimal,
        tax_amount: Decimal,
        target_margin_percent: Decimal,
        price_delta_percent: Decimal,
    ) -> BreakEvenRow:
        profile = BreakEvenExpenseProfile(
            tax_rate=_safe_rate(tax_amount, current_price),
            acquiring_rate=ZERO,
            advertising_rate=ZERO,
        )
        return self._calculate(
            product_id=product_id,
            title=title,
            seller_article=seller_article,
            sku=seller_article,
            brand="",
            marketplace=marketplace,
            category="",
            image_url=None,
            current_price=current_price,
            discounted_price=current_price,
            cost_price=cost_price,
            commission_rate=_safe_rate(commission_amount, current_price),
            logistics_cost=logistics_cost,
            order_other_cost=ZERO,
            profile=profile,
            target_margin_percent=target_margin_percent,
            price_delta_percent=price_delta_percent,
        )

    @staticmethod
    def row_to_dict(row: BreakEvenRow) -> dict[str, Any]:
        return {
            "product_id": row.product_id,
            "image_url": row.image_url,
            "seller_article": row.seller_article,
            "sku": row.sku,
            "brand": row.brand,
            "title": row.title,
            "marketplace": row.marketplace.value,
            "category": row.category,
            "current_price": _json_money(row.current_price),
            "discounted_price": _json_money(row.discounted_price),
            "cost_price": _json_money(row.cost_price),
            "commission_rate": _json_money(row.commission_rate),
            "commission_amount": _json_money(row.commission_amount),
            "logistics_cost": _json_money(row.logistics_cost),
            "advertising_cost": _json_money(row.advertising_cost),
            "acquiring_cost": _json_money(row.acquiring_cost),
            "storage_cost": _json_money(row.storage_cost),
            "tax_amount": _json_money(row.tax_amount),
            "other_cost": _json_money(row.other_cost),
            "break_even_price": _json_money(row.break_even_price),
            "min_profitable_price": _json_money(row.min_profitable_price),
            "current_margin_percent": _json_money(row.current_margin_percent),
            "current_profit": _json_money(row.current_profit),
            "recommended_price": _json_money(row.recommended_price),
            "roi_percent": _json_money(row.roi_percent),
            "gross_profit": _json_money(row.gross_profit),
            "net_profit": _json_money(row.net_profit),
            "status": row.status,
            "status_label": row.status_label,
            "recommendation": row.recommendation,
            "data_warning": row.data_warning,
        }

    async def _count_products(self, *, user_id: int) -> int:
        return int(
            (
                await self.session.execute(
                    select(func.count(Product.id)).where(
                        Product.user_id == user_id,
                        Product.is_active.is_(True),
                    )
                )
            ).scalar_one()
            or 0
        )

    async def _count_filtered_products(
        self,
        *,
        user_id: int,
        search: str,
        marketplace: str,
        category: str,
        brand: str,
    ) -> int:
        query = select(func.count(Product.id)).where(
            Product.user_id == user_id,
            Product.is_active.is_(True),
        )
        query = self._apply_product_filters(
            query,
            search=search,
            marketplace=marketplace,
            category=category,
            brand=brand,
        )
        return int((await self.session.execute(query)).scalar_one() or 0)

    async def _fetch_products(
        self,
        *,
        user_id: int,
        search: str,
        marketplace: str,
        category: str,
        brand: str,
        start: int,
        length: int,
    ) -> list[Product]:
        query = select(Product).where(Product.user_id == user_id, Product.is_active.is_(True))
        query = self._apply_product_filters(
            query,
            search=search,
            marketplace=marketplace,
            category=category,
            brand=brand,
        )
        query = (
            query.order_by(Product.updated_at.desc(), Product.id.desc())
            .offset(start)
            .limit(length)
        )
        return list((await self.session.execute(query)).scalars().all())

    def _apply_product_filters(
        self,
        query: Any,
        *,
        search: str,
        marketplace: str,
        category: str,
        brand: str,
    ) -> Any:
        if marketplace != "all":
            query = query.where(Product.marketplace == Marketplace(marketplace))
        if category:
            query = query.where(Product.category.ilike(f"%{category}%"))
        if brand:
            query = query.where(Product.brand.ilike(f"%{brand}%"))
        if search:
            like = f"%{search}%"
            query = query.where(
                or_(
                    cast(Product.id, String).ilike(like),
                    Product.seller_article.ilike(like),
                    Product.marketplace_article.ilike(like),
                    Product.title.ilike(like),
                    Product.brand.ilike(like),
                    Product.category.ilike(like),
                )
            )
        return query

    async def _latest_costs(self, product_ids: list[int]) -> dict[int, ProductCostHistory]:
        if not product_ids:
            return {}
        rows = (
            await self.session.execute(
                select(ProductCostHistory)
                .where(
                    ProductCostHistory.product_id.in_(product_ids),
                    ProductCostHistory.valid_to.is_(None),
                )
                .order_by(ProductCostHistory.valid_from.desc())
            )
        ).scalars()
        by_product: dict[int, ProductCostHistory] = {}
        for row in rows:
            by_product.setdefault(row.product_id, row)
        return by_product

    async def _order_economics(self, product_ids: list[int]) -> dict[int, dict[str, Decimal]]:
        if not product_ids:
            return {}
        rows = await self.session.execute(
            select(
                OrderItem.product_id,
                func.avg(OrderItem.discounted_price),
                func.avg(func.coalesce(OrderItem.commission_estimated, 0)),
                func.avg(func.coalesce(OrderItem.logistics_estimated, 0)),
                func.avg(func.coalesce(OrderItem.other_marketplace_expenses_estimated, 0)),
                func.avg(func.coalesce(OrderItem.cost_price_used, 0)),
                func.avg(func.coalesce(OrderItem.package_cost_used, 0)),
                func.avg(func.coalesce(OrderItem.tax_amount_estimated, 0)),
            )
            .join(Order, Order.id == OrderItem.order_id)
            .where(OrderItem.product_id.in_(product_ids))
            .group_by(OrderItem.product_id)
        )
        return {
            int(product_id): {
                "avg_price": _money(avg_price),
                "commission": _money(commission),
                "logistics": _money(logistics),
                "other": _money(other),
                "cost": _money(cost),
                "package": _money(package),
                "tax": _money(tax),
            }
            for (
                product_id,
                avg_price,
                commission,
                logistics,
                other,
                cost,
                package,
                tax,
            ) in rows.all()
            if product_id is not None
        }

    async def _latest_wb_prices(self, products: list[Product]) -> dict[int, WbProductPrice]:
        wb_keys = [
            (p.marketplace_account_id, p.external_product_id)
            for p in products
            if p.marketplace == Marketplace.WB and p.external_product_id
        ]
        if not wb_keys:
            return {}
        conditions = [
            and_(
                WbProductPrice.marketplace_account_id == account_id,
                cast(WbProductPrice.wb_nm_id, String) == external_id,
            )
            for account_id, external_id in wb_keys
        ]
        rows = (
            await self.session.execute(
                select(WbProductPrice).where(or_(*conditions)).order_by(WbProductPrice.synced_at.desc())
            )
        ).scalars()
        by_product: dict[int, WbProductPrice] = {}
        product_lookup = {
            (p.marketplace_account_id, p.external_product_id): p.id
            for p in products
            if p.marketplace == Marketplace.WB
        }
        for price in rows:
            product_id = product_lookup.get((price.marketplace_account_id, str(price.wb_nm_id)))
            if product_id:
                by_product.setdefault(product_id, price)
        return by_product

    async def _latest_ozon_prices(self, product_ids: list[int]) -> dict[int, OzonPriceSnapshot]:
        if not product_ids:
            return {}
        rows = (
            await self.session.execute(
                select(OzonPriceSnapshot)
                .where(OzonPriceSnapshot.product_id.in_(product_ids))
                .order_by(OzonPriceSnapshot.synced_at.desc())
            )
        ).scalars()
        by_product: dict[int, OzonPriceSnapshot] = {}
        for row in rows:
            if row.product_id:
                by_product.setdefault(row.product_id, row)
        return by_product

    async def _expense_settings(
        self, *, user_id: int, products: list[Product]
    ) -> dict[str, Any]:
        categories = {p.category for p in products if p.category}
        product_ids = [p.id for p in products]
        query = select(BreakEvenExpenseSetting).where(
            BreakEvenExpenseSetting.user_id == user_id,
            or_(
                BreakEvenExpenseSetting.scope == "global",
                BreakEvenExpenseSetting.category.in_(categories) if categories else False,
                BreakEvenExpenseSetting.product_id.in_(product_ids) if product_ids else False,
            ),
        )
        rows = list((await self.session.execute(query)).scalars().all())
        return {
            "global": next((r for r in rows if r.scope == "global"), None),
            "category": {r.category: r for r in rows if r.scope == "category" and r.category},
            "product": {r.product_id: r for r in rows if r.scope == "product" and r.product_id},
        }

    def _expense_profile_for(
        self, product: Product, settings: dict[str, Any]
    ) -> BreakEvenExpenseProfile:
        row = settings["product"].get(product.id)
        source = "product"
        if row is None and product.category:
            row = settings["category"].get(product.category)
            source = "category"
        if row is None:
            row = settings["global"]
            source = "global" if row else "default"
        if row is None:
            return BreakEvenExpenseProfile()
        return BreakEvenExpenseProfile(
            tax_rate=Decimal(row.tax_rate or 0),
            acquiring_rate=Decimal(row.acquiring_rate or 0),
            advertising_rate=Decimal(row.advertising_rate or 0),
            packaging_cost=_money(row.packaging_cost),
            storage_cost=_money(row.storage_cost),
            other_cost=_money(row.other_cost),
            source=source,
        )

    def _current_price(
        self,
        product: Product,
        wb_prices: dict[int, WbProductPrice],
        ozon_prices: dict[int, OzonPriceSnapshot],
    ) -> tuple[Decimal, Decimal, str | None]:
        if product.marketplace == Marketplace.WB:
            price = wb_prices.get(product.id)
            if price:
                return (
                    _money(price.price or price.discounted_price),
                    _money(price.discounted_price or price.price),
                    None,
                )
        if product.marketplace == Marketplace.OZON:
            price = ozon_prices.get(product.id)
            if price:
                return (
                    _money(price.price or price.marketing_price),
                    _money(price.marketing_price or price.price),
                    None,
                )
        return ZERO, ZERO, "Нет актуальной цены, использована экономика заказов"

    def _row_from_product(
        self,
        *,
        product: Product,
        latest_cost: ProductCostHistory | None,
        order_economics: dict[str, Decimal],
        current_price: tuple[Decimal, Decimal, str | None],
        expense_profile: BreakEvenExpenseProfile,
        target_margin_percent: Decimal,
        price_delta_percent: Decimal,
    ) -> BreakEvenRow:
        price, discounted_price, warning = current_price
        if discounted_price <= 0:
            discounted_price = order_economics.get("avg_price", ZERO)
            price = discounted_price
        cost_price = _money(
            latest_cost.cost_price if latest_cost else order_economics.get("cost", ZERO)
        )
        package_cost = _money(
            latest_cost.package_cost if latest_cost else order_economics.get("package", ZERO)
        )
        if latest_cost and latest_cost.tax_rate:
            expense_profile.tax_rate = Decimal(latest_cost.tax_rate)
        expense_profile.packaging_cost += package_cost
        commission_rate = Decimal(product.marketplace_commission_rate or 0)
        if commission_rate <= 0:
            commission_rate = _safe_rate(order_economics.get("commission", ZERO), discounted_price)
        return self._calculate(
            product_id=product.id,
            title=product.title or "Без названия",
            seller_article=product.seller_article or "н/д",
            sku=product.marketplace_article or product.external_product_id or "",
            brand=product.brand or "",
            marketplace=product.marketplace,
            category=product.category or "",
            image_url=product.image_url,
            current_price=price,
            discounted_price=discounted_price,
            cost_price=cost_price,
            commission_rate=commission_rate,
            logistics_cost=order_economics.get("logistics", ZERO),
            order_other_cost=order_economics.get("other", ZERO),
            profile=expense_profile,
            target_margin_percent=target_margin_percent,
            price_delta_percent=price_delta_percent,
            data_warning=warning,
        )

    def _calculate(
        self,
        *,
        product_id: int | None,
        title: str,
        seller_article: str,
        sku: str,
        brand: str,
        marketplace: Marketplace,
        category: str,
        image_url: str | None,
        current_price: Decimal,
        discounted_price: Decimal,
        cost_price: Decimal,
        commission_rate: Decimal,
        logistics_cost: Decimal,
        order_other_cost: Decimal,
        profile: BreakEvenExpenseProfile,
        target_margin_percent: Decimal,
        price_delta_percent: Decimal,
        data_warning: str | None = None,
    ) -> BreakEvenRow:
        target_margin_rate = target_margin_percent / PERCENT
        variable_rate = (
            commission_rate
            + profile.tax_rate
            + profile.acquiring_rate
            + profile.advertising_rate
        )
        fixed_cost = (
            cost_price
            + _money(logistics_cost)
            + profile.storage_cost
            + profile.packaging_cost
            + profile.other_cost
            + _money(order_other_cost)
        )
        break_even_price = _price_for_margin(fixed_cost, variable_rate, ZERO)
        target_margin_price = _price_for_margin(fixed_cost, variable_rate, target_margin_rate)
        min_profitable_price = _money(break_even_price + Decimal("1"))
        recommended_price = (
            target_margin_price if target_margin_price > current_price else current_price
        )
        simulated_price = _money(current_price * (Decimal("1") + price_delta_percent / PERCENT))
        current_profit = _profit(discounted_price, variable_rate, fixed_cost)
        simulated_profit = _profit(simulated_price, variable_rate, fixed_cost)
        margin = _margin(current_profit, discounted_price)
        simulated_margin = _margin(simulated_profit, simulated_price)
        commission_amount = _money(discounted_price * commission_rate)
        acquiring_cost = _money(discounted_price * profile.acquiring_rate)
        advertising_cost = _money(discounted_price * profile.advertising_rate)
        tax_amount = _money(discounted_price * profile.tax_rate)
        gross_profit = _money(discounted_price - commission_amount - logistics_cost - cost_price)
        roi = _roi(current_profit, cost_price + profile.packaging_cost + profile.other_cost)
        status = _status(discounted_price, break_even_price, current_profit, margin)
        label = _status_label(status)
        return BreakEvenRow(
            product_id=product_id,
            title=title,
            seller_article=seller_article,
            sku=sku,
            brand=brand,
            marketplace=marketplace,
            category=category,
            image_url=image_url,
            current_price=_money(current_price),
            discounted_price=_money(discounted_price),
            cost_price=_money(cost_price),
            commission_rate=(commission_rate * PERCENT).quantize(Decimal("0.1")),
            commission_amount=commission_amount,
            logistics_cost=_money(logistics_cost),
            advertising_cost=advertising_cost,
            acquiring_cost=acquiring_cost,
            storage_cost=profile.storage_cost,
            tax_rate=(profile.tax_rate * PERCENT).quantize(Decimal("0.1")),
            tax_amount=tax_amount,
            packaging_cost=profile.packaging_cost,
            other_cost=_money(profile.other_cost + order_other_cost),
            break_even_price=break_even_price,
            min_profitable_price=min_profitable_price,
            current_margin_percent=margin,
            current_profit=current_profit,
            recommended_price=_money(recommended_price),
            target_margin_price=target_margin_price,
            simulated_price=simulated_price,
            simulated_profit=simulated_profit,
            simulated_margin_percent=simulated_margin,
            roi_percent=roi,
            gross_profit=gross_profit,
            net_profit=current_profit,
            status=status,
            status_label=label,
            recommendation=_recommendation(discounted_price, break_even_price, target_margin_price),
            data_warning=data_warning,
        )

    def _filter_calculated_rows(
        self, rows: list[BreakEvenRow], **filters: Any
    ) -> list[BreakEvenRow]:
        status = filters["status"]
        result = [row for row in rows if status == "all" or row.status == status]
        for key, attr, op in (
            ("min_profit", "current_profit", "ge"),
            ("max_profit", "current_profit", "le"),
            ("min_margin", "current_margin_percent", "ge"),
            ("max_margin", "current_margin_percent", "le"),
            ("min_price", "discounted_price", "ge"),
            ("max_price", "discounted_price", "le"),
        ):
            value = filters[key]
            if value is None:
                continue
            if op == "ge":
                result = [row for row in result if getattr(row, attr) >= value]
            else:
                result = [row for row in result if getattr(row, attr) <= value]
        return result

    def _summary(self, rows: list[BreakEvenRow]) -> BreakEvenSummary:
        total = len(rows)
        if not rows:
            return BreakEvenSummary(0, 0, 0, 0, 0, ZERO, ZERO, ZERO, ZERO)
        loss = sum(1 for row in rows if row.status == "loss")
        risky = sum(1 for row in rows if row.status == "risk")
        high = sum(1 for row in rows if row.status == "high")
        profitable = total - loss - risky
        avg_margin = _money(sum((row.current_margin_percent for row in rows), ZERO) / total)
        avg_profit = _money(sum((row.current_profit for row in rows), ZERO) / total)
        lost = sum((max(ZERO, row.break_even_price - row.discounted_price) for row in rows), ZERO)
        optimized = sum(
            (
                max(
                    ZERO,
                    _profit(row.recommended_price, row.commission_rate / PERCENT, ZERO)
                    - row.current_profit,
                )
                for row in rows
            ),
            ZERO,
        )
        return BreakEvenSummary(
            total_products=total,
            loss_products=loss,
            risky_products=risky,
            profitable_products=profitable,
            high_margin_products=high,
            average_margin_percent=avg_margin,
            average_profit=avg_profit,
            potential_lost_profit=_money(lost),
            additional_profit_after_optimization=_money(optimized),
        )

    def _expense_part(self, label: str, amount: Decimal, total: Decimal) -> dict[str, Any]:
        percent = (amount / total * PERCENT).quantize(Decimal("0.1")) if total > 0 else ZERO
        return {"label": label, "amount": _json_money(amount), "percent": _json_money(percent)}

    def _sensitivity(self, row: BreakEvenRow) -> list[dict[str, Any]]:
        start = max(Decimal("1"), row.break_even_price * Decimal("0.75"))
        end = max(start + Decimal("1"), row.recommended_price * Decimal("1.25"))
        step = (end - start) / Decimal("11")
        variable_rate = (
            row.commission_rate / PERCENT
            + _safe_rate(row.acquiring_cost, row.discounted_price)
            + _safe_rate(row.advertising_cost, row.discounted_price)
            + _safe_rate(row.tax_amount, row.discounted_price)
        )
        fixed_cost = (
            row.cost_price
            + row.logistics_cost
            + row.storage_cost
            + row.packaging_cost
            + row.other_cost
        )
        return [
            {
                "price": _json_money(start + step * Decimal(i)),
                "profit": _json_money(
                    _profit(start + step * Decimal(i), variable_rate, fixed_cost)
                ),
            }
            for i in range(12)
        ]


def _price_for_margin(
    fixed_cost: Decimal,
    variable_rate: Decimal,
    target_margin_rate: Decimal,
) -> Decimal:
    denominator = Decimal("1") - variable_rate - target_margin_rate
    if denominator <= Decimal("0.01"):
        return ZERO
    return _money(fixed_cost / denominator)


def _profit(price: Decimal, variable_rate: Decimal, fixed_cost: Decimal) -> Decimal:
    return _money(price - (price * variable_rate) - fixed_cost)


def _safe_rate(amount: Decimal, price: Decimal) -> Decimal:
    if price <= 0:
        return ZERO
    return Decimal(amount or 0) / price


def _margin(profit: Decimal, price: Decimal) -> Decimal:
    if price <= 0:
        return ZERO
    return (profit / price * PERCENT).quantize(Decimal("0.1"))


def _roi(profit: Decimal, investment: Decimal) -> Decimal:
    if investment <= 0:
        return ZERO
    return (profit / investment * PERCENT).quantize(Decimal("0.1"))


def _status(price: Decimal, break_even: Decimal, profit: Decimal, margin: Decimal) -> str:
    if break_even and price < break_even:
        return "loss"
    if profit < DEFAULT_YELLOW_PROFIT:
        return "risk"
    if margin >= DEFAULT_BLUE_MARGIN:
        return "high"
    return "profit"


def _status_label(status: str) -> str:
    return {
        "loss": "Убыток",
        "risk": "Низкая прибыльность",
        "profit": "Нормальная прибыль",
        "high": "Высокая маржа",
    }.get(status, "Н/д")


def _recommendation(
    current_price: Decimal,
    break_even_price: Decimal,
    target_margin_price: Decimal,
) -> str:
    if break_even_price == 0:
        return "Недостаточно данных для рекомендации"
    if current_price < break_even_price:
        return "Поднять цену минимум до безубыточной"
    if target_margin_price and current_price < target_margin_price:
        return "Цена покрывает расходы, но ниже цели по марже"
    return "Цена соответствует целевой экономике"


def _money(value: Decimal | int | float | None) -> Decimal:
    return Decimal(value or 0).quantize(MONEY, rounding=ROUND_HALF_UP)


def _rate_from_percent(value: Decimal) -> Decimal:
    if value > Decimal("1"):
        value = value / PERCENT
    return Decimal(value or 0).quantize(RATE, rounding=ROUND_HALF_UP)


def _json_money(value: Decimal) -> str:
    return str(_money(value))


def _has_calculated_filters(
    *,
    status: str,
    min_profit: Decimal | None,
    max_profit: Decimal | None,
    min_margin: Decimal | None,
    max_margin: Decimal | None,
    min_price: Decimal | None,
    max_price: Decimal | None,
) -> bool:
    return status != "all" or any(
        value is not None
        for value in (min_profit, max_profit, min_margin, max_margin, min_price, max_price)
    )


def _simple_pdf(lines: list[str]) -> bytes:
    escaped_lines = [
        line.encode("latin-1", "replace")
        .decode("latin-1")
        .replace("\\", "\\\\")
        .replace("(", "\\(")
        .replace(")", "\\)")
        for line in lines[:55]
    ]
    text = (
        "BT /F1 10 Tf 40 790 Td 14 TL "
        + " T* ".join(f"({line}) Tj" for line in escaped_lines)
        + " ET"
    )
    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        (
            b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] "
            b"/Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>"
        ),
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        f"<< /Length {len(text.encode('latin-1'))} >>\nstream\n{text}\nendstream".encode("latin-1"),
    ]
    pdf = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(len(pdf))
        pdf.extend(f"{index} 0 obj\n".encode("ascii"))
        pdf.extend(obj)
        pdf.extend(b"\nendobj\n")
    xref = len(pdf)
    pdf.extend(f"xref\n0 {len(objects) + 1}\n0000000000 65535 f \n".encode("ascii"))
    for offset in offsets[1:]:
        pdf.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    pdf.extend(
        f"trailer << /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF".encode(
            "ascii"
        )
    )
    return bytes(pdf)
