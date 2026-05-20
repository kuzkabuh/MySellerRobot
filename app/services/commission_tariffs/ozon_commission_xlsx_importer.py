"""version: 1.0.0
description: Ozon commission XLSX file import service — validates and imports tariff data.
updated: 2026-05-20
"""

import hashlib
import logging
import re
from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.commission_tariffs import (
    MarketplaceCommissionImportLog,
    MarketplaceCommissionRate,
    MarketplaceCommissionVersion,
)
from app.models.enums import Marketplace

logger = logging.getLogger(__name__)

EXPECTED_SHEET_NAME = "Прайс РФ (БЗ)"
REQUIRED_COLUMNS = {"Категория", "Тип товара"}

FBO_PRICE_RANGES = [
    ("до 100 руб.", 0, 100, False),
    ("свыше 100 до 300 руб.", 100, 300, False),
    ("свыше 300 до 1500 руб.", 300, 1500, False),
    ("свыше 1500 до 5000 руб.", 1500, 5000, False),
    ("свыше 5000 до 10 000 руб.", 5000, 10000, False),
    ("свыше 10 000 руб.", 10000, 999999999, True),
]

FBO_FRESH_PRICE_RANGES = list(FBO_PRICE_RANGES)

FBS_PRICE_RANGES = list(FBO_PRICE_RANGES)

RFBS_PRICE_RANGES = [
    ("до 1500 руб.", 0, 1500, False),
    ("свыше 1500 до 5000 руб.", 1500, 5000, False),
    ("свыше 5000 до 10 000 руб.", 5000, 10000, False),
    ("свыше 10 000 руб.", 10000, 999999999, True),
]

SALES_MODEL_RANGES: dict[str, list[tuple[str, int, int, bool]]] = {
    "fbo": FBO_PRICE_RANGES,
    "fbo_fresh": FBO_FRESH_PRICE_RANGES,
    "fbs": FBS_PRICE_RANGES,
    "rfbs": RFBS_PRICE_RANGES,
}


def _extract_date_from_filename(file_name: str) -> date | None:
    """Try to extract a date from the filename like '06042026'."""
    match = re.search(r"(\d{2})(\d{2})(\d{4})", file_name)
    if match:
        day, month, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
        try:
            return date(year, month, day)
        except ValueError:
            pass
    return None


def _normalize_commission_value(value: Any) -> Decimal | None:
    """Normalize a commission cell value to a percentage (0-100 scale)."""
    if value is None:
        return None
    try:
        dec = Decimal(str(value).strip().replace(",", ".").replace("%", ""))
    except Exception:
        return None
    if dec < 0:
        return None
    if dec <= 1:
        return dec * Decimal("100")
    if dec <= 100:
        return dec
    return None


def _compute_file_sha256(file_bytes: bytes) -> str:
    return hashlib.sha256(file_bytes).hexdigest()


class OzonCommissionXlsxImporter:
    """Import Ozon commission tariffs from an XLSX file."""

    def __init__(self, session: AsyncSession) -> None:
        self.session = session

    async def validate_and_import(
        self,
        file_bytes: bytes,
        file_name: str,
        effective_from: date,
        version_label: str | None = None,
        uploaded_by_user_id: int | None = None,
    ) -> dict[str, Any]:
        """Validate the XLSX file and import commissions.

        Returns a summary dict with import results.
        """
        logger.info(
            "ozon_commission_import_started",
            extra={"file_name": file_name, "effective_from": str(effective_from)},
        )

        file_sha256 = _compute_file_sha256(file_bytes)

        existing = await self._find_import_by_sha256(file_sha256)
        if existing and existing.status == "imported":
            logger.info(
                "ozon_commission_import_duplicate_file",
                extra={"file_sha256": file_sha256},
            )
            return {
                "success": False,
                "duplicate": True,
                "message": "Этот файл уже был импортирован ранее.",
                "existing_import_id": existing.id,
            }

        import_log = MarketplaceCommissionImportLog(
            marketplace=Marketplace.OZON,
            file_name=file_name,
            file_sha256=file_sha256,
            uploaded_by_user_id=uploaded_by_user_id,
            status="uploaded",
        )
        self.session.add(import_log)
        await self.session.flush()

        try:
            import openpyxl
        except ImportError:
            error_msg = "Библиотека openpyxl не установлена."
            return await self._fail_import(import_log, error_msg)

        try:
            wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        except Exception as exc:
            return await self._fail_import(import_log, f"Не удалось открыть XLSX: {exc}")

        sheet = self._find_sheet(wb)
        if sheet is None:
            return await self._fail_import(
                import_log,
                f"Не найден лист '{EXPECTED_SHEET_NAME}'. Доступные: {wb.sheetnames}",
            )

        validation = self._validate_sheet_structure(sheet)
        if not validation["valid"]:
            import_log.status = "validation_failed"
            import_log.validation_errors = validation["errors"]
            import_log.error_message = "Ошибка валидации структуры файла."
            await self.session.commit()
            return {
                "success": False,
                "validation_errors": validation["errors"],
                "message": "Файл не прошёл валидацию: " + "; ".join(validation["errors"]),
            }

        rates_data = self._parse_rates(sheet)
        if not rates_data:
            return await self._fail_import(import_log, "Не найдено ни одной ставки комиссии.")

        import_log.status = "validated"
        import_log.rows_total = len(rates_data)
        await self.session.commit()

        logger.info(
            "ozon_commission_import_validated",
            extra={
                "import_log_id": import_log.id,
                "rows_total": len(rates_data),
            },
        )

        active_version = await self._get_active_ozon_version()
        if active_version:
            active_version.is_active = False
            active_version.effective_to = effective_from
            self.session.add(active_version)

        label = version_label or f"Ozon commissions from {effective_from.isoformat()}"
        new_version = MarketplaceCommissionVersion(
            marketplace=Marketplace.OZON,
            version_label=label,
            effective_from=effective_from,
            effective_to=None,
            source_type="ozon_manual_xlsx",
            source_url=None,
            source_file_name=file_name,
            source_file_sha256=file_sha256,
            imported_by_user_id=uploaded_by_user_id,
            is_active=True,
            imported_at=datetime.now(tz=UTC),
        )
        self.session.add(new_version)
        await self.session.flush()

        rows_imported = 0
        rows_failed = 0
        for rate_data in rates_data:
            try:
                rate = MarketplaceCommissionRate(
                    version_id=new_version.id,
                    marketplace=Marketplace.OZON,
                    **rate_data,
                )
                self.session.add(rate)
                rows_imported += 1
            except Exception:
                rows_failed += 1

        import_log.version_id = new_version.id
        import_log.status = "imported"
        import_log.rows_imported = rows_imported
        import_log.rows_failed = rows_failed
        import_log.finished_at = datetime.now(tz=UTC)
        await self.session.commit()

        logger.info(
            "ozon_commission_import_finished",
            extra={
                "version_id": new_version.id,
                "rows_imported": rows_imported,
                "rows_failed": rows_failed,
            },
        )

        diff_summary = {}
        if active_version:
            diff_summary = await self._compute_diff_summary(active_version.id, new_version.id)

        return {
            "success": True,
            "version_id": new_version.id,
            "version_label": new_version.version_label,
            "effective_from": effective_from,
            "file_name": file_name,
            "file_sha256": file_sha256,
            "rows_total": len(rates_data),
            "rows_imported": rows_imported,
            "rows_failed": rows_failed,
            "diff_summary": diff_summary,
            "message": (
                f"Импортировано {rows_imported} ставок комиссий Ozon. "
                f"Версия: {new_version.version_label}. "
                f"Дата начала: {effective_from.isoformat()}."
            ),
        }

    def _find_sheet(self, wb: Any) -> Any | None:
        if EXPECTED_SHEET_NAME in wb.sheetnames:
            return wb[EXPECTED_SHEET_NAME]
        for name in wb.sheetnames:
            if "прайс" in name.lower() or "price" in name.lower():
                return wb[name]
        return wb.active if wb.sheetnames else None

    def _validate_sheet_structure(self, sheet: Any) -> dict[str, Any]:
        """Validate that the sheet has the expected structure."""
        errors = []

        first_row = [cell.value for cell in sheet[1]]
        second_row = [cell.value for cell in sheet[2]]

        first_row_str = [str(v).strip() if v else "" for v in first_row]
        second_row_str = [str(v).strip() if v else "" for v in second_row]

        has_category = "Категория" in first_row_str or "Категория" in second_row_str
        has_type = "Тип товара" in first_row_str or "Тип товара" in second_row_str

        if not has_category:
            errors.append("Не найдена колонка 'Категория'")
        if not has_type:
            errors.append("Не найдена колонка 'Тип товара'")

        has_fbo = any("fbo" in str(v).lower() for v in first_row if v)
        has_fbs = any("fbs" in str(v).lower() for v in first_row if v)
        has_rfbs = any("rfbs" in str(v).lower() for v in first_row if v)

        if not has_fbo and not has_fbs and not has_rfbs:
            errors.append("Не найдены группы продаж (FBO, FBS, RFBS)")

        data_rows = 0
        for row in sheet.iter_rows(min_row=3, max_col=4, values_only=True):
            if row[0]:
                data_rows += 1
        if data_rows == 0:
            errors.append("Файл не содержит данных о категориях")

        return {"valid": len(errors) == 0, "errors": errors}

    def _parse_rates(self, sheet: Any) -> list[dict[str, Any]]:
        """Parse commission rates from the sheet."""
        rates = []

        first_row = [cell.value for cell in sheet[1]]
        second_row = [cell.value for cell in sheet[2]]

        col_groups = self._identify_column_groups(first_row, second_row)

        for row in sheet.iter_rows(min_row=3, values_only=True):
            category = str(row[0]).strip() if row[0] else None
            product_type = str(row[1]).strip() if row[1] else None
            if not category:
                continue

            for group in col_groups:
                sales_model = group["sales_model"]
                col_ranges = group["ranges"]

                for range_info in col_ranges:
                    col_idx = range_info["col_idx"]
                    if col_idx >= len(row):
                        continue

                    raw_value = row[col_idx]
                    commission = _normalize_commission_value(raw_value)
                    if commission is None:
                        continue

                    rates.append({
                        "category_name": category[:512],
                        "product_type_name": product_type[:512] if product_type else None,
                        "subject_name": None,
                        "object_name": None,
                        "sales_model": sales_model,
                        "price_from": Decimal(str(range_info["price_from"])),
                        "price_to": Decimal(str(range_info["price_to"])),
                        "price_to_inclusive": range_info["inclusive"],
                        "commission_percent": commission,
                        "raw_payload": {
                            "range_label": range_info["label"],
                            "column_index": col_idx,
                        },
                    })

        return rates

    def _identify_column_groups(
        self,
        first_row: list,
        second_row: list,
    ) -> list[dict[str, Any]]:
        """Identify sales model groups and their price range columns."""
        groups = []
        current_group = None

        for idx, header in enumerate(first_row):
            header_str = str(header).strip().lower() if header else ""
            if not header_str:
                if current_group:
                    current_group["end_idx"] = idx - 1
                continue

            if current_group:
                current_group["end_idx"] = idx - 1
                groups.append(current_group)
                current_group = None

            sales_model = None
            if "fbo" in header_str and "fresh" in header_str:
                sales_model = "fbo_fresh"
            elif "fbo" in header_str:
                sales_model = "fbo"
            elif "fbs" in header_str:
                sales_model = "fbs"
            elif "rfbs" in header_str:
                sales_model = "rfbs"

            if sales_model:
                current_group = {
                    "sales_model": sales_model,
                    "start_idx": idx,
                    "end_idx": idx,
                    "ranges": [],
                }

        if current_group:
            current_group["end_idx"] = len(first_row) - 1
            groups.append(current_group)

        for group in groups:
            range_labels = SALES_MODEL_RANGES.get(group["sales_model"], [])
            col_idx = group["start_idx"] + 2
            for label, price_from, price_to, inclusive in range_labels:
                if col_idx <= group["end_idx"]:
                    group["ranges"].append({
                        "col_idx": col_idx,
                        "label": label,
                        "price_from": price_from,
                        "price_to": price_to,
                        "inclusive": inclusive,
                    })
                    col_idx += 1

        return groups

    async def _get_active_ozon_version(self) -> MarketplaceCommissionVersion | None:
        result = await self.session.execute(
            select(MarketplaceCommissionVersion)
            .where(MarketplaceCommissionVersion.marketplace == Marketplace.OZON)
            .where(MarketplaceCommissionVersion.is_active.is_(True))
            .order_by(MarketplaceCommissionVersion.effective_from.desc())
            .limit(1)
        )
        return result.scalar_one_or_none()

    async def _find_import_by_sha256(self, sha256: str) -> MarketplaceCommissionImportLog | None:
        result = await self.session.execute(
            select(MarketplaceCommissionImportLog)
            .where(MarketplaceCommissionImportLog.file_sha256 == sha256)
            .order_by(MarketplaceCommissionImportLog.created_at.desc())
            .limit(1)
        )
        return result.scalar_one_or_none()

    async def _compute_diff_summary(
        self,
        old_version_id: int,
        new_version_id: int,
    ) -> dict[str, int]:
        """Compute a simple diff summary between two versions."""
        from sqlalchemy import select

        old_rates = await self.session.execute(
            select(
                MarketplaceCommissionRate.category_name,
                MarketplaceCommissionRate.product_type_name,
                MarketplaceCommissionRate.sales_model,
                MarketplaceCommissionRate.price_from,
                MarketplaceCommissionRate.price_to,
                MarketplaceCommissionRate.commission_percent,
            ).where(MarketplaceCommissionRate.version_id == old_version_id)
        )
        old_set = set(old_rates.all())

        new_rates = await self.session.execute(
            select(
                MarketplaceCommissionRate.category_name,
                MarketplaceCommissionRate.product_type_name,
                MarketplaceCommissionRate.sales_model,
                MarketplaceCommissionRate.price_from,
                MarketplaceCommissionRate.price_to,
                MarketplaceCommissionRate.commission_percent,
            ).where(MarketplaceCommissionRate.version_id == new_version_id)
        )
        new_set = set(new_rates.all())

        added = len(new_set - old_set)
        removed = len(old_set - new_set)
        changed = 0

        old_by_key = {r[:5]: r[5] for r in old_set}
        new_by_key = {r[:5]: r[5] for r in new_set}
        for key in old_by_key:
            if key in new_by_key and old_by_key[key] != new_by_key[key]:
                changed += 1

        return {"added": added, "removed": removed, "changed": changed}

    async def _fail_import(
        self,
        import_log: MarketplaceCommissionImportLog,
        error_message: str,
    ) -> dict[str, Any]:
        import_log.status = "failed"
        import_log.error_message = error_message
        import_log.finished_at = datetime.now(tz=UTC)
        await self.session.commit()

        logger.error(
            "ozon_commission_import_failed",
            extra={"error": error_message[:300]},
        )

        return {
            "success": False,
            "message": error_message,
        }
