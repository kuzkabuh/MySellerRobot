"""version: 1.0.0
description: Excel template and validation service for product cost imports.
updated: 2026-05-14
"""

from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

REQUIRED_COLUMNS = [
    "Маркетплейс",
    "Кабинет",
    "Артикул продавца",
    "Артикул маркетплейса / offer_id",
    "Себестоимость",
    "Упаковка",
    "Дополнительные расходы",
    "Налог, %",
    "Дата начала действия",
]
TEMPLATE_COLUMNS = [
    "Маркетплейс",
    "Кабинет",
    "Артикул продавца",
    "Артикул маркетплейса / offer_id",
    "Название товара",
    "Себестоимость",
    "Упаковка",
    "Дополнительные расходы",
    "Налог, %",
    "Дата начала действия",
]


@dataclass(slots=True)
class CostImportRow:
    marketplace: str
    account: str
    seller_article: str
    marketplace_article: str
    cost_price: Decimal
    package_cost: Decimal
    additional_cost: Decimal
    tax_rate: Decimal
    valid_from: datetime


@dataclass(slots=True)
class CostTemplateProductRow:
    marketplace: str
    account: str
    seller_article: str | None
    marketplace_article: str | None
    title: str | None


class ExcelCostImportService:
    """Create and parse cost history Excel files."""

    def create_template(self, path: Path) -> Path:
        workbook = Workbook()
        sheet = workbook.active
        if sheet is None:
            raise ValueError("Не удалось создать лист Excel")
        sheet.title = "Себестоимость"
        self._append_header(sheet)
        sheet.append(
            [
                "WB",
                "Основной WB",
                "SKU-001",
                "123456789",
                "Полотенце Fresh",
                520,
                25,
                0,
                6,
                "2026-05-14",
            ]
        )
        workbook.save(path)
        return path

    def create_template_for_products(
        self,
        path: Path,
        products: list[CostTemplateProductRow],
    ) -> Path:
        workbook = Workbook()
        sheet = workbook.active
        if sheet is None:
            raise ValueError("Не удалось создать лист Excel")
        sheet.title = "Себестоимость"
        self._append_header(sheet)
        for product in products:
            sheet.append(
                [
                    product.marketplace,
                    product.account,
                    product.seller_article or "",
                    product.marketplace_article or "",
                    product.title or "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            )
        self._autosize_columns(sheet)
        workbook.save(path)
        return path

    def parse(self, path: Path, max_rows: int = 10_000) -> tuple[list[CostImportRow], list[str]]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        if sheet is None:
            return [], ["Не найден активный лист Excel"]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return [], ["Файл пустой"]
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        missing = [column for column in REQUIRED_COLUMNS if column not in headers]
        if missing:
            return [], [f"Не найдены колонки: {', '.join(missing)}"]
        index = {header: headers.index(header) for header in REQUIRED_COLUMNS}
        parsed: list[CostImportRow] = []
        errors: list[str] = []
        for row_number, row in enumerate(rows[1 : max_rows + 1], start=2):
            if not any(row):
                continue
            try:
                parsed.append(
                    CostImportRow(
                        marketplace=str(row[index["Маркетплейс"]]).strip(),
                        account=str(row[index["Кабинет"]]).strip(),
                        seller_article=str(row[index["Артикул продавца"]]).strip(),
                        marketplace_article=str(
                            row[index["Артикул маркетплейса / offer_id"]]
                        ).strip(),
                        cost_price=self._decimal(row[index["Себестоимость"]]),
                        package_cost=self._decimal(row[index["Упаковка"]]),
                        additional_cost=self._decimal(row[index["Дополнительные расходы"]]),
                        tax_rate=self._decimal(row[index["Налог, %"]]) / Decimal("100"),
                        valid_from=self._date(row[index["Дата начала действия"]]),
                    )
                )
            except (InvalidOperation, ValueError, TypeError) as exc:
                errors.append(f"Строка {row_number}: {exc}")
        return parsed, errors

    @staticmethod
    def _decimal(value: object) -> Decimal:
        return Decimal(str(value or "0")).quantize(Decimal("0.01"))

    @staticmethod
    def _date(value: object) -> datetime:
        if isinstance(value, datetime):
            return value
        return datetime.fromisoformat(str(value))

    @staticmethod
    def _append_header(sheet: Worksheet) -> None:
        sheet.append(TEMPLATE_COLUMNS)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")

    @staticmethod
    def _autosize_columns(sheet: Worksheet) -> None:
        for column in sheet.columns:
            max_length = 0
            letter = get_column_letter(column[0].column or 1)
            for cell in column:
                max_length = max(max_length, len(str(cell.value or "")))
            sheet.column_dimensions[letter].width = min(max_length + 3, 45)
