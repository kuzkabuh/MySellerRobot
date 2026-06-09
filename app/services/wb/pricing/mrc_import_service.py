"""version: 1.1.0
description: MRC bulk import/export service for Excel files.
    Shared between Telegram bot and WEB.
    Preview stored in DB (mrc_imports / mrc_import_rows).
updated: 2026-05-21
"""

import asyncio
import logging
from dataclasses import dataclass
from datetime import UTC, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import get_settings
from app.models.domain import (
    MarketplaceAccount,
    MrcImport,
    MrcImportRow,
    Product,
    WbPromotion,
    WbPromotionNomenclature,
)
from app.models.enums import Marketplace

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

TEMPLATE_COLUMNS = [
    "product_id",
    "wb_nm_id",
    "seller_sku",
    "barcode",
    "brand",
    "product_name",
    "current_wb_price",
    "current_mrc_price",
    "new_mrc_price",
    "min_price",
    "promo_name",
    "promo_plan_price",
    "calculated_price_preview",
    "comment",
]

REQUIRED_COLUMNS = {"product_id", "wb_nm_id", "new_mrc_price"}

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SERVICE_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

PREVIEW_TTL_MINUTES = 30


@dataclass(slots=True)
class MrcImportPreview:
    """Preview result after parsing import file."""

    import_id: int
    user_id: int
    file_name: str
    total_rows: int
    valid_rows: int
    skipped_rows: int
    warning_rows: int
    error_rows: int
    created_at: datetime


@dataclass(slots=True)
class MrcImportResult:
    """Final result after applying import."""

    import_id: int
    user_id: int
    updated_count: int
    cleared_count: int
    skipped_count: int
    error_count: int
    warnings: list[str]
    errors: list[str]


class MrcImportService:
    """Service for bulk MRC import/export via Excel files."""

    def __init__(self, session: AsyncSession) -> None:
        self.session = session
        self.settings = get_settings()

    def _max_file_size(self) -> int:
        return int(self.settings.mrc_import_max_file_size_mb) * 1024 * 1024

    def _max_rows(self) -> int:
        return int(self.settings.mrc_import_max_rows)

    def _allow_clear(self) -> bool:
        return bool(self.settings.mrc_import_allow_clear)

    async def generate_mrc_template(
        self,
        user_id: int,
        account_id: int | None = None,
    ) -> Path:
        """Generate Excel template with user's WB products."""
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl not installed")

        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise RuntimeError("Excel workbook has no active worksheet")
        ws.title = "МРЦ WB"

        service_cols = {0, 1, 2, 3}
        input_col = 8

        for col_idx, header in enumerate(TEMPLATE_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

            if col_idx - 1 in service_cols:
                cell.fill = SERVICE_FILL
            elif col_idx - 1 == input_col:
                cell.fill = INPUT_FILL

        query = (
            select(Product)
            .join(MarketplaceAccount, MarketplaceAccount.id == Product.marketplace_account_id)
            .where(MarketplaceAccount.user_id == user_id)
            .where(Product.marketplace == Marketplace.WB)
            .where(Product.is_active.is_(True))
            .order_by(Product.seller_article)
        )
        if account_id:
            query = query.where(Product.marketplace_account_id == account_id)

        result = await self.session.execute(query)
        products = list(result.scalars().all())

        promo_map = await self._build_promo_map(user_id, products)

        for row_idx, product in enumerate(products, 2):
            wb_nm_id = _extract_nm_id(product)
            ws.cell(row=row_idx, column=1, value=product.id).fill = SERVICE_FILL
            ws.cell(row=row_idx, column=2, value=wb_nm_id).fill = SERVICE_FILL
            ws.cell(row=row_idx, column=3, value=product.seller_article).fill = SERVICE_FILL
            ws.cell(row=row_idx, column=4, value="").fill = SERVICE_FILL
            ws.cell(row=row_idx, column=5, value=product.brand)
            ws.cell(row=row_idx, column=6, value=(product.title or "")[:100])
            ws.cell(row=row_idx, column=7, value="")
            ws.cell(
                row=row_idx, column=8, value=float(product.mrc_price) if product.mrc_price else ""
            )
            ws.cell(row=row_idx, column=9, value="").fill = INPUT_FILL
            ws.cell(row=row_idx, column=10, value="").fill = SERVICE_FILL

            promo_key = (product.marketplace_account_id, wb_nm_id)
            promo_info = promo_map.get(promo_key)
            if promo_info:
                ws.cell(row=row_idx, column=11, value=promo_info["name"])
                ws.cell(
                    row=row_idx,
                    column=12,
                    value=float(promo_info["plan_price"]) if promo_info["plan_price"] else "",
                )

        ws.auto_filter.ref = f"A1:{_col_letter(len(TEMPLATE_COLUMNS))}{len(products) + 1}"
        ws.freeze_panes = "A2"

        widths = [12, 14, 18, 18, 16, 40, 14, 14, 14, 14, 24, 16, 18, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[_col_letter(i)].width = w

        ws_instr = wb.create_sheet("Инструкция")
        instructions = [
            "Инструкция по заполнению МРЦ",
            "",
            "1. Заполните только колонку 'new_mrc_price' (жёлтый фон).",
            "2. МРЦ — это целевая цена товара со скидкой на Wildberries.",
            "3. Цена продавца до скидки будет рассчитана автоматически как МРЦ × 4.",
            "4. Если товар участвует в подходящей акции WB, цена может быть снижена, "
            "но не более чем на 10% от МРЦ.",
            "5. Не меняйте product_id и wb_nm_id — это служебные поля.",
            "6. Если оставить new_mrc_price пустым, МРЦ по товару не изменится.",
            "7. Чтобы очистить МРЦ, укажите CLEAR в колонке new_mrc_price.",
            "8. Значение 0 или отрицательное число вызовет ошибку.",
        ]
        for i, line in enumerate(instructions, 1):
            ws_instr.cell(row=i, column=1, value=line)

        output_dir = Path("/tmp")
        await asyncio.to_thread(output_dir.mkdir, exist_ok=True)
        file_path = output_dir / f"mrc_prices_wb_{datetime.now(tz=UTC).strftime('%Y-%m-%d')}.xlsx"
        wb.save(str(file_path))
        return file_path

    async def create_preview(
        self,
        file_path: Path,
        user_id: int,
        source: str,
        account_id: int | None = None,
        original_file_name: str | None = None,
    ) -> MrcImportPreview:
        """Parse import file, validate rows, and store preview in DB."""
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl not installed")

        logger.info(
            "mrc_import_preview_started",
            extra={
                "user_id": user_id,
                "source": source,
                "file_name": original_file_name or file_path.name,
            },
        )

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError("Excel-файл не содержит листов")

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        headers = [str(h).strip() if h else "" for h in headers]

        missing = REQUIRED_COLUMNS - set(headers)
        if missing:
            raise ValueError(f"Отсутствуют обязательные колонки: {', '.join(sorted(missing))}")

        col_map = {h: i for i, h in enumerate(headers, 1)}

        expires_at = datetime.now(tz=UTC).replace(second=0, microsecond=0)
        expires_at = expires_at + timedelta(minutes=PREVIEW_TTL_MINUTES)

        mrc_import = MrcImport(
            user_id=user_id,
            account_id=account_id,
            source=source,
            original_file_name=original_file_name or file_path.name,
            status="preview",
            expires_at=expires_at,
        )
        self.session.add(mrc_import)
        await self.session.flush()

        seen_product_ids: set[int] = set()
        seen_wb_nm_ids: set[int] = set()
        total_rows = 0
        valid_rows = 0
        skipped_rows = 0
        warning_rows = 0
        error_rows = 0

        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for col_name, col_idx in col_map.items():
                row_data[col_name] = ws.cell(row=row_idx, column=col_idx).value

            product_id = _safe_int(row_data.get("product_id"))
            wb_nm_id = _safe_int(row_data.get("wb_nm_id"))
            new_mrc_raw = _safe_str(row_data.get("new_mrc_price"))

            total_rows += 1

            if total_rows > self._max_rows():
                self.session.add(
                    MrcImportRow(
                        import_id=mrc_import.id,
                        row_number=row_idx,
                        product_id=product_id,
                        wb_nm_id=wb_nm_id,
                        seller_sku=_safe_str(row_data.get("seller_sku")),
                        product_name=_safe_str(row_data.get("product_name")),
                        status="error",
                        message=f"Превышен лимит строк (макс. {self._max_rows()}).",
                    )
                )
                error_rows += 1
                continue

            if not new_mrc_raw or new_mrc_raw.strip() == "":
                self.session.add(
                    MrcImportRow(
                        import_id=mrc_import.id,
                        row_number=row_idx,
                        product_id=product_id,
                        wb_nm_id=wb_nm_id,
                        seller_sku=_safe_str(row_data.get("seller_sku")),
                        product_name=_safe_str(row_data.get("product_name")),
                        old_mrc_price=None,
                        new_mrc_price=None,
                        status="skipped_empty",
                        message="Пустое значение, МРЦ не изменится.",
                    )
                )
                skipped_rows += 1
                continue

            if new_mrc_raw.strip().upper() == "CLEAR":
                if not self._allow_clear():
                    self.session.add(
                        MrcImportRow(
                            import_id=mrc_import.id,
                            row_number=row_idx,
                            product_id=product_id,
                            wb_nm_id=wb_nm_id,
                            status="error",
                            message="Очистка МРЦ запрещена настройкой.",
                        )
                    )
                    error_rows += 1
                    continue
                self.session.add(
                    MrcImportRow(
                        import_id=mrc_import.id,
                        row_number=row_idx,
                        product_id=product_id,
                        wb_nm_id=wb_nm_id,
                        seller_sku=_safe_str(row_data.get("seller_sku")),
                        product_name=_safe_str(row_data.get("product_name")),
                        status="valid_clear",
                    )
                )
                valid_rows += 1
                continue

            try:
                mrc_val = Decimal(new_mrc_raw.replace(",", "."))
            except (InvalidOperation, ValueError):
                self.session.add(
                    MrcImportRow(
                        import_id=mrc_import.id,
                        row_number=row_idx,
                        product_id=product_id,
                        wb_nm_id=wb_nm_id,
                        status="error",
                        message="МРЦ должна быть числом.",
                    )
                )
                error_rows += 1
                continue

            if mrc_val <= 0:
                self.session.add(
                    MrcImportRow(
                        import_id=mrc_import.id,
                        row_number=row_idx,
                        product_id=product_id,
                        wb_nm_id=wb_nm_id,
                        status="error",
                        message="МРЦ должна быть больше 0.",
                    )
                )
                error_rows += 1
                continue

            if not product_id:
                self.session.add(
                    MrcImportRow(
                        import_id=mrc_import.id,
                        row_number=row_idx,
                        product_id=product_id,
                        wb_nm_id=wb_nm_id,
                        status="error",
                        message="product_id должен быть числом.",
                    )
                )
                error_rows += 1
                continue

            if product_id in seen_product_ids:
                self.session.add(
                    MrcImportRow(
                        import_id=mrc_import.id,
                        row_number=row_idx,
                        product_id=product_id,
                        wb_nm_id=wb_nm_id,
                        status="error",
                        message="Дубликат product_id в файле.",
                    )
                )
                error_rows += 1
                continue
            seen_product_ids.add(product_id)

            if wb_nm_id and wb_nm_id in seen_wb_nm_ids:
                self.session.add(
                    MrcImportRow(
                        import_id=mrc_import.id,
                        row_number=row_idx,
                        product_id=product_id,
                        wb_nm_id=wb_nm_id,
                        status="error",
                        message="Дубликат wb_nm_id в файле.",
                    )
                )
                error_rows += 1
                continue
            if wb_nm_id:
                seen_wb_nm_ids.add(wb_nm_id)

            product = await self.session.get(Product, product_id)
            if product is None:
                self.session.add(
                    MrcImportRow(
                        import_id=mrc_import.id,
                        row_number=row_idx,
                        product_id=product_id,
                        wb_nm_id=wb_nm_id,
                        status="error",
                        message="Товар не найден в базе.",
                    )
                )
                error_rows += 1
                continue

            if product.user_id != user_id:
                self.session.add(
                    MrcImportRow(
                        import_id=mrc_import.id,
                        row_number=row_idx,
                        product_id=product_id,
                        wb_nm_id=wb_nm_id,
                        status="error",
                        message="Товар принадлежит другому пользователю.",
                    )
                )
                error_rows += 1
                continue

            if product.marketplace != Marketplace.WB:
                self.session.add(
                    MrcImportRow(
                        import_id=mrc_import.id,
                        row_number=row_idx,
                        product_id=product_id,
                        wb_nm_id=wb_nm_id,
                        status="error",
                        message="Товар не относится к Wildberries.",
                    )
                )
                error_rows += 1
                continue

            product_nm_id = _extract_nm_id(product)
            if wb_nm_id and product_nm_id and wb_nm_id != product_nm_id:
                self.session.add(
                    MrcImportRow(
                        import_id=mrc_import.id,
                        row_number=row_idx,
                        product_id=product_id,
                        wb_nm_id=wb_nm_id,
                        seller_sku=_safe_str(row_data.get("seller_sku")),
                        product_name=_safe_str(row_data.get("product_name")),
                        old_mrc_price=product.mrc_price,
                        new_mrc_price=mrc_val.quantize(Decimal("0.01")),
                        status="warning",
                        message=(
                            f"wb_nm_id в файле ({wb_nm_id}) не совпадает "
                            f"с товаром ({product_nm_id})."
                        ),
                    )
                )
                warning_rows += 1
                continue

            if product.mrc_price and mrc_val == product.mrc_price:
                self.session.add(
                    MrcImportRow(
                        import_id=mrc_import.id,
                        row_number=row_idx,
                        product_id=product_id,
                        wb_nm_id=wb_nm_id,
                        seller_sku=_safe_str(row_data.get("seller_sku")),
                        product_name=_safe_str(row_data.get("product_name")),
                        old_mrc_price=product.mrc_price,
                        new_mrc_price=mrc_val.quantize(Decimal("0.01")),
                        status="skipped_no_change",
                        message="МРЦ не изменилась.",
                    )
                )
                skipped_rows += 1
                continue

            self.session.add(
                MrcImportRow(
                    import_id=mrc_import.id,
                    row_number=row_idx,
                    product_id=product_id,
                    wb_nm_id=wb_nm_id,
                    seller_sku=_safe_str(row_data.get("seller_sku")),
                    product_name=_safe_str(row_data.get("product_name")),
                    old_mrc_price=product.mrc_price,
                    new_mrc_price=mrc_val.quantize(Decimal("0.01")),
                    status="valid",
                )
            )
            valid_rows += 1

        mrc_import.total_rows = total_rows
        mrc_import.valid_rows = valid_rows
        mrc_import.skipped_rows = skipped_rows
        mrc_import.warning_rows = warning_rows
        mrc_import.error_rows = error_rows

        await self.session.flush()
        await self.session.commit()
        await self.session.refresh(mrc_import)

        logger.info(
            "mrc_import_preview_created",
            extra={
                "user_id": user_id,
                "source": source,
                "import_id": mrc_import.id,
                "file_name": original_file_name or file_path.name,
                "total_rows": total_rows,
                "valid_rows": valid_rows,
                "error_rows": error_rows,
            },
        )

        return MrcImportPreview(
            import_id=mrc_import.id,
            user_id=user_id,
            file_name=original_file_name or file_path.name,
            total_rows=total_rows,
            valid_rows=valid_rows,
            skipped_rows=skipped_rows,
            warning_rows=warning_rows,
            error_rows=error_rows,
            created_at=mrc_import.created_at or datetime.now(tz=UTC),
        )

    async def apply_mrc_import(
        self,
        import_id: int,
        user_id: int,
        source: str = "unknown",
    ) -> MrcImportResult:
        """Apply validated MRC import to database."""
        logger.info(
            "mrc_import_apply_started",
            extra={"user_id": user_id, "source": source, "import_id": import_id},
        )

        result = await self.session.execute(select(MrcImport).where(MrcImport.id == import_id))
        mrc_import = result.scalar_one_or_none()

        if mrc_import is None:
            # Diagnostic: check for recent imports by this user to help debug
            recent_result = await self.session.execute(
                select(MrcImport)
                .where(MrcImport.user_id == user_id)
                .where(MrcImport.source == source)
                .order_by(MrcImport.created_at.desc())
                .limit(3)
            )
            recent_imports = recent_result.scalars().all()
            recent_info = [
                {
                    "id": imp.id,
                    "status": imp.status,
                    "created_at": str(imp.created_at),
                    "expires_at": str(imp.expires_at),
                }
                for imp in recent_imports
            ]
            logger.warning(
                "mrc_import_not_found_by_id",
                extra={
                    "import_id": import_id,
                    "user_id": user_id,
                    "source": source,
                    "now_utc": str(datetime.now(tz=UTC)),
                    "recent_imports": recent_info,
                },
            )
            raise ValueError(
                "Предварительная проверка файла устарела или не найдена. Загрузите файл заново."
            )

        if mrc_import.user_id != user_id:
            logger.warning(
                "mrc_import_confirm_rejected",
                extra={
                    "import_id": import_id,
                    "expected_user_id": user_id,
                    "actual_user_id": mrc_import.user_id,
                    "reason": "wrong_user_id",
                },
            )
            raise ValueError("Доступ запрещён.")

        if mrc_import.status == "applied":
            logger.warning(
                "mrc_import_confirm_rejected",
                extra={
                    "import_id": import_id,
                    "status": mrc_import.status,
                    "applied_at": str(mrc_import.applied_at),
                    "reason": "already_applied",
                },
            )
            raise ValueError("Этот файл уже был сохранён ранее.")

        if mrc_import.status in ("cancelled", "failed"):
            logger.warning(
                "mrc_import_confirm_rejected",
                extra={
                    "import_id": import_id,
                    "status": mrc_import.status,
                    "reason": "wrong_status",
                },
            )
            raise ValueError(f"Импорт имеет статус '{mrc_import.status}' и не может быть применён.")

        now_utc = datetime.now(tz=UTC)
        if mrc_import.expires_at and mrc_import.expires_at < now_utc:
            logger.warning(
                "mrc_import_confirm_rejected",
                extra={
                    "import_id": import_id,
                    "expires_at": str(mrc_import.expires_at),
                    "now_utc": str(now_utc),
                    "reason": "expired",
                },
            )
            raise ValueError("Предварительная проверка файла истекла. Загрузите файл заново.")

        if mrc_import.valid_rows == 0:
            raise ValueError(
                "В файле нет строк для сохранения. Заполните колонку "
                "new_mrc_price и загрузите файл заново."
            )

        rows_result = await self.session.execute(
            select(MrcImportRow)
            .where(MrcImportRow.import_id == import_id)
            .where(MrcImportRow.status.in_(("valid", "valid_clear", "warning")))
            .order_by(MrcImportRow.row_number)
        )
        rows = list(rows_result.scalars().all())

        updated_count = 0
        cleared_count = 0
        error_count = 0
        warnings: list[str] = []
        errors: list[str] = []

        for row in rows:
            if not row.product_id:
                error_count += 1
                continue

            product = await self.session.get(Product, row.product_id)
            if product is None or product.user_id != user_id:
                error_count += 1
                continue

            old_mrc = product.mrc_price

            if row.status == "valid_clear":
                product.mrc_price = None
                cleared_count += 1
            else:
                if row.new_mrc_price is None:
                    error_count += 1
                    continue
                product.mrc_price = row.new_mrc_price
                updated_count += 1

                if row.status == "warning" and row.message:
                    warnings.append(f"Строка {row.row_number}: {row.message}")

            logger.info(
                "mrc_price_updated_import",
                extra={
                    "user_id": user_id,
                    "product_id": product.id,
                    "wb_nm_id": _extract_nm_id(product),
                    "old_mrc_price": str(old_mrc),
                    "new_mrc_price": str(product.mrc_price),
                    "source": "import",
                },
            )

        try:
            mrc_import.status = "applied"
            mrc_import.updated_rows = updated_count
            mrc_import.cleared_rows = cleared_count
            mrc_import.applied_at = datetime.now(tz=UTC)
            await self.session.commit()

            logger.info(
                "mrc_import_apply_completed",
                extra={
                    "user_id": user_id,
                    "source": source,
                    "import_id": import_id,
                    "updated_rows": updated_count,
                    "cleared_rows": cleared_count,
                    "error_rows": error_count,
                },
            )
        except Exception as exc:
            await self.session.rollback()
            mrc_import.status = "failed"
            mrc_import.error_text = "Database error during apply"
            await self.session.commit()
            logger.exception(
                "mrc_import_apply_failed",
                extra={"user_id": user_id, "source": source, "import_id": import_id},
            )
            raise ValueError(
                "Не удалось сохранить МРЦ из-за ошибки базы данных. Ошибка записана в лог."
            ) from exc

        return MrcImportResult(
            import_id=import_id,
            user_id=user_id,
            updated_count=updated_count,
            cleared_count=cleared_count,
            skipped_count=mrc_import.skipped_rows,
            error_count=error_count + mrc_import.error_rows,
            warnings=warnings,
            errors=errors,
        )

    async def cancel_import(
        self,
        import_id: int,
        user_id: int,
    ) -> None:
        """Cancel a pending import."""
        result = await self.session.execute(select(MrcImport).where(MrcImport.id == import_id))
        mrc_import = result.scalar_one_or_none()

        if mrc_import is None or mrc_import.user_id != user_id:
            return

        if mrc_import.status == "preview":
            mrc_import.status = "cancelled"
            await self.session.commit()

    async def get_import(self, import_id: int, user_id: int) -> MrcImport | None:
        """Get import record by ID, checking ownership."""
        result = await self.session.execute(
            select(MrcImport).where(MrcImport.id == import_id, MrcImport.user_id == user_id)
        )
        return result.scalar_one_or_none()

    async def get_import_rows(self, import_id: int, user_id: int) -> list[MrcImportRow]:
        """Get rows for an import, checking ownership."""
        import_record = await self.get_import(import_id, user_id)
        if import_record is None:
            return []

        result = await self.session.execute(
            select(MrcImportRow)
            .where(MrcImportRow.import_id == import_id)
            .order_by(MrcImportRow.row_number)
        )
        return list(result.scalars().all())

    async def _build_promo_map(
        self,
        user_id: int,
        products: list[Product],
    ) -> dict[tuple[int, int | None], dict[str, Any]]:
        """Build map of (account_id, wb_nm_id) → promo info for template."""
        now_utc = datetime.now(tz=UTC)

        account_ids = {p.marketplace_account_id for p in products}
        result = await self.session.execute(
            select(WbPromotionNomenclature, WbPromotion.name)
            .join(
                WbPromotion,
                (WbPromotion.wb_promotion_id == WbPromotionNomenclature.wb_promotion_id)
                & (
                    WbPromotion.marketplace_account_id
                    == WbPromotionNomenclature.marketplace_account_id
                ),
            )
            .where(
                WbPromotionNomenclature.marketplace_account_id.in_(account_ids),
                WbPromotion.is_active_today.is_(True),
                WbPromotion.start_datetime <= now_utc,
                WbPromotion.end_datetime >= now_utc,
                WbPromotionNomenclature.plan_price.isnot(None),
                WbPromotionNomenclature.plan_price > 0,
            )
        )

        promo_map: dict[tuple[int, int | None], dict[str, Any]] = {}
        for nom, promo_name in result.all():
            key = (nom.marketplace_account_id, nom.wb_nm_id)
            if key not in promo_map:
                promo_map[key] = {"name": promo_name, "plan_price": nom.plan_price}

        return promo_map

    def generate_import_report(self, rows: list[MrcImportRow]) -> Path:
        """Generate Excel report of import results."""
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl not installed")

        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise RuntimeError("Excel workbook has no active worksheet")
        ws.title = "Результат импорта"

        report_cols = [
            "row_number",
            "product_id",
            "wb_nm_id",
            "seller_sku",
            "product_name",
            "old_mrc_price",
            "new_mrc_price",
            "status",
            "message",
        ]

        for col_idx, header in enumerate(report_cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        for row_idx, row in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=row.row_number)
            ws.cell(row=row_idx, column=2, value=row.product_id)
            ws.cell(row=row_idx, column=3, value=row.wb_nm_id)
            ws.cell(row=row_idx, column=4, value=row.seller_sku)
            ws.cell(row=row_idx, column=5, value=row.product_name)
            ws.cell(
                row=row_idx, column=6, value=float(row.old_mrc_price) if row.old_mrc_price else ""
            )
            ws.cell(
                row=row_idx, column=7, value=float(row.new_mrc_price) if row.new_mrc_price else ""
            )
            ws.cell(row=row_idx, column=8, value=row.status)
            ws.cell(row=row_idx, column=9, value=row.message)

        output_dir = Path("/tmp")
        output_dir.mkdir(exist_ok=True)
        ts = datetime.now(tz=UTC).strftime("%Y-%m-%d_%H-%M")
        file_path = output_dir / f"mrc_import_result_{ts}.xlsx"
        wb.save(str(file_path))
        return file_path


def _extract_nm_id(product: Product) -> int | None:
    """Extract WB nmID from product."""
    if product.marketplace_article and product.marketplace_article.isdigit():
        return int(product.marketplace_article)
    if product.external_product_id and product.external_product_id.isdigit():
        return int(product.external_product_id)
    return None


def _safe_int(value: Any) -> int | None:
    """Safely convert to int."""
    if value is None:
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


def _safe_str(value: Any) -> str | None:
    """Safely convert to string."""
    if value is None:
        return None
    return str(value).strip()


def _col_letter(n: int) -> str:
    """Convert column number to Excel letter (1→A, 2→B, 27→AA)."""
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result
