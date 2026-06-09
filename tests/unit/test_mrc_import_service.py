"""Tests for MRC import service with DB-backed preview storage."""

from datetime import UTC, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from unittest.mock import MagicMock

import pytest

from app.models.domain import MrcImport, MrcImportRow, Product
from app.models.enums import Marketplace
from app.services.wb.pricing.mrc_import_service import HAS_OPENPYXL, MrcImportService

if not HAS_OPENPYXL:
    pytest.skip("openpyxl not installed", allow_module_level=True)


def _create_test_excel(file_path: Path, rows: list[dict]) -> None:
    """Create a test Excel file with given rows."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "МРЦ WB"

    headers = [
        "product_id",
        "wb_nm_id",
        "seller_sku",
        "barcode",
        "brand",
        "product_name",
        "current_wb_price",
        "current_mrc_price",
        "new_mrc_price",
        "min_price",
        "promo_name",
        "promo_plan_price",
        "calculated_price_preview",
        "comment",
    ]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))

    wb.save(str(file_path))


class MockAsyncSession:
    """Mock async session that supports add/flush/commit/execute/get."""

    def __init__(self) -> None:
        self.added_objects: list = []
        self.products: dict[int, MagicMock] = {}
        self.imports: dict[int, MagicMock] = {}
        self.rows: list[MagicMock] = []
        self._next_import_id = 1

    def add(self, obj: object) -> None:
        self.added_objects.append(obj)
        if isinstance(obj, MrcImport):
            obj.id = self._next_import_id
            obj.created_at = datetime.now(tz=UTC)
            self.imports[self._next_import_id] = obj
            self._next_import_id += 1
        elif isinstance(obj, MrcImportRow):
            self.rows.append(obj)

    async def flush(self) -> None:
        pass

    async def commit(self) -> None:
        pass

    async def rollback(self) -> None:
        pass

    async def refresh(self, _obj: object) -> None:
        pass

    async def get(self, model: type, pk: int) -> MagicMock | None:
        return self.products.get(pk)

    async def execute(self, query: object) -> MagicMock:
        result = MagicMock()

        result.scalars = MagicMock()
        result.scalars.return_value = MagicMock()
        result.scalars.return_value.all = MagicMock(return_value=self.rows)

        result.scalar_one_or_none = MagicMock()

        if self.imports:
            first_import = next(iter(self.imports.values()))
            result.scalar_one_or_none.return_value = first_import
        else:
            result.scalar_one_or_none.return_value = None

        return result


@pytest.fixture
def mock_session() -> MockAsyncSession:
    """Create a mock async session."""
    return MockAsyncSession()


@pytest.fixture
def mock_product() -> MagicMock:
    """Create a mock product."""
    product = MagicMock(spec=Product)
    product.id = 1
    product.user_id = 123
    product.marketplace = Marketplace.WB
    product.marketplace_account_id = 1
    product.mrc_price = Decimal("100.00")
    product.seller_article = "TEST-001"
    product.brand = "Test Brand"
    product.title = "Test Product"
    product.external_product_id = "12345"
    product.marketplace_article = None
    product.is_active = True
    return product


@pytest.mark.asyncio
async def test_create_preview_stores_in_db(
    mock_session: MockAsyncSession, mock_product: MagicMock, tmp_path: Path
) -> None:
    """WEB import creates preview in DB."""
    test_file = tmp_path / "test_import.xlsx"
    _create_test_excel(
        test_file,
        [
            {"product_id": 1, "wb_nm_id": 12345, "new_mrc_price": "150"},
        ],
    )

    mock_session.products[1] = mock_product

    service = MrcImportService(mock_session)
    preview = await service.create_preview(
        test_file, user_id=123, source="web", original_file_name="test.xlsx"
    )

    assert preview.import_id is not None
    assert preview.user_id == 123
    assert preview.total_rows == 1
    assert preview.valid_rows >= 0
    assert len(mock_session.added_objects) > 0


@pytest.mark.asyncio
async def test_apply_mrc_import_uses_import_id(
    mock_session: MockAsyncSession, mock_product: MagicMock, tmp_path: Path
) -> None:
    """BOT confirm finds import_id and applies MRC."""
    test_file = tmp_path / "test_import.xlsx"
    _create_test_excel(
        test_file,
        [
            {"product_id": 1, "wb_nm_id": 12345, "new_mrc_price": "150"},
        ],
    )

    mock_session.products[1] = mock_product

    service = MrcImportService(mock_session)
    preview = await service.create_preview(
        test_file, user_id=123, source="bot", original_file_name="test.xlsx"
    )

    assert preview.valid_rows >= 1, f"Expected at least 1 valid row, got {preview.valid_rows}"

    result = await service.apply_mrc_import(preview.import_id, user_id=123, source="bot")

    assert result.import_id == preview.import_id
    assert result.user_id == 123


@pytest.mark.asyncio
async def test_confirm_with_nonexistent_import_id() -> None:
    """BOT confirm with nonexistent import_id shows clear error."""
    session = MockAsyncSession()

    service = MrcImportService(session)
    service.session.imports = {}

    with pytest.raises(ValueError, match="Предварительная проверка файла устарела или не найдена"):
        await service.apply_mrc_import(99999, user_id=123, source="bot")


@pytest.mark.asyncio
async def test_already_applied_import() -> None:
    """Repeat confirm on already applied import shows clear message."""
    session = MockAsyncSession()
    mock_import = MagicMock(spec=MrcImport)
    mock_import.id = 1
    mock_import.user_id = 123
    mock_import.status = "applied"
    mock_import.expires_at = None
    session.imports[1] = mock_import

    service = MrcImportService(session)

    with pytest.raises(ValueError, match="Этот файл уже был сохранён ранее"):
        await service.apply_mrc_import(1, user_id=123, source="bot")


@pytest.mark.asyncio
async def test_import_of_different_user_forbidden() -> None:
    """Import of another user_id is forbidden."""
    session = MockAsyncSession()
    mock_import = MagicMock(spec=MrcImport)
    mock_import.id = 1
    mock_import.user_id = 456
    mock_import.status = "preview"
    mock_import.expires_at = None
    session.imports[1] = mock_import

    service = MrcImportService(session)

    with pytest.raises(ValueError, match="Доступ запрещён"):
        await service.apply_mrc_import(1, user_id=123, source="bot")


@pytest.mark.asyncio
async def test_file_without_required_columns(tmp_path: Path) -> None:
    """File without required columns gives clear error."""
    import openpyxl

    test_file = tmp_path / "bad_import.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="wrong_column")
    wb.save(str(test_file))

    session = MockAsyncSession()
    service = MrcImportService(session)

    with pytest.raises(ValueError, match="Отсутствуют обязательные колонки"):
        await service.create_preview(test_file, user_id=123, source="web")


@pytest.mark.asyncio
async def test_empty_new_mrc_price_skipped(
    mock_session: MockAsyncSession, mock_product: MagicMock, tmp_path: Path
) -> None:
    """Empty new_mrc_price rows are skipped."""
    test_file = tmp_path / "test_import.xlsx"
    _create_test_excel(
        test_file,
        [
            {"product_id": 1, "wb_nm_id": 12345, "new_mrc_price": ""},
        ],
    )

    mock_session.products[1] = mock_product

    service = MrcImportService(mock_session)
    preview = await service.create_preview(test_file, user_id=123, source="web")

    assert preview.total_rows == 1
    assert preview.valid_rows == 0
    assert preview.skipped_rows == 1


@pytest.mark.asyncio
async def test_clear_mrc(
    mock_session: MockAsyncSession, mock_product: MagicMock, tmp_path: Path
) -> None:
    """CLEAR in new_mrc_price clears MRC."""
    test_file = tmp_path / "test_import.xlsx"
    _create_test_excel(
        test_file,
        [
            {"product_id": 1, "wb_nm_id": 12345, "new_mrc_price": "CLEAR"},
        ],
    )

    mock_session.products[1] = mock_product

    service = MrcImportService(mock_session)
    preview = await service.create_preview(test_file, user_id=123, source="web")

    assert preview.valid_rows == 1

    result = await service.apply_mrc_import(preview.import_id, user_id=123, source="web")
    assert result.cleared_count == 1


@pytest.mark.asyncio
async def test_cancel_import() -> None:
    """Cancel import changes status to cancelled."""
    session = MockAsyncSession()
    mock_import = MagicMock(spec=MrcImport)
    mock_import.id = 1
    mock_import.user_id = 123
    mock_import.status = "preview"
    session.imports[1] = mock_import

    service = MrcImportService(session)
    await service.cancel_import(1, user_id=123)

    assert mock_import.status == "cancelled"


@pytest.mark.asyncio
async def test_expired_preview() -> None:
    """Expired preview gives clear error."""
    session = MockAsyncSession()
    mock_import = MagicMock(spec=MrcImport)
    mock_import.id = 1
    mock_import.user_id = 123
    mock_import.status = "preview"
    mock_import.expires_at = datetime.now(tz=UTC) - timedelta(hours=1)
    mock_import.valid_rows = 5
    session.imports[1] = mock_import

    service = MrcImportService(session)

    with pytest.raises(ValueError, match="Предварительная проверка файла истекла"):
        await service.apply_mrc_import(1, user_id=123, source="bot")
