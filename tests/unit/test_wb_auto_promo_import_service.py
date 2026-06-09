"""Tests for WbAutoPromoImportService - Excel/CSV import for auto promotion conditions."""

import csv
import math
import os
from decimal import Decimal
from pathlib import Path
from tempfile import NamedTemporaryFile
from unittest.mock import AsyncMock, MagicMock, patch

import openpyxl
import pytest

from app.services.wb.pricing.wb_auto_promo_import_service import (
    WbAutoPromoImportService,
)


def _create_excel_file(rows: list[list], headers: list[str] | None = None) -> Path:
    """Create a temporary Excel file with given rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Условия автоакций"

    if headers:
        ws.append(headers)
    for row in rows:
        ws.append(row)

    tmp = NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return Path(tmp.name)


def _create_wb_discount_report(rows: list[list]) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчёт по скидкам"
    ws.append(
        [
            "Товар уже участвует в акции",
            "Бренд",
            "Предмет",
            "Наименование",
            "Артикул поставщика",
            "Артикул WB",
            "Последний баркод",
            "",
            "",
            "",
            "",
            "Плановая цена для акции",
            "Текущая розничная цена",
            "Валюта",
            "Текущая скидка на сайте, %",
            "Загружаемая скидка для участия в акции",
            "Статус",
        ]
    )
    for row in rows:
        ws.append(row)
    tmp = NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return Path(tmp.name)


def _remove_file(path: Path) -> None:
    try:
        os.remove(path)
    except FileNotFoundError:
        pass


# Test 1: Import conditions from file row
@pytest.mark.asyncio
async def test_import_conditions_from_file():
    """File row with wb_nm_id=345455998, required_price=846 should be saved."""
    session = AsyncMock()
    session.execute = AsyncMock(
        return_value=MagicMock(scalar_one_or_none=MagicMock(return_value=None))
    )
    session.add = MagicMock()
    session.flush = AsyncMock()

    service = WbAutoPromoImportService(session)

    preview_rows = [
        {
            "row_num": 2,
            "wb_nm_id": 345455998,
            "seller_article": "2461.RoeRue",
            "title": "Test Product",
            "promotion_name": "Модная распродажа",
            "required_price": Decimal("846"),
            "current_wb_price": Decimal("930"),
            "is_participating": False,
            "product_id": None,
            "status": "valid",
            "message": None,
        }
    ]

    saved = await service.apply_import(
        preview_rows,
        user_id=1,
        marketplace_account_id=2,
    )

    assert saved == 1
    assert session.add.called
    condition = session.add.call_args[0][0]
    assert condition.wb_nm_id == 345455998
    assert condition.required_price == Decimal("846")
    assert condition.promotion_name == "Модная распродажа"
    assert condition.source == "file_import"


# Test 2: Preview with valid row finds product
@pytest.mark.asyncio
async def test_preview_valid_row_finds_product():
    """Preview should find product by nmID and mark as valid."""
    session = AsyncMock()

    mock_product = MagicMock()
    mock_product.id = 98
    mock_product.mrc_price = Decimal("930")

    service = WbAutoPromoImportService(session)

    # Patch the _find_product_by_nm_id method directly on the class
    with patch.object(
        WbAutoPromoImportService, "_find_product_by_nm_id", new=AsyncMock(return_value=mock_product)
    ):
        headers = [
            "nmID",
            "Артикул продавца",
            "Название товара",
            "Название автоакции",
            "Цена для участия",
            "Текущая цена WB",
            "Участвует",
        ]
        rows = [[345455998, "2461.RoeRue", "Test Product", "Модная распродажа", 846, 930, "нет"]]

        file_path = _create_excel_file(rows, headers)

        preview, preview_rows = await service.create_preview(
            file_path,
            user_id=1,
            marketplace_account_id=2,
        )

        assert preview.total_rows == 1
        assert preview.valid_rows == 1
        assert preview.warning_rows == 0
        assert preview.error_rows == 0
        assert preview_rows[0]["status"] == "valid"
        assert preview_rows[0]["wb_nm_id"] == 345455998
        assert preview_rows[0]["required_price"] == Decimal("846")

        try:
            _remove_file(file_path)
        except Exception:
            pass


# Test 3: Preview with missing product shows warning
@pytest.mark.asyncio
async def test_preview_missing_product_shows_warning():
    """Preview should show warning when product not found."""
    session = AsyncMock()

    service = WbAutoPromoImportService(session)

    # Patch the _find_product_by_nm_id method to return None
    with patch.object(
        WbAutoPromoImportService, "_find_product_by_nm_id", new=AsyncMock(return_value=None)
    ):
        headers = ["nmID", "Цена для участия"]
        rows = [[999999999, 846]]

        file_path = _create_excel_file(rows, headers)

        preview, preview_rows = await service.create_preview(
            file_path,
            user_id=1,
            marketplace_account_id=2,
        )

        assert preview.total_rows == 1
        assert preview.valid_rows == 0
        assert preview.warning_rows == 1
        assert preview_rows[0]["status"] == "warning"

        try:
            _remove_file(file_path)
        except Exception:
            pass


# Test 4: Preview with invalid nmID shows error
@pytest.mark.asyncio
async def test_preview_invalid_nm_id_shows_error():
    """Preview should show error when nmID is not a number."""
    session = AsyncMock()

    service = WbAutoPromoImportService(session)

    headers = ["nmID", "Цена для участия"]
    rows = [["not_a_number", 846]]

    file_path = _create_excel_file(rows, headers)

    preview, preview_rows = await service.create_preview(
        file_path,
        user_id=1,
        marketplace_account_id=2,
    )

    assert preview.error_rows == 1
    assert preview_rows[0]["status"] == "error"

    try:
        _remove_file(file_path)
    except Exception:
        pass


# Test 5: Preview with missing required_price shows error
@pytest.mark.asyncio
async def test_preview_missing_required_price_shows_error():
    """Preview should show error when required_price is missing."""
    session = AsyncMock()

    service = WbAutoPromoImportService(session)

    headers = ["nmID", "Цена для участия"]
    rows = [[345455998, ""]]

    file_path = _create_excel_file(rows, headers)

    preview, preview_rows = await service.create_preview(
        file_path,
        user_id=1,
        marketplace_account_id=2,
    )

    assert preview.error_rows == 1
    assert preview_rows[0]["status"] == "error"

    try:
        _remove_file(file_path)
    except Exception:
        pass


# Test 6: Parse boolean values correctly
def test_parse_bool_values():
    """_parse_bool should recognize various boolean representations."""
    assert WbAutoPromoImportService._parse_bool("да") is True
    assert WbAutoPromoImportService._parse_bool("участвует") is True
    assert WbAutoPromoImportService._parse_bool("true") is True
    assert WbAutoPromoImportService._parse_bool("1") is True
    assert WbAutoPromoImportService._parse_bool("нет") is False
    assert WbAutoPromoImportService._parse_bool("не участвует") is False
    assert WbAutoPromoImportService._parse_bool("false") is False
    assert WbAutoPromoImportService._parse_bool("0") is False
    assert WbAutoPromoImportService._parse_bool(None) is None
    assert WbAutoPromoImportService._parse_bool("unknown") is None


@pytest.mark.asyncio
async def test_wb_discount_report_uses_plan_price_as_auto_promo_price():
    session = AsyncMock()
    mock_product = MagicMock()
    mock_product.id = 303
    service = WbAutoPromoImportService(session)

    with patch.object(
        WbAutoPromoImportService,
        "_find_product_by_nm_id",
        new=AsyncMock(return_value=mock_product),
    ):
        file_path = _create_wb_discount_report(
            [
                [
                    "нет",
                    "Brand",
                    "Cream",
                    "Крем Wai Ora",
                    "SUP-1",
                    303892412,
                    "barcode",
                    "",
                    "",
                    "",
                    "",
                    446,
                    1820,
                    "RUB",
                    75,
                    76,
                    "Можно участвовать",
                ]
            ]
        )

        preview, rows = await service.create_preview(
            file_path,
            user_id=1,
            marketplace_account_id=2,
        )

        assert preview.valid_rows == 1
        row = rows[0]
        assert row["wb_nm_id"] == 303892412
        assert row["required_price"] == Decimal("446")
        assert row["current_full_price"] == Decimal("1820")
        assert row["current_discount"] == Decimal("75")
        assert row["current_discounted_price"] == Decimal("455.00")
        assert row["wb_upload_discount_percent"] == Decimal("76")
        assert row["fallback_discounted_price"] == Decimal("436.80")
        assert row["condition_type"] == "max_price"
        assert row["raw_payload"]["wb_upload_discount_is_diagnostic"] is True
        assert math.ceil((1 - 446 / 1820) * 100) == 76

        _remove_file(file_path)


@pytest.mark.asyncio
async def test_wb_discount_report_apply_persists_upload_discount_as_diagnostic():
    session = AsyncMock()
    session.execute = AsyncMock(
        return_value=MagicMock(scalar_one_or_none=MagicMock(return_value=None))
    )
    session.add = MagicMock()
    session.flush = AsyncMock()
    service = WbAutoPromoImportService(session)
    preview_rows = [
        {
            "row_num": 2,
            "wb_nm_id": 303892412,
            "seller_article": "SUP-1",
            "title": "Крем Wai Ora",
            "promotion_name": None,
            "required_price": Decimal("446"),
            "current_wb_price": Decimal("455"),
            "current_full_price": Decimal("1820"),
            "current_discount": Decimal("75"),
            "current_discounted_price": Decimal("455"),
            "wb_upload_discount_percent": Decimal("76"),
            "fallback_discounted_price": Decimal("436.80"),
            "condition_type": "max_price",
            "wb_status": "Можно участвовать",
            "is_participating": False,
            "product_id": 303,
            "status": "valid",
            "message": None,
            "raw_payload": {"wb_upload_discount_is_diagnostic": True},
        }
    ]

    saved = await service.apply_import(preview_rows, user_id=1, marketplace_account_id=2)

    assert saved == 1
    condition = session.add.call_args[0][0]
    assert condition.required_price == Decimal("446")
    assert condition.wb_condition_discount_percent == Decimal("76")
    assert condition.current_full_price == Decimal("1820")
    assert condition.current_discount == 75
    assert condition.current_discounted_price == Decimal("455")
    assert condition.candidate_discounted_price == Decimal("446")
    assert condition.condition_type == "max_price"
    assert condition.raw_payload["wb_upload_discount_is_diagnostic"] is True


# Test 7: Template generation
@pytest.mark.asyncio
async def test_template_generation():
    """Template should have correct headers and example row."""
    session = AsyncMock()
    service = WbAutoPromoImportService(session)

    file_path = await service.generate_template(user_id=1)

    wb = None
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter)
        example = next(rows_iter)

        assert "nmID" in headers
        assert "Цена для участия" in headers
        assert example[0] == 345455998  # nmID
        assert example[4] == 980  # required_price
    finally:
        if wb:
            wb.close()
        try:
            _remove_file(file_path)
        except Exception:
            pass


def _create_csv_file(rows: list[list], headers: list[str] | None = None) -> Path:
    """Create a temporary CSV file with given rows."""
    tmp = NamedTemporaryFile(suffix=".csv", delete=False, mode="w", encoding="utf-8", newline="")
    writer = csv.writer(tmp)
    if headers:
        writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    tmp.close()
    return Path(tmp.name)


# Test 8: Import from CSV file
@pytest.mark.asyncio
async def test_import_from_csv_file():
    """CSV file with wb_nm_id=345455998 should be imported without BadZipFile."""
    session = AsyncMock()

    mock_product = MagicMock()
    mock_product.id = 98
    mock_product.mrc_price = Decimal("930")

    service = WbAutoPromoImportService(session)

    with patch.object(
        WbAutoPromoImportService, "_find_product_by_nm_id", new=AsyncMock(return_value=mock_product)
    ):
        headers = [
            "nmID",
            "Артикул продавца",
            "Название товара",
            "Название автоакции",
            "Цена для участия",
            "Текущая цена WB",
            "Участвует",
        ]
        rows = [[345455998, "2461.RoeRue", "Test Product", "Модная распродажа", 846, 930, "нет"]]

        file_path = _create_csv_file(rows, headers)

        preview, preview_rows = await service.create_preview(
            file_path,
            user_id=1,
            marketplace_account_id=2,
        )

        assert preview.total_rows == 1
        assert preview.valid_rows == 1
        assert preview.warning_rows == 0
        assert preview.error_rows == 0
        assert preview_rows[0]["status"] == "valid"
        assert preview_rows[0]["wb_nm_id"] == 345455998
        assert preview_rows[0]["required_price"] == Decimal("846")

        try:
            _remove_file(file_path)
        except Exception:
            pass


# Test 9: Invalid XLSX file returns user-friendly error
@pytest.mark.asyncio
async def test_import_invalid_xlsx_returns_user_friendly_error():
    """Corrupted XLSX file should return user-friendly error, not traceback."""
    session = AsyncMock()
    service = WbAutoPromoImportService(session)

    # Create a file that is NOT a valid XLSX (just plain text)
    tmp = NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.write(b"This is not a valid XLSX file")
    tmp.close()
    file_path = Path(tmp.name)

    with pytest.raises(ValueError) as exc_info:
        await service.create_preview(
            file_path,
            user_id=1,
            marketplace_account_id=2,
        )

    assert "повреждён" in str(exc_info.value) or "XLSX" in str(exc_info.value)

    try:
        _remove_file(file_path)
    except Exception:
        pass


# Test 10: Unsupported file format returns error
@pytest.mark.asyncio
async def test_import_unsupported_format_returns_error():
    """File with unsupported extension should return clear error."""
    session = AsyncMock()
    service = WbAutoPromoImportService(session)

    tmp = NamedTemporaryFile(suffix=".pdf", delete=False)
    tmp.write(b"PDF content")
    tmp.close()
    file_path = Path(tmp.name)

    with pytest.raises(ValueError) as exc_info:
        await service.create_preview(
            file_path,
            user_id=1,
            marketplace_account_id=2,
        )

    assert ".xlsx" in str(exc_info.value) and ".csv" in str(exc_info.value)

    try:
        _remove_file(file_path)
    except Exception:
        pass
